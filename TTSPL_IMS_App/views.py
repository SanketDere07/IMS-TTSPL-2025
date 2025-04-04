from django.shortcuts import render,get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Location,Employee,Company,Category,SubCategory, Box, Rank,Product,ProductImage, Exhibition,Customer,Supplier,StockEntry,StockBarcode,TempStockBarcode
import json
from django.shortcuts import render, redirect
import random
import string
from django.db.models import Count,Sum
from django.core.exceptions import ValidationError
from django.conf import settings
from django.db.models import Count, Sum, Case, When, IntegerField
from datetime import datetime
import barcode
from barcode.writer import ImageWriter
import os
from django.db.models import Max
from django.utils.text import slugify
from PIL import Image
from io import BytesIO
import re
import shutil
from django.core.files.storage import default_storage
from django.db.models import Sum
from django.utils.timezone import now
from reportlab.lib.pagesizes import A3, A4, A5
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageTemplate, Frame
from datetime import datetime
from django.db.models import Min, Max, F, Prefetch

import os
from django.conf import settings
from datetime import datetime
from django.http import HttpResponse
import csv
import openpyxl
from openpyxl.styles import Font
from django.utils.dateparse import parse_date
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

from django.http import JsonResponse
from django.contrib.auth.hashers import make_password
from django.core.validators import EmailValidator, RegexValidator, MinLengthValidator, MaxLengthValidator, ValidationError
from django.db import IntegrityError
from django.contrib import messages
from django.core.mail import send_mail
from .models import User, Profile, Role, UserRole, UserAuditLog,Permission,RolePermissionAuditLog,UserSession,SecondaryEmailConfig, DefaultRolePermission,ScheduledBackup, ScheduledBackupDetails
from .forms import RoleForm, UserForm, ProfileForm, UserRoleForm, UserPermissionForm

from TTSPL_IMS.custom_email_backend import CustomEmailBackend
from django.contrib.auth import get_user_model
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_control
from django.contrib.auth import logout
from django.contrib.sessions.models import Session
from django.utils import timezone
from django.db.models import Q
from django.views.decorators.http import require_GET

from reportlab.lib.pagesizes import letter
from reportlab.platypus import LongTable, KeepTogether
import qrcode
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus.flowables import Image
from django.utils.timezone import localtime
from django.db import connection
import psycopg2
import io
from datetime import timedelta
from django.http import FileResponse, Http404
from django.db.models import Min, Max
from django.core.paginator import Paginator
from django.core.files.storage import default_storage
from django.http import JsonResponse
import uuid
from decimal import Decimal 
from django.core.mail import EmailMultiAlternatives
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Expense, NewExpenseGenerated,VerifyProcessExpense, ExpenseFinanceProcess, AssignProduct, ReturnProductHistory
from django.db.models import Q
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.core.files.base import ContentFile
import traceback 

import time
import random
import uuid

import json

User = get_user_model()


def login_view(request):
    if request.user.is_authenticated:
        # Redirect to dashboard if user is already authenticated
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        if not username:
            return render(request, 'auth/login.html', {'error_message': 'Please enter your email'})
        elif not password:
            return render(request, 'auth/login.html', {'error_message': 'Please enter your password'})
        else:
            user = authenticate(request, email=username, password=password)
            if user is not None:
                # Invalidate previous sessions
                try:
                    user_session = UserSession.objects.get(user=user.id)
                    session = Session.objects.get(
                        session_key=user_session.session_key)
                    session.delete()
                    user_session.delete()
                except UserSession.DoesNotExist:
                    pass
                except Session.DoesNotExist:
                    pass

                # Log in the user and create a new session
                login(request, user)
                UserSession.objects.create(
                    user=user, session_key=request.session.session_key)
                return redirect('/dashboard_page')
            else:
                # Check if username exists
                if User.objects.filter(email=username).exists():
                    return render(request, 'auth/login.html', {'error_message': 'Incorrect password'})
                else:
                    return render(request, 'auth/login.html', {'error_message': 'Invalid email'})
    else:
        return render(request, 'auth/login.html')


def logout_view(request):
    try:
        user_session = UserSession.objects.get(user=request.user.id)
        session = Session.objects.get(session_key=user_session.session_key)
        session.delete()
        user_session.delete()
    except UserSession.DoesNotExist:
        pass
    except Session.DoesNotExist:
        pass
    logout(request)
    return redirect('login')  # Redirect to the login page after logout
    


def send_reset_otp(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        is_resend = request.POST.get('resend') == 'true'
        
        try:
            user = User.objects.get(email=email)
            now = timezone.now()

            # Check resend limits
            if is_resend:
                if not user.can_resend_otp():
                    remaining_time = (user.otp_blocked_until - now).seconds // 60
                    return JsonResponse({
                        'success': False,
                        'message': f'Maximum resend attempts reached. Please try again after {remaining_time} minutes.'
                    }, status=429)

            # Generate and save OTP
            otp = random.randint(100000, 999999)
            user.otp = otp
            user.otp_generated_time = now

            if is_resend:
                user.otp_resend_attempts += 1
                user.last_otp_resend = now
                
                # Block for 30 minutes after 4 attempts
                if user.otp_resend_attempts >= 4:
                    user.otp_blocked_until = now + timedelta(minutes=30)
            
            user.save()
            user.refresh_from_db()

            # Email sending logic (same as before)
            # Prepare email context
            context = {
                'username': user.username,
                'otp': otp,
                'company_name': 'Trisnota Technical Services Pvt. Ltd.',
                'company_phone': '+91 22 4605 5448',
                'company_website': 'https://ttspl.co.in'
            }

            # Render email template
            html_message = render_to_string('auth/email_send_otp_format.html', context)
            
            # Send email
            connection = CustomEmailBackend(use_secondary=False)
            connection.open()
            
            send_mail(
                subject='Password Reset OTP',
                message='',  # Text version can be empty since we're using html_message
                from_email='noreply@ttspl.co.in',
                recipient_list=[email],
                connection=connection,
                html_message=html_message
            )
            
            connection.close()


            if is_resend:
                remaining_attempts = max(0, 4 - user.otp_resend_attempts)
                return JsonResponse({
                    'success': True,
                    'message': f'OTP resent successfully. Remaining attempts: {remaining_attempts}',
                    'clear_timer': True,
                    'remaining_attempts': remaining_attempts
                })

            request.session['email'] = email
            messages.success(request, f"OTP sent to {email}. Please check your email.")
            return redirect('verify_otp')

        except User.DoesNotExist:
            if is_resend:
                return JsonResponse({
                    'success': False,
                    'message': 'Email does not exist!'
                }, status=400)
            messages.error(request, "Email does not exist!")
            return redirect('send_reset_otp')

    return render(request, 'auth/send_reset_otp.html')


def verify_otp(request):
    if request.method == 'POST':
        email = request.session.get('email')
        otp_parts = [request.POST.get(f'otp{i}') for i in range(6)]
        otp = ''.join(otp_parts)

        if not otp.isdigit() or len(otp) != 6:
            messages.error(
                request, "Invalid OTP format. Please enter a 6-digit OTP.")
            return redirect('verify_otp')

        try:
            user = User.objects.get(email=email, otp=otp)

            # Check OTP expiry
            if user.otp_generated_time:
                if timezone.now() - user.otp_generated_time > timedelta(minutes=3):
                    messages.error(
                        request, "OTP has expired. Please request a new one.")
                    user.otp = None  # Clear expired OTP
                    user.otp_generated_time = None  # Clear OTP generation time
                    user.save()
                    return redirect('verify_otp')
            else:
                messages.error(
                    request, "OTP generation time is missing. Please request a new OTP.")
                return redirect('verify_otp')

            # OTP verified
            user.otp = None  # Clear OTP after successful verification
            user.otp_generated_time = None  # Clear OTP generation time
            user.save()
            messages.success(request, "OTP verified successfully!")
            return redirect('reset_password')
        except User.DoesNotExist:
            messages.error(request, "Invalid OTP or OTP expired!")
            return redirect('verify_otp')

    return render(request, 'auth/verify_otp.html')


def reset_password(request):
    if request.method == 'POST':
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        email = request.session.get('email')

        if not password or not confirm_password:
            messages.error(request, "Password fields cannot be empty.")
            return redirect('reset_password')

        if password == confirm_password:
            try:
                user = User.objects.get(email=email)
                user.set_password(password)
                user.save()

               # Prepare email context
                context = {
                    'username': user.username,
                    'company_name': 'Trisnota Technical Services Pvt. Ltd.',
                    'company_phone': '+91 22 4605 5448',
                    'company_website': 'https://ttspl.co.in',
                    'support_email': 'support@ttspl.co.in',
                    'system_name': 'TTSPL IMS'
                }

                # Render email template
                html_message = render_to_string('auth/email_password_reset_confirmation.html', context)
                
                # Send email
                connection = CustomEmailBackend(use_secondary=False)
                connection.open()
                
                send_mail(
                    subject='Password Reset Confirmation',
                    message='Your password has been successfully reset.',  # Plain text fallback
                    from_email='noreply@ttspl.co.in',
                    recipient_list=[email],
                    connection=connection,
                    html_message=html_message
                )
                
                connection.close()
                messages.success(
                    request, "Password has been reset successfully!")
                return redirect('reset_password')
            except User.DoesNotExist:
                messages.error(request, "Something went wrong!")
        else:
            messages.error(request, "Passwords do not match!")

    return render(request, 'auth/reset_password.html')


def resend_otp(request):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        email = data.get('email')
        if not email:
            return JsonResponse({'success': False, 'message': 'Email is required.'})

        try:
            user = User.objects.get(email=email)
            otp = random.randint(100000, 999999)  # Generate a new 6-digit OTP
            user.otp = otp
            user.otp_generated_time = timezone.now()  # Set the current time
            user.save()

            # Send OTP via email using the custom email backend
            # Or use True if you want to use secondary settings
            connection = CustomEmailBackend(use_secondary=False)
            connection.open()

            # Create the HTML message
            subject = 'Password Reset OTP - Resent'
            html_message = f"""
            <!DOCTYPE html>
            <html>
            <head></head>
            <body style="font-family: Arial, sans-serif; line-height: 1.6;">
                <p>
                    Dear {user.username},  <!-- Optional: Add user's first name -->
                </p>

                <p>
                    As requested, we have resent your One-Time Password (OTP) to reset your password. Please use the OTP provided below to complete the process:
                </p>

                <p style="text-align: center; font-size: 1.5em; font-weight: bold;">
                    {otp}
                </p>

                <p>
                    <strong>Note:</strong> This OTP is valid for 3 minutes.
                </p>

                <p>
                    If you did not request this OTP, please ignore this email or contact our support team for assistance.
                </p>

                <p>
                    Best Regards,<br>
                    Trisnota Technical Services Pvt. Ltd.<br>
                    Website: <a href="https://ttspl.co.in">ttspl.co.in</a><br>
                    Contact: +91&nbsp;22&nbsp;4605&nbsp;5448
                </p>
            </body>
            </html>
            """

            from_email = 'noreply@example.com'  # Change to your email address
            recipient_list = [email]

            # Send the email with HTML content
            send_mail(
                subject,
                '',  # Leave the plain text version empty
                from_email,
                recipient_list,
                connection=connection,
                fail_silently=False,
                html_message=html_message  # Use the HTML content for the email
            )
            connection.close()
            return JsonResponse({'success': True})
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Email does not exist!'})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})



def login_page(request):
    return render(request, "auth/login.html")



def reset_password_page(request):
    return render(request, "auth/reset_password.html")


    
def dashboard_page(request):
    return render(request, "dashboard.html")

def location_list_page(request):
    return render(request, "location/list_location.html")

def location_add_page(request):
    return render(request, "location/add_location.html")

def generate_shortcode(name):
    """Generate a two-letter shortcode from the location name."""
    words = re.findall(r'\b\w', name)  # Get the first letter of each word
    if len(words) >= 2:
        shortcode = ''.join(words[:2]).upper()  # Take the first two letters from separate words
    else:
        shortcode = name[:2].upper()  # If single word, take first two letters
    return shortcode

@csrf_exempt
def add_location(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            name = data.get('name')
            details = data.get('details')
            status = data.get('status', 'active')  # Default to 'active' if not provided

            if not name:
                return JsonResponse({'error': 'Location name is required.'}, status=400)

            # Generate the shortcode
            shortcode = generate_shortcode(name)

            # Create a new Location instance with the shortcode
            location = Location.objects.create(name=name, details=details, status=status, shortcode=shortcode)

            return JsonResponse({
                'message': 'Location added successfully!',
                'shortcode': shortcode
            }, status=201)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)



@csrf_exempt
def get_locations(request):
    if request.method == "GET":
        location_name = request.GET.get("location_name", "")
        shortcode = request.GET.get("shortcode", "")
        status = request.GET.get("status", "")

        locations = Location.objects.all()

        if location_name:
            locations = locations.filter(name__icontains=location_name)

        if shortcode:
            locations = locations.filter(shortcode__iexact=shortcode)

        if status:
            locations = locations.filter(status__iexact=status)

        location_list = [
            {
                "id": location.id,
                "name": location.name,
                'shortcode': location.shortcode,
                'details': location.details,
                'status': location.status,
                'created_at': location.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
            for location in locations
        ]

        return JsonResponse({"locations": location_list}, status=200)

    return JsonResponse({"error": "Invalid request method."}, status=405)


def location_view_page(request, location_id):
    # Fetch the specific location by ID
    location = get_object_or_404(Location, id=location_id)

    # Pass the location data to the template
    context = {
        'location': {
            'id': location.id,
            'name': location.name,
            'shortcode': location.shortcode,
            'status': location.status,
            'details': location.details,
            'created_at': location.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        }
    }
    return render(request, "location/view_location.html", context)


@csrf_exempt
def location_update_page(request, id):  
    location = get_object_or_404(Location, id=id)  # Use id instead of location_id
    
    if request.method == "GET":
        # Render the update form with existing data
        return render(request, 'location/update_location.html', {'location': location})
    
    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            new_name = data.get('name', location.name)

            # Check if the name is being updated
            if new_name != location.name:
                location.shortcode = generate_shortcode(new_name)  # Update shortcode
            
            location.name = new_name
            location.details = data.get('details', location.details)
            location.status = data.get('status', location.status)
            location.save()

            return JsonResponse({'message': 'Location updated successfully!', 'shortcode': location.shortcode}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def delete_location(request, id):
    if request.method == 'POST':  # Ensure you're handling a POST request
        try:
            location = get_object_or_404(Location, id=id)
            location.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method'}, status=405)


def employee_list_page(request):
    return render(request, "employee/list_employee.html")

def employee_add_page(request):
    return render(request, "employee/add_employee.html")

def generate_employee_code():
    # Get the highest employee ID in the database
    highest_employee = Employee.objects.all().order_by('-employee_id').first()

    # If there are no employees, start from EMP0001
    if highest_employee is None:
        return "EMP0001"
    
    # Increment the last employee ID by 1
    next_employee_id = highest_employee.employee_id + 1
    
    # Format the employee ID with 'EMP' prefix and zero-padded 4-digit number
    return f"EMP{next_employee_id:04d}"


@csrf_exempt
def add_employee(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            name = data.get('name')
            designation = data.get('designation')
            address = data.get('address')
            location = data.get('location')
            date_of_birth = data.get('date_of_birth')
            mobile_number = data.get('mobile_number')
            email = data.get('email')
            aadhaar_card = data.get('aadhaar_card')
            pan_card = data.get('pan_card', None)
            work_location_id = data.get('work_location')
            employee_code = data.get('employee_code', None)  # Optionally, get employee_code if user inputs it
            
             # Validate work_location ID
            work_location = Location.objects.get(id=work_location_id) if work_location_id else None

            # Generate employee code if not provided
            if not employee_code:
                employee_code = generate_employee_code()

            # Create a new Employee instance
            Employee.objects.create(
                employee_code=employee_code,
                name=name,
                designation=designation,
                address=address,
                location=location,
                work_location=work_location,
                date_of_birth=date_of_birth,
                mobile_number=mobile_number,
                email=email,
                aadhaar_card=aadhaar_card,
                pan_card=pan_card,
                created_at=now(),  
                updated_at=now(),
            )

            return JsonResponse({'message': 'Employee added successfully!'}, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def employee_list(request):
    if request.method == "GET":
        employee_code = request.GET.get("employee_code", "")
        employee_name = request.GET.get("employee_name", "")
        work_location = request.GET.get("work_location", "")  

        employees = Employee.objects.all()

        if employee_code:
            employees = employees.filter(employee_code__iexact=employee_code)

        if employee_name:
            employees = employees.filter(name__icontains=employee_name)

        if work_location:
            employees = employees.filter(work_location__id=work_location)

        employee_list = [
            {
                "id": employee.employee_id,
                "employee_code": employee.employee_code,
                "name": employee.name,
                "designation": employee.designation,
                "mobile_number": employee.mobile_number,
                "email": employee.email,
                "location": employee.location,
                "work_location": employee.work_location.name if employee.work_location else None, 
            }
            for employee in employees
        ]

        return JsonResponse({"employees": employee_list}, status=200)

    return JsonResponse({"error": "Invalid request method."}, status=405)


@csrf_exempt
def get_work_locations(request):
    if request.method == "GET":
        locations = Location.objects.all().values("id", "name")
        return JsonResponse({"work_locations": list(locations)}, status=200)

def employee_view_page(request, employee_id):
    # Fetch the specific employee by ID
    employee = get_object_or_404(Employee, employee_id=employee_id)

    # Pass the employee data to the template
    context = {
        'employee': {
            'id': employee.employee_id,
            'employee_code': employee.employee_code,
            'name': employee.name,
            'designation': employee.designation,
            'address': employee.address,
            'location': employee.location,
            'work_location':employee.work_location,
            'date_of_birth': employee.date_of_birth.strftime('%Y-%m-%d'),
            'mobile_number': employee.mobile_number,
            'email': employee.email,
            'aadhaar_card': employee.aadhaar_card,
            'pan_card': employee.pan_card,
        }
    }
    return render(request, "employee/view_employee.html", context)


def employee_update_page(request, id):
    employee = get_object_or_404(Employee, employee_id=id)

    if request.method == "GET":
        return render(request, 'employee/update_employee.html', {'employee': employee})

    elif request.method == "POST":
        try:
            data = json.loads(request.body)

            # Assign values safely
            employee.name = data.get('name', employee.name)
            employee.designation = data.get('designation', employee.designation)
            employee.address = data.get('address', employee.address)
            employee.location = data.get('location', employee.location)
            employee.date_of_birth = data.get('date_of_birth', employee.date_of_birth)
            employee.mobile_number = data.get('mobile_number', employee.mobile_number)
            employee.email = data.get('email', employee.email)
            employee.aadhaar_card = data.get('aadhaar_card', employee.aadhaar_card)
            employee.pan_card = data.get('pan_card', employee.pan_card)

            work_location_id = data.get('work_location')
            if work_location_id:
                employee.work_location = get_object_or_404(Location, id=int(work_location_id))

            # Update timestamp
            employee.updated_at = now()
            
            employee.save()
            return JsonResponse({'message': 'Employee updated successfully!'}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)


@csrf_exempt
def delete_employee(request, id):
    if request.method == 'POST':
        try:
            employee = get_object_or_404(Employee, employee_id=id)
            employee.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def company_list_page(request):
    return render(request, "company/list_company.html")

def company_add_page(request):
    return render(request, "company/add_company.html")


def add_company(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            company = Company.objects.create(
                company_name=data['company_name'],
                CIN_number=data['cin_number'],
                GST_number=data['gst_number'],
                location=data['location'],
                address=data['address'],
                email=data['email'],
                phone_number=data['phone_number']
            )
            return JsonResponse({'success': True, 'message': 'Company added successfully!'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def company_list(request):
    if request.method == "GET":
        # Get filters from request
        company_name = request.GET.get('company_name', '').strip()
        location = request.GET.get('location', '').strip()

        # Start with all companies
        companies = Company.objects.all()

        # Apply filters
        if company_name:
            companies = companies.filter(company_name__icontains=company_name)
        if location:
            companies = companies.filter(location__icontains=location)

        # Format data for DataTable
        company_list = [
            {
                'id': company.company_id,
                'name': company.company_name,
                'CIN_number': company.CIN_number,
                'GST_number': company.GST_number,
                'location': company.location,
                'email': company.email,
                'phone_number': company.phone_number
            }
            for company in companies
        ]

        return JsonResponse({'companies': company_list}, status=200)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)

def company_view_page(request, company_id):
    # Fetch the specific company by ID
    company = get_object_or_404(Company, company_id=company_id)

    # Pass the company data to the template
    context = {
        'company': {
            'id': company.company_id,
            'name': company.company_name,
            'cin': company.CIN_number,
            'gst': company.GST_number,
            'location': company.location,
            'address': company.address,
            'email': company.email,
            'phone_number': company.phone_number,
        }
    }
    return render(request, "company/view_company.html", context)


def company_update_page(request, id):
    # Fetch company data
    company = get_object_or_404(Company, company_id=id)

    if request.method == "GET":
        # Render the update form with existing data
        return render(request, 'company/update_company.html', {'company': company})

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            # Update company fields
            company.company_name = data.get('company_name', company.company_name)
            company.CIN_number = data.get('CIN_number', company.CIN_number)
            company.GST_number = data.get('GST_number', company.GST_number)
            company.location = data.get('location', company.location)
            company.address = data.get('address', company.address)
            company.email = data.get('email', company.email)
            company.phone_number = data.get('phone_number', company.phone_number)
            company.save()

            return JsonResponse({'message': 'Company updated successfully!'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)


@csrf_exempt
def delete_company(request, id):
    if request.method == 'POST':
        try:
            company = get_object_or_404(Company, company_id=id)
            company.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def category_list_page(request):
    return render(request, "category/list_category.html")

def category_add_page(request):
    return render(request, "category/add_category.html")


def generate_shortcode_category(name):
    """Generate a four-letter shortcode from the category name."""
    shortcode = name[:4].upper()  # Take the first four characters of the name, convert to uppercase
    return shortcode

def add_category(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            category_name = data.get('category_name')
            details = data.get('details')

            if not category_name:
                return JsonResponse({'success': False, 'error': 'Category name is required.'})

            # Generate the shortcode
            shortcode = generate_shortcode_category(category_name)

            # Create a new Category instance and save it
            category = Category.objects.create(
                category_name=category_name,
                details=details,
                shortcode=shortcode  # Save the generated shortcode
            )

            return JsonResponse({'success': True, 'message': 'Category added successfully!'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def category_list(request):
    if request.method == "GET":
        category_name = request.GET.get("category_name", "")
        shortcode = request.GET.get("shortcode", "")

        categories = Category.objects.all()

        if category_name:
            categories = categories.filter(category_name__icontains=category_name)

        if shortcode:
            categories = categories.filter(shortcode__iexact=shortcode)

        category_list = [
            {
                "id": category.category_id,
                "name": category.category_name,
                "shortcode": category.shortcode,
                "description": category.details,
                "created_at": category.created_on.strftime('%Y-%m-%d %H:%M:%S')
            }
            for category in categories
        ]

        return JsonResponse({"categories": category_list}, status=200)

    return JsonResponse({"error": "Invalid request method."}, status=405)


def category_view_page(request, category_id):
    # Fetch the specific category by ID
    category = get_object_or_404(Category, category_id=category_id)

    # Pass the category data to the template
    context = {
        'category': {
            'id': category.category_id,
            'name': category.category_name,
            'shortcode': category.shortcode,
            'description': category.details,
        }
    }
    return render(request, "category/view_category.html", context)


def category_update_page(request, id):
    # Fetch category data
    category = get_object_or_404(Category, category_id=id)

    if request.method == "GET":
        # Render the update form with existing data
        return render(request, 'category/update_category.html', {'category': category})

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            # Update category fields
            new_category_name = data.get('category_name', category.category_name)

            # If the category name has changed, update the shortcode
            if new_category_name != category.category_name:
                category.shortcode = generate_shortcode_category(new_category_name)  # Generate a new shortcode
            
            category.category_name = new_category_name
            category.details = data.get('details', category.details)
            category.save()

            # Send the updated shortcode back in the response
            return JsonResponse({
                'message': 'Category updated successfully!',
                'shortcode': category.shortcode  # Include updated shortcode in response
            }, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)


@csrf_exempt
def delete_category(request, id):
    if request.method == 'POST':
        try:
            category = get_object_or_404(Category, category_id=id)
            category.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def subcategory_list_page(request):
    return render(request, "subcategory/list_subcategory.html")

def subcategory_add_page(request):
    return render(request, "subcategory/add_subcategory.html")

def generate_shortcode_subcategory(name):
    """Generate a four-letter shortcode from the subcategory name."""
    shortcode = name[:4].upper()  # Take the first four characters and convert to uppercase
    return shortcode

@csrf_exempt
def add_subcategory(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            category = Category.objects.get(category_id=data['category_id'])  # Get Category instance
            
            # Generate shortcode from subcategory name
            shortcode = generate_shortcode_subcategory(data['subcategory_name'])

            # Create SubCategory with the generated shortcode
            subcategory = SubCategory.objects.create(
                category=category,
                subcategory_name=data['subcategory_name'],
                details=data['details'],
                shortcode=shortcode  
            )

            return JsonResponse({
                'success': True, 
                'message': 'Subcategory added successfully!',
                'shortcode': shortcode  # Return shortcode in response
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def get_categories(request): 
    categories = list(Category.objects.values('category_id', 'category_name','shortcode'))  # Use 'category_id'
    return JsonResponse({'categories': categories})


@csrf_exempt
def subcategory_list(request):
    if request.method == "GET":
        subcategory_name = request.GET.get("subcategory_name", "").strip()
        shortcode = request.GET.get("shortcode", "").strip()
        category_name = request.GET.get("category_name", "").strip()

        subcategories = SubCategory.objects.all()

        if subcategory_name:
            subcategories = subcategories.filter(subcategory_name__icontains=subcategory_name)

        if shortcode:
            subcategories = subcategories.filter(shortcode__iexact=shortcode)

        if category_name:
            subcategories = subcategories.filter(category__category_name__icontains=category_name)

        subcategory_list = [
            {
                "id": subcategory.subcategory_id,
                "name": subcategory.subcategory_name,
                "shortcode": subcategory.shortcode,
                "details": subcategory.details,
                "category_name": subcategory.category.category_name,
                "created_on": subcategory.created_on.strftime('%Y-%m-%d %H:%M:%S'),
            }
            for subcategory in subcategories
        ]

        return JsonResponse({"subcategories": subcategory_list}, status=200)

    return JsonResponse({"error": "Invalid request method."}, status=405)


def subcategory_view_page(request, subcategory_id):
    # Fetch the specific subcategory by ID
    subcategory = get_object_or_404(SubCategory, subcategory_id=subcategory_id)

    # Pass the subcategory data to the template
    context = {
        'subcategory': {
            'id': subcategory.subcategory_id,
            'name': subcategory.subcategory_name,
            'shortcode': subcategory.shortcode,
            'details': subcategory.details,
            'category_name': subcategory.category.category_name,
        }
    }
    return render(request, "subcategory/view_subcategory.html", context)


def subcategory_update_page(request, id):
    # Fetch subcategory data
    subcategory = get_object_or_404(SubCategory, subcategory_id=id)

    if request.method == "GET":
        return render(request, 'subcategory/update_subcategory.html', {'subcategory': subcategory})

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            
            # Check if subcategory name has changed
            new_name = data.get('subcategory_name', subcategory.subcategory_name)
            if new_name != subcategory.subcategory_name:
                subcategory.shortcode = generate_shortcode_subcategory(new_name)  # Update shortcode
            
            # Update subcategory fields
            subcategory.subcategory_name = new_name
            subcategory.category_id = data.get('category_id', subcategory.category_id)
            subcategory.details = data.get('details', subcategory.details)
            subcategory.save()

            return JsonResponse({
                'message': 'Subcategory updated successfully!',
                'shortcode': subcategory.shortcode  # Return updated shortcode
            }, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)


@csrf_exempt
def delete_subcategory(request, id):
    if request.method == 'POST':
        try:
            subcategory = get_object_or_404(SubCategory, subcategory_id=id)
            subcategory.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method.'}, status=405)



def box_list_page(request):
    return render(request, "box/list_box.html")

def box_add_page(request):
    return render(request, "box/add_box.html")


def add_box(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            box = Box.objects.create(
                box_name=data['box_name'],
                details=data['details']
            )
            return JsonResponse({'success': True, 'message': 'Box added successfully!'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def box_list(request):
    if request.method == "GET":
        box_id = request.GET.get('box_id', '')
        box_name = request.GET.get('box_name', '')

        # Filtering logic
        boxes = Box.objects.all()
        if box_id:
            boxes = boxes.filter(box_id__icontains=box_id)
        if box_name:
            boxes = boxes.filter(box_name__icontains=box_name)

        box_list = [
            {
                'id': box.box_id,
                'name': box.box_name,
                'description': box.details,
            }
            for box in boxes
        ]
        return JsonResponse({'boxes': box_list}, status=200)
    
    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def box_view_page(request, box_id):
    # Fetch the specific box by ID
    box = get_object_or_404(Box, box_id=box_id)

    # Pass the box data to the template
    context = {
        'box': {
            'id': box.box_id,
            'name': box.box_name,
            'details': box.details,
            'created_on': box.created_on,
        }
    }
    return render(request, "box/view_box.html", context)


def box_update_page(request, id):
    # Fetch box data by ID
    box = get_object_or_404(Box, box_id=id)

    if request.method == "GET":
        # Render the update form with existing box data
        return render(request, 'box/update_box.html', {'box': box})

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            # Update box fields with the new data
            box.box_name = data.get('box_name', box.box_name)
            box.details = data.get('details', box.details)
            box.save()

            return JsonResponse({'message': 'Box updated successfully!'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)


@csrf_exempt
def delete_box(request, id):
    if request.method == 'POST':
        try:
            box = get_object_or_404(Box, box_id=id)
            box.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def rank_list_page(request):
    return render(request, "rank/list_rank.html")

def rank_add_page(request):
    return render(request, "rank/add_rank.html")


@csrf_exempt
def add_rank(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            rank = Rank.objects.create(
                rank_name=data['rank_name'],
                details=data['details']
            )
            return JsonResponse({'success': True, 'message': 'Rack added successfully!'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def rank_list(request):
    if request.method == "GET":
        rack_id = request.GET.get('rack_id', '')
        rack_name = request.GET.get('rack_name', '')

        # Filtering logic
        ranks = Rank.objects.all()
        if rack_id:
            ranks = ranks.filter(rank_id__icontains=rack_id)
        if rack_name:
            ranks = ranks.filter(rank_name__icontains=rack_name)

        rank_list = [
            {
                'id': rank.rank_id,
                'name': rank.rank_name,
                'description': rank.details,  
            }
            for rank in ranks
        ]
        return JsonResponse({'ranks': rank_list}, status=200)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def rank_view_page(request, rank_id):
    # Fetch the specific rank by ID
    rank = get_object_or_404(Rank, rank_id=rank_id)

    # Pass the rank data to the template
    context = {
        'rank': {
            'id': rank.rank_id,
            'name': rank.rank_name,
            'details': rank.details,
            'created_on': rank.created_on,
        }
    }
    return render(request, "rank/view_rank.html", context)


def rank_update_page(request, id):
    # Fetch rank data by ID
    rank = get_object_or_404(Rank, rank_id=id)

    if request.method == "GET":
        # Render the update form with existing rank data
        return render(request, 'rank/update_rank.html', {'rank': rank})

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            # Update rank fields with the new data
            rank.rank_name = data.get('rank_name', rank.rank_name)
            rank.details = data.get('details', rank.details)
            rank.save()

            return JsonResponse({'message': 'Rack updated successfully!'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)


@csrf_exempt
def delete_rank(request, id):
    if request.method == 'POST':
        try:
            # Use rank_id instead of id
            rank = get_object_or_404(Rank, rank_id=id)
            rank.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method.'}, status=405)



def product_list_page(request):
    return render(request, "product/list_product.html")

def product_add_page(request):
    return render(request, "product/add_product.html")


def customer_list_page(request):
    return render(request, "customer/list_customer.html")

def customer_add_page(request):
    return render(request, "customer/add_customer.html")


def add_customer(request):
    if request.method == 'POST':
        try:
            # Parse the JSON data from the request body
            data = json.loads(request.body)

            # Fetch the company using the company_id (not company_name)
            company_id = data['company_name']  # This should be the company_id sent from the frontend
            company = Company.objects.get(company_id=company_id)  # Search by company_id
            
            # Create a new customer instance and save it
            customer = Customer.objects.create(
                customer_name=data['customer_name'],
                email=data['email'],
                company=company,  # Linking the customer to the company
                phone_number=data['phone_number'],
                address=data['address'],
                city=data['city'],
                state=data['state'],
                zip_code=data['zip_code'],
                country=data['country'],
                details=data['details']
            )

            return JsonResponse({'success': True, 'message': 'Customer added successfully!'})

        except Company.DoesNotExist:
            # Handle the case where the company does not exist
            return JsonResponse({'success': False, 'error': 'Company not found'})

        except Exception as e:
            # Catch any other exceptions and return the error
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def get_companies(request):
    companies = Company.objects.all()
    company_data = [{"company_id": company.company_id, "company_name": company.company_name} for company in companies]
    return JsonResponse(company_data, safe=False)


def customer_list(request):
    if request.method == "GET":
        customer_name = request.GET.get("customer_name", "")
        company_id = request.GET.get("company_id", "")

        customers = Customer.objects.all()

        if customer_name:
            customers = customers.filter(customer_name__icontains=customer_name)

        if company_id:
            customers = customers.filter(company_id=company_id)

        customer_list = [
            {
                'id': customer.customer_id,
                'name': customer.customer_name,
                'email': customer.email,
                'company_name': customer.company.company_name if customer.company else None,
                'phone_number': customer.phone_number,
                'address': customer.address,
                'city': customer.city,
                'state': customer.state,
                'zip_code': customer.zip_code,
                'country': customer.country,
                'details': customer.details,
            }
            for customer in customers
        ]
        return JsonResponse({'customers': customer_list}, status=200)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)



def customer_view_page(request, customer_id):
    # Fetch the specific customer by ID
    customer = get_object_or_404(Customer, customer_id=customer_id)

    # Pass the customer data to the template
    context = {
        'customer': {
            'id': customer.customer_id,
            'name': customer.customer_name,
            'email': customer.email,
            'phone_number': customer.phone_number,
            'address': customer.address,
            'city': customer.city,
            'state': customer.state,
            'zip_code': customer.zip_code,
            'country': customer.country,
            'details': customer.details,
        }
    }
    return render(request, "customer/view_customer.html", context)


def customer_update_page(request, id):
    # Fetch customer data
    customer = get_object_or_404(Customer, customer_id=id)

    if request.method == "GET":
        # Render the update form with existing data
        return render(request, 'customer/update_customer.html', {'customer': customer})

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            # Update customer fields
            customer.customer_name = data.get('name', customer.customer_name)
            customer.email = data.get('email', customer.email)
            customer.phone_number = data.get('phone_number', customer.phone_number)
            customer.address = data.get('address', customer.address)
            customer.city = data.get('city', customer.city)
            customer.state = data.get('state', customer.state)
            customer.zip_code = data.get('zip_code', customer.zip_code)
            customer.save()

            return JsonResponse({'message': 'Customer updated successfully!'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)



# Delete Customer View
@csrf_exempt
def delete_customer(request, id):
    if request.method == 'POST':
        try:
            customer = get_object_or_404(Customer, customer_id=id)
            customer.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def supplier_list_page(request):
    return render(request, "supplier/list_supplier.html")

def supplier_add_page(request):
    return render(request, "supplier/add_supplier.html")


def add_supplier(request):
    if request.method == 'POST':
        try:
            # Parse the JSON data from the request body
            data = json.loads(request.body)

            # Fetch the company using the company_id (not company_name)
            company_id = data['company_name']  # This should be the company_id sent from the frontend
            company = Company.objects.get(company_id=company_id)  # Search by company_id
            
            # Create a new supplier instance and save it
            supplier = Supplier.objects.create(
                supplier_name=data['supplier_name'],
                email=data['email'],
                company=company,  # Linking the supplier to the company
                phone_number=data['phone_number'],
                address=data['address'],
                city=data['city'],
                state=data['state'],
                zip_code=data['zip_code'],
                country=data['country'],
                details=data['details']
            )

            return JsonResponse({'success': True, 'message': 'Supplier added successfully!'})

        except Company.DoesNotExist:
            # Handle the case where the company does not exist
            return JsonResponse({'success': False, 'error': 'Company not found'})

        except Exception as e:
            # Catch any other exceptions and return the error
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def supplier_list(request):
    if request.method == "GET":
        supplier_name = request.GET.get('supplier_name', '').strip()
        company_name = request.GET.get('company_name', '').strip()

        suppliers = Supplier.objects.all()

        if supplier_name:
            suppliers = suppliers.filter(supplier_name__icontains=supplier_name)
        if company_name:  
            suppliers = suppliers.filter(company__company_id=company_name)  

        supplier_list = [
            {
                'id': supplier.supplier_id,
                'name': supplier.supplier_name,
                'email': supplier.email,
                'company_name': supplier.company.company_name if supplier.company else '',
                'phone_number': supplier.phone_number,
                'address': supplier.address,
                'city': supplier.city,
                'state': supplier.state,
                'zip_code': supplier.zip_code,
                'country': supplier.country,
                'details': supplier.details,
            }
            for supplier in suppliers
        ]
        return JsonResponse({'suppliers': supplier_list}, status=200)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)




def supplier_view_page(request, supplier_id):
    # Fetch the specific supplier by ID
    supplier = get_object_or_404(Supplier, supplier_id=supplier_id)

    # Pass the supplier data to the template
    context = {
        'supplier': {
            'id': supplier.supplier_id,
            'name': supplier.supplier_name,
            'email': supplier.email,
            'phone_number': supplier.phone_number,
            'address': supplier.address,
            'city': supplier.city,
            'state': supplier.state,
            'zip_code': supplier.zip_code,
            'country': supplier.country,
            'details': supplier.details,
        }
    }
    return render(request, "supplier/view_supplier.html", context)


def supplier_update_page(request, id):
    # Fetch supplier data
    supplier = get_object_or_404(Supplier, supplier_id=id)

    if request.method == "GET":
        # Fetch companies data to populate the company dropdown
        companies = Company.objects.all()

        # Render the update form with existing data
        return render(request, 'supplier/update_supplier.html', {'supplier': supplier, 'companies': companies})

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            # Update supplier fields
            supplier.supplier_name = data.get('supplier_name', supplier.supplier_name)
            supplier.email = data.get('email', supplier.email)
            supplier.phone_number = data.get('phone_number', supplier.phone_number)
            supplier.address = data.get('address', supplier.address)
            supplier.city = data.get('city', supplier.city)
            supplier.state = data.get('state', supplier.state)
            supplier.zip_code = data.get('zip_code', supplier.zip_code)
            supplier.country = data.get('country', supplier.country)
            supplier.details = data.get('details', supplier.details)
            supplier.company_id = data.get('company_id', supplier.company_id)  # Update company field
            supplier.save()

            return JsonResponse({'message': 'Supplier updated successfully!'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)



# Delete Supplier View
@csrf_exempt
def delete_supplier(request, id):
    if request.method == 'POST':
        try:
            supplier = get_object_or_404(Supplier, supplier_id=id)
            supplier.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method.'}, status=405)



def validate_image(file):
    # Check if the file type is an image (e.g., jpg, png)
    valid_image_types = ['image/jpeg', 'image/png', 'image/gif']
    if file.content_type not in valid_image_types:
        raise ValidationError("Invalid image format. Only JPG, PNG, and GIF are allowed.")
    
    # Check the file size (max 5MB in this example)
    max_size = 5 * 1024 * 1024  # 5 MB
    if file.size > max_size:
        raise ValidationError("Image size exceeds the maximum limit of 5MB.")

    # Open the image using Pillow
    img = Image.open(file)

    # Get the original image dimensions
    width, height = img.size

    # You can set a max width and height for resizing if necessary
    max_width, max_height = 1024, 1024

    # Resize the image if it's larger than the maximum allowed dimensions
    if width > max_width or height > max_height:
        img.thumbnail((max_width, max_height))  # Resize while maintaining aspect ratio

    # Save the image to a BytesIO object to handle file size reduction
    img_io = BytesIO()
    # Compress the image to reduce file size (quality can be adjusted)
    img.save(img_io, format='JPEG', quality=85)  # Adjust quality as needed (between 0 and 100)

    # Rewind the BytesIO object to the beginning
    img_io.seek(0)

    # Return the resized image file
    file.image = img_io  # Replace original file with the resized one (optional)

    # Check the new file size after resizing
    if img_io.tell() > max_size:
        raise ValidationError("Image size exceeds the maximum limit of 5MB after compression.")

    return file


def generate_product_code():
    """ Generate the next sequential product code in the format PROD-000001 """
    last_product = Product.objects.aggregate(Max('product_id'))
    last_id = last_product['product_id__max'] or 0  # If no product exists, start from 0
    new_id = last_id + 1
    return f"PROD-{new_id:06d}"  # Formats as PROD-000001, PROD-000002, etc.


def generate_barcode_image(barcode_text, product_id, category_shortcode):
    """ Generate and save a barcode image """
    barcode_class = barcode.get_barcode_class('code128')
    barcode_instance = barcode_class(barcode_text, writer=ImageWriter())

    # Create file path using product_id and category_shortcode
    file_path = os.path.join(settings.MEDIA_ROOT, 'barcodes', f'barcode_{product_id}_{category_shortcode}')
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    # Save the barcode without the .png extension
    barcode_instance.save(file_path)

    # Return relative path with .png extension
    return f'barcodes/barcode_{product_id}_{category_shortcode}.png'


def generate_barcode(product_code, product_name, category_shortcode):
    """ Generate a barcode string using product_code, name, and category shortcode """
    barcode_text = f"{product_code}-{slugify(product_name[:5])}-{slugify(category_shortcode[:5])}".upper()
    return barcode_text


def add_product(request):
    if request.method == 'POST':
        try:
            product_name = request.POST['product_name']
            category_id = request.POST['category_id']
            subcategory_id = request.POST['subcategory_id']
            length = request.POST['length']
            breadth = request.POST['breadth']
            height = request.POST['height']
            product_weight = request.POST['product_weight']
            manufacture_name = request.POST['manufacture_name']
            description = request.POST['description']
            vendor = request.POST['vendor']
            purchase_date = request.POST['purchase_date']
            purchase_amount = request.POST['purchase_amount']
            images = request.FILES.getlist('images')

            category = Category.objects.get(category_id=category_id)
            subcategory = SubCategory.objects.get(subcategory_id=subcategory_id)

            # Generate sequential product_code
            product_code = generate_product_code()

            # Create product without barcode initially
            product = Product.objects.create(
                product_code=product_code,
                product_name=product_name,
                category=category,
                subcategory=subcategory,
                product_size_length=length,
                product_size_breadth=breadth,
                product_size_height=height,
                product_weight=product_weight,
                manufacture_name=manufacture_name,
                description=description,
                vendor=vendor,
                purchase_date=purchase_date,
                purchase_amount=purchase_amount,
                barcode=product_code  # Assign generated product_code
            )

            # Generate barcode text and image using product_code instead of product_id
            barcode_text = generate_barcode(product_code, product_name, category.shortcode)
            barcode_image_path = generate_barcode_image(barcode_text, product_code,category.shortcode)

            # Update product with barcode details
            product.barcode = barcode_text
            product.barcode_image = barcode_image_path
            product.save()

            # Save multiple images
            for image in images:
                product_image = ProductImage.objects.create(image=image)
                product.product_images.add(product_image)

            return JsonResponse({'success': True, 'message': 'Product added successfully!'})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    categories = Category.objects.all()
    return render(request, 'add_product.html', {'categories': categories})


def get_subcategories(request, category_id):
    subcategories = list(SubCategory.objects.filter(category_id=category_id).values('subcategory_id', 'subcategory_name','shortcode'))
    return JsonResponse({'subcategories': subcategories})



def get_subcategories(request, category_id):
    subcategories = list(SubCategory.objects.filter(category_id=category_id).values('subcategory_id', 'subcategory_name','shortcode'))
    return JsonResponse({'subcategories': subcategories})


@csrf_exempt
def get_products(request):
    if request.method == "GET":
        try:
            name = request.GET.get('name', '').strip()
            pcode = request.GET.get('pcode', '').strip()
            category_id = request.GET.get('category', '').strip()
            subcategory_id = request.GET.get('subcategory', '').strip()

            products = Product.objects.all()

            if name:
                products = products.filter(product_name__icontains=name)

            if pcode:
                products = products.filter(product_code__icontains=pcode)

            if category_id:
                products = products.filter(category_id=category_id)

            if subcategory_id:
                products = products.filter(subcategory_id=subcategory_id)

            product_list = [
                {
                    'id': product.product_id,
                    'product_code': product.product_code,
                    'name': product.product_name,
                    'category': product.category.category_name if product.category else None,
                    'purchase_amount': product.purchase_amount,
                    'purchase_date': product.purchase_date.strftime('%Y-%m-%d'),
                    'image': product.product_images.first().image.url if product.product_images.exists() else None
                }
                for product in products
            ]

            return JsonResponse({'products': product_list}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def search_products(request):
    query = request.GET.get('q', '').strip()
    product_code_query = request.GET.get('pcode', '').strip()
    
    if query or product_code_query:
        products = Product.objects.all()

        if query:
            products = products.filter(product_name__icontains=query)
        if product_code_query:
            products = products.filter(product_code__icontains=product_code_query)
        
        product_list = [
            {
                'product_id': product.product_id,
                'product_name': product.product_name,
                'product_code': product.product_code,
                'category_id': product.category.category_id if product.category else None,
                'subcategory_id': product.subcategory.subcategory_id if product.subcategory else None
            }
            for product in products[:10]  # Limit to 10 results
        ]
        return JsonResponse({'products': product_list}, status=200)

    return JsonResponse({'products': []}, status=200)


def product_view_page(request, product_id):
    # Fetch the specific product by ID
    product = get_object_or_404(Product, product_id=product_id)

    # Pass the product data to the template
    context = {
        'product': {
            'id': product.product_id,
            'name': product.product_name,
            'product_code': product.product_code,
            'category': product.category.category_name,
            'subcategory': product.subcategory.subcategory_name if product.subcategory else 'N/A',
            'size': f"{product.product_size_length} x {product.product_size_breadth} x {product.product_size_height}",
            'weight': product.product_weight,
            'manufacture_name': product.manufacture_name,
            'description': product.description,
            'vendor': product.vendor,
            'barcode': product.barcode,
            'barcode_image': product.barcode_image.url if product.barcode_image else None,
            'purchase_date': product.purchase_date.strftime('%Y-%m-%d'),
            'purchase_amount': product.purchase_amount,
            'images': product.product_images.all(),
        }
    }
    return render(request, "product/view_product.html", context)


def update_product(request, id):
    product = get_object_or_404(Product, product_id=id)

    if request.method == "GET":
        categories = Category.objects.all()
        subcategories = SubCategory.objects.filter(category_id=product.category_id)
        return render(request, 'product/update_product.html', {
            'product': product,
            'categories': categories,
            'subcategories': subcategories
        })

    elif request.method == "POST":
        try:
            # Get product details from the form
            old_category_id = product.category_id  # Store the old category for comparison
            product.product_code = request.POST.get('product_code', product.product_code)
            product.product_name = request.POST.get('product_name', product.product_name)
            product.barcode = request.POST.get('barcode', product.barcode)

            # Use 'category' and 'subcategory' instead of 'category_id' and 'subcategory_id'
            new_category = get_object_or_404(Category, category_id=request.POST.get('category'))
            product.category = new_category
            product.subcategory = get_object_or_404(SubCategory, subcategory_id=request.POST.get('subcategory'))

            product.product_size_length = request.POST.get('length', product.product_size_length)
            product.product_size_breadth = request.POST.get('breadth', product.product_size_breadth)
            product.product_size_height = request.POST.get('height', product.product_size_height)
            product.product_weight = request.POST.get('product_weight', product.product_weight)
            product.manufacture_name = request.POST.get('manufacture_name', product.manufacture_name)
            product.description = request.POST.get('description', product.description)
            product.vendor = request.POST.get('vendor', product.vendor)

            # Handle purchase date
            purchase_date_str = request.POST.get('purchase_date')
            if purchase_date_str:
                product.purchase_date = datetime.strptime(purchase_date_str, '%Y-%m-%d').date()

            product.purchase_amount = request.POST.get('purchase_amount', product.purchase_amount)

            # Handle image deletions
            remove_image_ids = request.POST.getlist('remove_image_ids[]')  # Get IDs of images to remove
            if remove_image_ids:
                ProductImage.objects.filter(id__in=remove_image_ids).delete()

            # Handle new images upload
            images = request.FILES.getlist('images')
            for image in images:
                # Validate image before saving
                validate_image(image)
                product_image = ProductImage.objects.create(image=image)
                product.product_images.add(product_image)

           # Regenerate barcode if category or subcategory has changed
            if product.category != old_category_id:
                # Generate barcode text using the product's code, name, and the new category's shortcode
                barcode_text = generate_barcode(product.product_code, product.product_name, new_category.shortcode)

                # Regenerate barcode image
                barcode_image_path = generate_barcode_image(barcode_text, product.product_id, new_category.shortcode)
                product.barcode_image = barcode_image_path


            product.save()

            # Prepare response data
            product_images = [{'id': img.id, 'image_url': img.image.url} for img in product.product_images.all()]
            response_data = {
                'success': True,
                'message': 'Product updated successfully!',
                'barcode_image': product.barcode_image.url if product.barcode_image else None,
                'product_images': product_images  # Updated image list
            }

            return JsonResponse(response_data)

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def remove_product_image(request, image_id):
    try:
        # Get the image object by its ID
        image = ProductImage.objects.get(id=image_id)
        
        # Delete the image
        image.delete()

        return JsonResponse({'success': True, 'message': 'Image removed successfully.'})

    except ProductImage.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Image not found.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    
@csrf_exempt
def delete_product(request, id):
    if request.method == 'POST':
        try:
            product = get_object_or_404(Product, product_id=id)
            product.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method'}, status=405)


def exhibition_list_page(request):
    return render(request, "exhibition/list_exhibition.html")

def exhibition_add_page(request):
    return render(request, "exhibition/add_exhibition.html")


def add_exhibition(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            Exhibition.objects.create(
                exhibition_name=data['exhibition_name'],
                location=data['location'],
                address=data['address'],
                city=data['city'],
                state=data['state'],
                pincode=data['pincode'],
                start_date=data['start_date'],
                end_date=data['end_date'],
                details=data['details'],
            )
            return JsonResponse({'success': True, 'message': 'Exhibition added successfully!'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def exhibition_list(request):
    if request.method == "GET":
        exhibition_name = request.GET.get("exhibition_name", "").strip()
        location = request.GET.get("location", "").strip()

        exhibitions = Exhibition.objects.all()

        if exhibition_name:
            exhibitions = exhibitions.filter(exhibition_name__icontains=exhibition_name)

        if location:
            exhibitions = exhibitions.filter(location__icontains=location)

        exhibition_list = [
            {
                'id': exhibition.exhibition_id,
                'name': exhibition.exhibition_name,
                'location': exhibition.location,
                'city': exhibition.city,
                'state': exhibition.state,
                'start_date': exhibition.start_date.strftime('%Y-%m-%d'),
                'end_date': exhibition.end_date.strftime('%Y-%m-%d'),
            }
            for exhibition in exhibitions
        ]

        return JsonResponse({'exhibitions': exhibition_list}, status=200)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)



def exhibition_view_page(request, exhibition_id):
    # Fetch the specific exhibition by ID
    exhibition = get_object_or_404(Exhibition, exhibition_id=exhibition_id)

    # Pass the exhibition data to the template
    context = {
        'exhibition': {
            'id': exhibition.exhibition_id,
            'name': exhibition.exhibition_name,
            'location': exhibition.location,
            'address': exhibition.address,
            'city': exhibition.city,
            'state': exhibition.state,
            'pincode': exhibition.pincode,
            'start_date': exhibition.start_date,
            'end_date': exhibition.end_date,
            'details': exhibition.details,
        }
    }
    return render(request, "exhibition/view_exhibition.html", context)


def exhibition_update_page(request, id):
    exhibition = get_object_or_404(Exhibition, exhibition_id=id)

    if request.method == "GET":
        # Render the update form with existing data
        return render(request, 'exhibition/update_exhibition.html', {'exhibition': exhibition})

    elif request.method == "POST":
        try:
            data = json.loads(request.body)

            # Parse and update the fields
            exhibition.exhibition_name = data.get('exhibition_name', exhibition.exhibition_name)
            exhibition.location = data.get('location', exhibition.location)
            exhibition.address = data.get('address', exhibition.address)
            exhibition.city = data.get('city', exhibition.city)
            exhibition.state = data.get('state', exhibition.state)
            exhibition.pincode = data.get('pincode', exhibition.pincode)

            # Parse and update start_date and end_date
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            if start_date:
                exhibition.start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            if end_date:
                exhibition.end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

            exhibition.details = data.get('details', exhibition.details)
            exhibition.save()

            return JsonResponse({'message': 'Exhibition updated successfully!'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)


@csrf_exempt
def delete_exhibition(request, id):
    if request.method == 'POST':
        try:
            exhibition = get_object_or_404(Exhibition, exhibition_id=id)
            exhibition.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method.'}, status=405)




def stock_list_page(request):
    return render(request, "stock/stock_list.html")

def stock_entry_page(request):
    return render(request, "stock/stock_entry.html")


# Fetch products based on selected subcategory
def get_products_stock(request, subcategory_id):
    products = list(Product.objects.filter(subcategory_id=subcategory_id)
                    .values('product_id', 'product_name', 'barcode_image'))
    return JsonResponse({'products': products})

# Fetch locations for Select2 dropdown
def get_locations_entry(request):
    locations = list(Location.objects.filter(status="active").values('id', 'name', 'shortcode'))
    return JsonResponse({'locations': locations})


from PIL import Image, ImageDraw, ImageFont
import qrcode
import os
from django.conf import settings

def generate_barcode_stock_image(product_code, product_name, product_size_length, product_size_breadth, 
                               product_size_height, product_weight, manufacture_name, category_shortcode, 
                               location_shortcode, quantity_number):
    """ Generate and save a QR code with barcode text displayed below. """

    # Encode full details inside the QR code
    qr_data = (
        f"Barcode Text: {product_code}\n"
        f"Name: {product_name}\n"
        f"Size: {product_size_length}x{product_size_breadth}x{product_size_height} cm\n"
        f"Weight: {product_weight} kg\n"
        f"Manufacturer: {manufacture_name}\n"
        f"Category: {category_shortcode}\n"
        f"Location: {location_shortcode}\n"
        f"Quantity: {quantity_number}"
    )

    # Generate QR Code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill='black', back_color='white')

    # Load font
    font_path = os.path.join(settings.BASE_DIR, 'static', 'assets', 'fonts', 'Arial.ttf')
    try:
        font = ImageFont.truetype(font_path, 22)
    except IOError:
        font = ImageFont.load_default()

    # Get QR code size
    qr_width, qr_height = qr_img.size

    # Get text size
    barcode_text = f"{product_code}-{category_shortcode}-{location_shortcode}-{str(quantity_number).zfill(2)}"
    temp_img = Image.new("RGB", (1000, 1000), "white")
    draw = ImageDraw.Draw(temp_img)
    bbox = draw.textbbox((0, 0), barcode_text, font=font)
    text_width, text_height = bbox[2] - bbox[0], bbox[3] - bbox[1]

    # Create final image canvas
    img_width = max(qr_width + 40, text_width + 40)
    img_height = qr_height + text_height + 50
    final_img = Image.new("RGB", (img_width, img_height), "white")

    # Center QR code
    qr_x = (img_width - qr_width) // 2
    qr_y = 10
    final_img.paste(qr_img, (qr_x, qr_y))

    # Draw barcode text
    draw = ImageDraw.Draw(final_img)
    text_x = (img_width - text_width) // 2
    text_y = qr_y + qr_height + 10
    draw.text((text_x, text_y), barcode_text, font=font, fill="black")

    # Save barcode image
    temp_folder = os.path.join(settings.MEDIA_ROOT, 'temp_barcodes')
    os.makedirs(temp_folder, exist_ok=True)

    file_name = f"{barcode_text}.png"
    file_path = os.path.join(temp_folder, file_name)

    final_img.save(file_path)

    return f'temp_barcodes/{file_name}'




def move_barcode_to_stock(temp_image_url, stock_entry):
    """ Move barcode image from temporary folder to stock folder and delete it from temp_barcodes """
    
    # Resolve full path of the temporary image
    temp_image_path = os.path.join(settings.MEDIA_ROOT, temp_image_url.replace(settings.MEDIA_URL, ""))
    
    # Ensure the file exists before proceeding
    if not os.path.exists(temp_image_path):
        print(f"Warning: Temp image {temp_image_path} not found!")
        return None

    # Define new folder for storing barcode images permanently
    new_folder = os.path.join(settings.MEDIA_ROOT, 'stock_barcodes')
    os.makedirs(new_folder, exist_ok=True)  # Create the folder if it doesn't exist

    # Get the file name and define the new storage path
    new_image_name = os.path.basename(temp_image_path)
    new_image_path = os.path.join(new_folder, new_image_name)

    try:
        # Move the file from temp_barcodes to stock_barcodes
        shutil.move(temp_image_path, new_image_path)

        print(f"Barcode image moved to: {new_image_path}")
        
        # Return the new relative path for saving in the database
        return f'stock_barcodes/{new_image_name}'

    except Exception as e:
        print(f"Error moving barcode image: {e}")
        return None



def add_stock(request):   
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        location_id = request.POST.get('location_id')
        category_id = request.POST.get('category_id')
        subcategory_id = request.POST.get('subcategory_id')
        quantity = int(request.POST.get('quantity', 1))
        serial_numbers = request.POST.getlist('serial_number[]')

        if not product_id or not location_id or not category_id:
            return JsonResponse({"status": "error", "message": "Missing required fields"}, status=400)

        try:
            product = Product.objects.get(product_id=product_id)
            category = Category.objects.filter(category_id=category_id).first()
            subcategory = SubCategory.objects.filter(subcategory_id=subcategory_id).first()
            location = Location.objects.get(id=location_id)

            if not category:
                return JsonResponse({"status": "error", "message": "Invalid Category ID"}, status=400)

            barcode_list = []

            # Get existing barcodes from TempStockBarcode
            existing_temp_barcodes = TempStockBarcode.objects.filter(
                product=product, location=location
            )

            if existing_temp_barcodes.exists():
                barcode_list = [
                    {
                        "barcode": temp_barcode.barcode_text,
                        "image": temp_barcode.barcode_image.url if temp_barcode.barcode_image else None,
                        "product_name": temp_barcode.product.product_name,
                        "location_name": temp_barcode.location.name,
                        "serial_number": temp_barcode.serial_number
                    }
                    for temp_barcode in existing_temp_barcodes
                ]

                return JsonResponse({
                    "status": "success",
                    "barcodes": barcode_list,
                    "product_name": product.product_name,
                    "location_name": location.name
                })

            # If no barcodes exist in TempStockBarcode, check StockBarcode
            existing_barcodes = list(StockBarcode.objects.filter(
                barcode_text__startswith=f"{product.product_code}-{category.shortcode}-{location.shortcode}-"
            ).values_list('barcode_text', flat=True))

            # Extract numerical part of existing barcodes
            quantity_numbers = []
            pattern = re.compile(rf"{product.product_code}-{category.shortcode}-{location.shortcode}-(\d+)")

            for barcode in existing_barcodes:
                match = pattern.match(barcode)
                if match:
                    quantity_numbers.append(int(match.group(1)))

            last_number = max(quantity_numbers) if quantity_numbers else 0

            for i in range(1, quantity + 1):
                quantity_number = last_number + i

                while True:
                    barcode_text = f"{product.product_code}-{category.shortcode}-{location.shortcode}-{str(quantity_number).zfill(2)}".upper()
                    if barcode_text not in existing_barcodes:
                        break
                    quantity_number += 1

                # Generate barcode image
                temp_barcode_path = generate_barcode_stock_image(
                    barcode_text,
                    product.product_name,  # ✅ Add missing product name
                    product.product_size_length,  # ✅ Add missing size length
                    product.product_size_breadth,  # ✅ Add missing size breadth
                    product.product_size_height,  # ✅ Add missing size height
                    product.product_weight,  # ✅ Product weight
                    product.manufacture_name,  # ✅ Manufacture name
                    category.shortcode,  # ✅ Category shortcode
                    location.shortcode,  # ✅ Location shortcode
                    quantity_number  # ✅ Quantity number
                )

                # Get corresponding serial number from user input
                serial_number = serial_numbers[i - 1] if i - 1 < len(serial_numbers) else ""


                # ✅ Save Product & Location in TempStockBarcode
                temp_barcode = TempStockBarcode.objects.create(
                    barcode_text=barcode_text,
                    barcode_image=temp_barcode_path,
                    product=product,  
                    location=location,
                    serial_number=serial_number  
                )

                barcode_list.append({
                    "barcode": temp_barcode.barcode_text,
                    "image": temp_barcode.barcode_image.url,
                    "product_name": product.product_name,
                    "location_name": location.name,
                    "serial_number": temp_barcode.serial_number
                })

            return JsonResponse({
                "status": "success",
                "barcodes": barcode_list,
                "product_name": product.product_name,
                "location_name": location.name
            })

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

    return JsonResponse({"status": "error", "message": "Invalid request"}, status=400)


def get_existing_barcodes(request):
    product_id = request.GET.get('product_id')
    location_id = request.GET.get('location_id')

    if not product_id or not location_id:
        return JsonResponse({"status": "error", "message": "Missing required fields"}, status=400)

    try:
        product = Product.objects.get(product_id=product_id)
        location = Location.objects.get(id=location_id)

        existing_barcodes = TempStockBarcode.objects.filter(product=product, location=location)

        barcode_list = [
            {
                "barcode": barcode.barcode_text,
                "image": barcode.barcode_image.url
            } 
            for barcode in existing_barcodes
        ]

        return JsonResponse({
            "status": "success",
            "barcodes": barcode_list,
            "product_name": product.product_name,
            "location_name": location.name,
        })

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=400)
    
    

@csrf_exempt
def submit_stock(request):
    """ Save Stock Entry and Barcodes to Database and delete TempStockBarcode """
    if request.method == 'POST':
        try:
            # Debugging: Log received data
            print("Received Data:", request.POST)

            # Extract form data
            product_id = request.POST.get('product_id')
            location_id = request.POST.get('location_id')
            category_id = request.POST.get('category_id')
            subcategory_id = request.POST.get('subcategory_id')
            quantity = int(request.POST.get('quantity', 1))
            product_status_list = request.POST.getlist('product_status[]')  # Extract as list
            import ast
            barcodes_json = request.POST.get('barcodes', '[]')
            barcode_list = ast.literal_eval(barcodes_json) if isinstance(barcodes_json, str) else []

            if not barcode_list:
                return JsonResponse({"status": "error", "message": "No barcodes received"}, status=400)

            # Debugging: Print extracted values
            print(f"Product ID: {product_id}, Location ID: {location_id}, Quantity: {quantity}, Barcodes: {barcode_list}")

            product = Product.objects.get(product_id=product_id)
            location = Location.objects.get(id=location_id)
            category = Category.objects.get(category_id=category_id)
            subcategory = SubCategory.objects.get(subcategory_id=subcategory_id)

            # Process each barcode
            for index, barcode in enumerate(barcode_list):
                rank_id = barcode.get("rank_id")
                box_id = barcode.get("box_id")
                serial_number = barcode.get("serial_number", "")
                product_status = product_status_list[index]  # Get product_status for this barcode

                rack = Rank.objects.get(rank_id=rank_id)
                box = Box.objects.get(box_id=box_id)

                stock_entry = StockEntry.objects.create(
                    product=product,
                    location=location,
                    category=category,
                    subcategory=subcategory,
                    rack=rack,
                    box=box,
                    quantity=1,  # Each barcode corresponds to 1 quantity
                    serial_number=serial_number,
                    product_status=product_status  # Pass product_status here
                )

                # Fetch the temp barcode object
                temp_barcode = TempStockBarcode.objects.get(barcode_text=barcode['barcode'])

                # Move barcode image and save in StockBarcode
                stock_image_path = move_barcode_to_stock(temp_barcode.barcode_image.url, stock_entry)
                StockBarcode.objects.create(
                    stock_entry=stock_entry,
                    barcode_text=temp_barcode.barcode_text,
                    barcode_image=stock_image_path
                )

                # Delete the temp barcode
                temp_barcode.delete()

            return JsonResponse({"status": "success", "message": "Stock submitted successfully"})

        except Exception as e:
            print("Error:", str(e))
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

    return JsonResponse({"status": "error", "message": "Invalid request"}, status=400)


def delete_temp_barcode(request):
    """ Delete a barcode from temp folder and remove its entry from the database """
    if request.method == 'POST':
        temp_image_path = request.POST.get('image_path')  # Get full URL from AJAX
        barcode_text = request.POST.get('barcode_text')  # Get barcode text
        
        # Convert full URL to relative path
        media_url = settings.MEDIA_URL  # Example: '/media/'
        if temp_image_path.startswith(media_url):
            relative_path = temp_image_path[len(media_url):]  # Remove '/media/' prefix
        else:
            relative_path = temp_image_path  # Fallback

        try:
            # Find the barcode in the database using barcode_text (safer than image path)
            barcode_obj = TempStockBarcode.objects.filter(barcode_text=barcode_text).first()
            
            if barcode_obj:
                # Delete the barcode record from the database
                barcode_obj.delete()

                # Delete the barcode image from the filesystem
                full_path = os.path.join(settings.MEDIA_ROOT, relative_path)
                if default_storage.exists(full_path):
                    default_storage.delete(full_path)

                return JsonResponse({"status": "success", "message": "Barcode deleted from DB and filesystem"})

            return JsonResponse({"status": "error", "message": "Barcode not found in DB"}, status=404)

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

    return JsonResponse({"status": "error", "message": "Invalid request"}, status=400)


def remove_barcode(request):
    """ Remove barcode entry from database """
    if request.method == 'POST':
        barcode_text = request.POST.get('barcode')

        try:
            barcode = StockBarcode.objects.get(barcode_text=barcode_text)
            barcode.barcode_image.delete()  # Delete image file
            barcode.delete()
            return JsonResponse({"status": "success", "message": "Barcode removed successfully"})
        except StockBarcode.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Barcode not found"}, status=400)

    return JsonResponse({"status": "error", "message": "Invalid request"}, status=400)

def get_boxes(request):
    boxes = list(Box.objects.values('box_id', 'box_name'))
    return JsonResponse({'boxes': boxes})

def get_ranks(request):
    ranks = list(Rank.objects.values('rank_id', 'rank_name'))
    return JsonResponse({'ranks': ranks})

    
def stock_entry_fetch_barcode(request):
    category_id = request.GET.get("category")
    subcategory_id = request.GET.get("subcategory")
    product_id = request.GET.get("product")
    location_id = request.GET.get("location")

    print(f"Stock Entry Received Parameters - Category: {category_id}, SubCategory: {subcategory_id}, Product: {product_id}, Location: {location_id}")
    stock_entries = StockEntry.objects.filter(
        category_id=category_id,
        subcategory_id=subcategory_id,
        product_id=product_id,
        location_id=location_id,
    ).select_related("product", "location", "rack", "box")

    total_quantity = stock_entries.aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0

    if stock_entries.exists():
        print(f"Stock Entries Found: {stock_entries.count()} | Total Quantity: {total_quantity}")

        barcode_list = []

        for stock_entry in stock_entries:
            # Fetch all barcodes associated with each stock entry
            barcodes = stock_entry.barcodes.all()  
            barcode_data = [
                {
                    "barcode_text": barcode.barcode_text,
                    "barcode_image": barcode.barcode_image.url,
                    "stock_id": stock_entry.stock_id,
                    "serial_number": stock_entry.serial_number,
                    "product_name": stock_entry.product.product_name,
                    "location": stock_entry.location.name,
                    "rank": stock_entry.rack.rank_name,  
                    "box": stock_entry.box.box_name,
                    "product_status": stock_entry.product_status,
                }
                for barcode in barcodes
            ]
            barcode_list.extend(barcode_data)

        print(f"Barcodes Found: {barcode_list}")  
        return JsonResponse({"barcodes": barcode_list, "available_quantity": total_quantity})

    else:
        print("No Stock Entries Found!")
        return JsonResponse({"barcodes": [], "available_quantity": 0})

def fetch_barcode(request):
    category_id = request.GET.get("category")
    subcategory_id = request.GET.get("subcategory")
    product_id = request.GET.get("product")
    location_id = request.GET.get("location")

    print(f"Received Parameters - Category: {category_id}, SubCategory: {subcategory_id}, Product: {product_id}, Location: {location_id}")

    # Fetch matching stock entries
    stock_entries = StockEntry.objects.filter(
        category_id=category_id,
        subcategory_id=subcategory_id,
        product_id=product_id,
        location_id=location_id,
        product_status ="INSTOCK"
    ).select_related("product", "location", "rack", "box")

    total_quantity = stock_entries.aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0

    if stock_entries.exists():
        print(f"Stock Entries Found: {stock_entries.count()} | Total Quantity: {total_quantity}")

        barcode_list = []

        for stock_entry in stock_entries:
            # Fetch all barcodes associated with each stock entry
            barcodes = stock_entry.barcodes.all()  
            barcode_data = [
                {
                    "barcode_text": barcode.barcode_text,
                    "barcode_image": barcode.barcode_image.url,
                    "stock_id": stock_entry.stock_id,
                    "serial_number": stock_entry.serial_number,
                    "product_name": stock_entry.product.product_name,
                    "location": stock_entry.location.name,
                    "rank": stock_entry.rack.rank_name,  
                    "box": stock_entry.box.box_name,
                    "product_status": stock_entry.product_status,
                }
                for barcode in barcodes
            ]
            barcode_list.extend(barcode_data)

        print(f"Barcodes Found: {barcode_list}")  
        return JsonResponse({"barcodes": barcode_list, "available_quantity": total_quantity})

    else:
        print("No Stock Entries Found!")
        return JsonResponse({"barcodes": [], "available_quantity": 0})

@csrf_exempt
def delete_stock(request, stock_id):
    """ Deletes stock entry, related barcodes, and barcode images """
    if request.method == "DELETE":
        try:
            stock_entry = StockEntry.objects.get(stock_id=stock_id)

            # Delete associated barcodes and remove their images
            for barcode in stock_entry.barcodes.all():
                # Resolve the barcode image path
                barcode_image_path = os.path.join(settings.MEDIA_ROOT, str(barcode.barcode_image))

                # Remove image file if it exists
                if os.path.exists(barcode_image_path):
                    os.remove(barcode_image_path)

                # Delete barcode entry from the database
                barcode.delete()

            # Delete the stock entry
            stock_entry.delete()

            return JsonResponse({"status": "success", "message": "Stock entry deleted successfully"})
        except StockEntry.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Stock entry not found"}, status=404)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

    return JsonResponse({"status": "error", "message": "Invalid request method"}, status=400)


@csrf_exempt
def update_stock(request):
    """ Update stock's rank and box based on user selection """
    if request.method == "POST":
        try:
            data = json.loads(request.body)  # Parse JSON data
            stock_id = data.get("stock_id")
            new_rank_id = data.get("rank_id")  # Corrected field name
            new_box_id = data.get("box_id")

            # Ensure required data is provided
            if not stock_id or not new_rank_id or not new_box_id:
                return JsonResponse({"status": "error", "message": "Missing required fields"}, status=400)

            # Fetch the stock entry
            stock_entry = StockEntry.objects.get(stock_id=stock_id)  # Use stock_id as primary key
            
            # Fetch the new rank and box using correct primary keys
            new_rank = Rank.objects.get(rank_id=new_rank_id)  # Use rank_id instead of id
            new_box = Box.objects.get(box_id=new_box_id)  # Use box_id instead of id

            # Update the stock entry
            stock_entry.rack = new_rank  # Ensure this matches ForeignKey field
            stock_entry.box = new_box
            stock_entry.save()

            return JsonResponse({"status": "success", "message": "Stock updated successfully!"})

        except StockEntry.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Stock entry not found"}, status=404)
        except Rank.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Rank not found"}, status=404)
        except Box.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Box not found"}, status=404)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

    return JsonResponse({"status": "error", "message": "Invalid request"}, status=400)


def fetch_stock_list(request):
    if request.method == "GET":
        category_id = request.GET.get("category", None)
        subcategory_id = request.GET.get("subcategory", None)
        product_id = request.GET.get("product", None)
        location_id = request.GET.get("location", None)

        stock_entries = StockEntry.objects.select_related(
            "product","category", "subcategory", "location", "rack", "box"
        ).all()

        # Apply filters only if values exist
        if category_id:
            stock_entries = stock_entries.filter(category_id=category_id)
        if subcategory_id:
            stock_entries = stock_entries.filter(subcategory_id=subcategory_id)
        if product_id:
            stock_entries = stock_entries.filter(product_id=product_id)
        if location_id:
            stock_entries = stock_entries.filter(location_id=location_id)

        stock_list = [
            {
                "stock_id": stock.stock_id,
                "product_name": stock.product.product_name,
                "category_name": stock.category.category_name,
                "subcategory_name": stock.subcategory.subcategory_name,
                "barcode_text": barcode.barcode_text,
                "barcode_image": barcode.barcode_image.url if barcode.barcode_image else "",
                "location": stock.location.name,
                "rack": stock.rack.rank_name if stock.rack else "N/A",
                "box": stock.box.box_name if stock.box else "N/A",
                "product_status" : stock.product_status
            }
            for stock in stock_entries
            for barcode in stock.barcodes.all()  # Assuming a ForeignKey or related_name="barcodes"
        ]

        return JsonResponse({"stocks": stock_list}, status=200)

    return JsonResponse({"error": "Invalid request method."}, status=405)

def stock_view_page(request, stock_id): 
    # Fetch the specific stock entry by ID
    stock = get_object_or_404(StockEntry, stock_id=stock_id)

    # Fetch all related barcodes from StockBarcode
    barcodes = StockBarcode.objects.filter(stock_entry=stock)

    # Prepare stock data for template rendering
    context = {
        'stock': {
            'stock_id': stock.stock_id,
            'product_name': stock.product.product_name,
            'category_name': f"{stock.category.category_name} ({stock.category.shortcode})" if stock.category else 'N/A',
            'subcategory_name': f"{stock.subcategory.subcategory_name} ({stock.subcategory.shortcode})" if stock.subcategory else 'N/A',
            'location': f"{stock.location.name} ({stock.location.shortcode})" if stock.location else 'N/A',
            'rack': stock.rack.rank_name if stock.rack else 'N/A',
            'box': stock.box.box_name if stock.box else 'N/A',
        },
        'barcodes': barcodes  # Pass all barcode entries to the template
    }
    
    return render(request, "stock/view_stock.html", context)


def stock_update_page(request, stock_id):
    stock = get_object_or_404(StockEntry, stock_id=stock_id)

    if request.method == "GET":
        # Fetch related barcodes including proper image URLs
        barcodes = StockBarcode.objects.filter(stock_entry=stock)
        barcode_data = [
            {
                'barcode_text': barcode.barcode_text,
                'barcode_image_url': barcode.barcode_image.url if barcode.barcode_image else None
            }
            for barcode in barcodes
        ]

        # Fetch available racks and boxes
        racks = Rank.objects.all().values('rank_id', 'rank_name')
        boxes = Box.objects.all().values('box_id', 'box_name')

        context = {
            'stock': {
                'stock_id': stock.stock_id,
                'product_name': stock.product.product_name,
                'category_name': stock.category.category_name if stock.category else 'N/A',
                'subcategory_name': stock.subcategory.subcategory_name if stock.subcategory else 'N/A',
                'location': stock.location.name if stock.location else 'N/A',
                'rack': stock.rack.rank_id if stock.rack else '',
                'box': stock.box.box_id if stock.box else '',
            },
            'racks': list(racks),
            'boxes': list(boxes),
            'barcodes': barcode_data,  # Now correctly formatted with image URLs
        }

        return render(request, "stock/update_stock.html", context)

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            rack_id = data.get('rack')
            box_id = data.get('box')

            stock.rack = Rank.objects.get(rank_id=rack_id) if rack_id else None
            stock.box = Box.objects.get(box_id=box_id) if box_id else None
            stock.save()

            return JsonResponse({'message': 'Stock updated successfully!'}, status=200)

        except Rank.DoesNotExist:
            return JsonResponse({'error': 'Invalid rack selected'}, status=400)
        except Box.DoesNotExist:
            return JsonResponse({'error': 'Invalid box selected'}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def stock_scan_page(request):
    return render(request, "scanstock/scan_stock.html")


def assign_list_page(request):
    return render(request, "assign/list_assign.html")


def assign_stock_page(request):
    return render(request, "assign/assign_stock.html")


def get_employees(request):
    employees = Employee.objects.all().values("employee_id", "name", "designation")
    return JsonResponse({"employees": list(employees)})  # Return all employees

def get_exhibitions(request):
    exhibitions = Exhibition.objects.all().values("exhibition_id", "exhibition_name", "city")
    return JsonResponse({"exhibitions": list(exhibitions)})


def get_companies(request):
    """
    Fetch all companies and return as JSON response.
    """
    companies = Company.objects.all().values("company_id", "company_name")
    return JsonResponse({"companies": list(companies)})


def get_customers(request):
    customers = Customer.objects.all().values('customer_id', 'customer_name', 'email', 'company_id', 'phone_number', 'address', 'city', 'state', 'zip_code', 'country', 'details')
    
    return JsonResponse({"customers": list(customers)})


from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings  # Import settings

def save_assigned_stock(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            assign_id = uuid.uuid4()
            print("The product Assign Code Is : ", data)

            # Dictionary to group products by user
            user_products = {}

            for item in data:
                print("The item is:", item)
                stock_product = StockEntry.objects.select_related(
                    'product',  # Fetch related product
                    'product__category',  # Fetch related category of the product
                    'product__subcategory'  # Fetch related subcategory of the product
                ).get(stock_id=item['stockId'])

                product = stock_product.product
                assign_product_mode = item['assign_product_mode']

                assign_employee = None
                assign_company = None
                assign_exhibition = None
                assign_customer = None

                assign_stock_id = stock_product
                if item['mode'] == "Users":
                    assign_employee = Employee.objects.get(employee_id=item['relatedFields']['employee'])
                    user_key = f"employee_{assign_employee.employee_id}"
                elif item['mode'] == "Company":
                    assign_company = Company.objects.get(company_id=item['relatedFields']['company'])
                    user_key = f"company_{assign_company.company_id}"
                elif item['mode'] == "Exhibition":
                    assign_exhibition = Exhibition.objects.get(exhibition_id=item['relatedFields']['exhibition'])
                    assign_employee = Employee.objects.get(employee_id=item['relatedFields']['employee'])
                    user_key = f"exhibition_{assign_exhibition.exhibition_id}"
                elif item['mode'] == "Customers":
                    assign_customer = Customer.objects.get(customer_id=item['relatedFields']['customer'])
                    user_key = f"customer_{assign_customer.customer_id}"
                else:
                    raise ValueError("Invalid mode provided")

                # Create and save the AssignProduct instance
                assign_product = AssignProduct(
                    assign_id=assign_id,
                    stock_id=assign_stock_id,
                    assign_mode=item['mode'],
                    assign_product=product,
                    assign_employee=assign_employee,
                    assign_company=assign_company,
                    assign_exhibition=assign_exhibition,
                    assign_customer=assign_customer,
                    assign_quantity=1,  # Assuming 1 quantity per item
                    assign_customer_details=item['relatedFields'].get('customerDetails', ''),  # Optional field
                    assign_at=timezone.now()
                )
                assign_product.save()

                return_assign_product = ReturnProductHistory(
                    assign_return_id=assign_id,
                    return_id="",
                    return_stock_id=assign_stock_id,
                    return_mode=item['mode'],
                    return_product=product,
                    return_employee=assign_employee,
                    return_company=assign_company,
                    return_exhibition=assign_exhibition,
                    return_customer=assign_customer,
                    return_quantity=0,  # Assuming 1 quantity per item
                    return_customer_details=item['relatedFields'].get('customerDetails', ''),  # Optional field
                    return_product_status=assign_product_mode,
                    reuturn_assign_at=timezone.now(),
                    return_at="-",
                )
                return_assign_product.save()

                stock_product.product_status = assign_product_mode
                stock_product.save()

                # Group products by user
                if user_key not in user_products:
                    user_products[user_key] = {
                        'user': assign_employee or assign_company or assign_exhibition or assign_customer,
                        'products': []
                    }
                user_products[user_key]['products'].append({
                    'product': product,
                    'stock_id': item['stockId'],
                    'barcode_text': item['barcodeText'],
                    'assign_product_mode': item['assign_product_mode']
                })

            # Send email to each user with their assigned products
            for user_key, user_data in user_products.items():
                user = user_data['user']
                products = user_data['products']

                # Prepare email content
                subject = "Your Assigned Products"
                html_message = render_to_string('assign_product_email.html', {
                    'user': user,
                    'products': products,  # Pass grouped products with item details
                    'assign_id': assign_id,
                })
                plain_message = strip_tags(html_message)
                from_email = settings.EMAIL_HOST_USER
                to_email = user.email  # Ensure the user model has an email field

                # Send email
                send_mail(subject, plain_message, from_email, [to_email], html_message=html_message)

            return JsonResponse({"success": True})
        except Exception as e:
            traceback_str = traceback.format_exc()
            print("Traceback:", traceback_str)
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid request method"})

@csrf_exempt
def get_all_assigned_data(request):
    category_id = request.GET.get("category")
    subcategory_id = request.GET.get("subcategory")
    product_id = request.GET.get("product")
    location_id = request.GET.get("location")
    mode = request.GET.get("mode")
    assigned_product_count = AssignProduct.objects.filter(assign_mode=mode).count()
    print("assigned_product_count",assigned_product_count)
    stock = AssignProduct.objects.filter(
        assign_product__category__category_id=category_id,
        assign_product__subcategory__subcategory_id=subcategory_id, 
        assign_product__product_id=product_id,
        stock_id__location = location_id,
        assign_mode = mode,
    )
    stock_data = []
    for entry in stock:
            entry_dict = {
                "assign_id": entry.assign_id,
                "product_name": entry.assign_product.product_name,
                "category_name": entry.assign_product.category.category_name,
                "subcategory_name": entry.assign_product.subcategory.subcategory_name,
                "assign_mode": entry.assign_mode,
                "product_status": entry.stock_id.product_status,
            }
            stock_data.append(entry_dict)
    return JsonResponse(stock_data, safe=False)

def return_stock(request):
    return render(request, "stock/stock_return.html")

@csrf_exempt
def get_assign_data_all(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            type = data.get("type")
            value = data.get("value")
            print("The Type Is:", type, "The Value Is:", value)
            assigned_products = []
            if type == "User":
                assign_type ="Users"
                assign_employee = AssignProduct.objects.filter(assign_employee__employee_id=value,assign_mode = assign_type)
                for i in assign_employee:
                    barcode_text = i.stock_id.barcodes.first().barcode_text if i.stock_id.barcodes.exists() else " - "
                    assigned_products.append({
                        "assign_id": i.assign_id,
                        "stock_id": i.stock_id.stock_id,
                        "barcode_text": barcode_text,
                        "product_name": i.assign_product.product_name,
                        "product_category": i.assign_product.category.category_name,
                        "product_subcategory": i.assign_product.subcategory.subcategory_name,
                        "employee_name": i.assign_employee.name,
                        "mode": assign_type,
                        "product_status": i.stock_id.product_status,
                        
                    })
            elif type == "Company":
                assign_company = AssignProduct.objects.filter(assign_company__company_id=value,assign_mode = type)
                for i in assign_company:
                    barcode_text = i.stock_id.barcodes.first().barcode_text if i.stock_id.barcodes.exists() else " - "
                    assigned_products.append({
                        "assign_id": i.assign_id,
                        "stock_id": i.stock_id.stock_id,
                        "barcode_text": barcode_text,
                        "product_name": i.assign_product.product_name,
                        "product_category": i.assign_product.category.category_name,
                        "product_subcategory": i.assign_product.subcategory.subcategory_name,
                        "company_name": i.assign_company.company_name,
                        "mode": type,
                        "product_status": i.stock_id.product_status,
                    })
            elif type == "Exhibition":
                try:
                    assign_exhibition = AssignProduct.objects.filter(assign_exhibition__exhibition_id=value,assign_mode = type)
                except:
                    assign_exhibition = AssignProduct.objects.filter(assign_mode = type)
                for i in assign_exhibition:
                    barcode_text = i.stock_id.barcodes.first().barcode_text if i.stock_id.barcodes.exists() else " - "
                    assigned_products.append({
                        "assign_id": i.assign_id,
                        "stock_id": i.stock_id.stock_id,
                        "barcode_text": barcode_text,
                        "product_name": i.assign_product.product_name,
                        "product_category": i.assign_product.category.category_name,
                        "product_subcategory": i.assign_product.subcategory.subcategory_name,
                        "exhibition_name": i.assign_exhibition.exhibition_name,
                        "mode": type,
                        "product_status": i.stock_id.product_status,
                        "employee_name": i.assign_employee.name,
                    })
            elif type == "Customer":
                assign_mode = "Customers"
                assign_customer = AssignProduct.objects.filter(assign_customer__customer_id=value,assign_mode = assign_mode)
                for i in assign_customer:
                    barcode_text = i.stock_id.barcodes.first().barcode_text if i.stock_id.barcodes.exists() else " - "
                    assigned_products.append({
                        "assign_id": i.assign_id,
                        "stock_id": i.stock_id.stock_id,
                        "barcode_text": barcode_text,
                        "product_name": i.assign_product.product_name,
                        "product_category": i.assign_product.category.category_name,
                        "product_subcategory": i.assign_product.subcategory.subcategory_name,
                        "customer_name": i.assign_customer.customer_name,
                        "mode": type,
                        "product_status": i.stock_id.product_status,
                    })
            elif type == "ALL":
                assign_employee = AssignProduct.objects.filter(assign_employee__employee_id=value)
                for i in assign_employee:
                    barcode_text = i.stock_id.barcodes.first().barcode_text if i.stock_id.barcodes.exists() else " - "
                    assigned_products.append({
                        "assign_id": i.assign_id,
                        "stock_id": i.stock_id.stock_id,
                        "barcode_text": barcode_text,
                        "product_name": i.assign_product.product_name,
                        "product_category": i.assign_product.category.category_name,
                        "product_subcategory": i.assign_product.subcategory.subcategory_name,
                        "employee_name": i.assign_employee.name,
                        "mode": i.assign_mode,
                        "product_status": i.stock_id.product_status,
                        "exhibition_name": i.assign_exhibition.exhibition_name if i.assign_exhibition else " - ",
                    })
            else:
                raise ValueError("Invalid type provided")
            # Return the assigned products in the response
            return JsonResponse({
                "status": "success",
                "type": type,
                "value": value,
                "assigned_products": assigned_products,  # Include assigned products data
            })
        except json.JSONDecodeError:
            return JsonResponse({"status": "error", "message": "Invalid JSON data"}, status=400)
        except ValueError as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
    else:
        return JsonResponse({"status": "error", "message": "Invalid request method"}, status=400)
    

@csrf_exempt
def update_product_status(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid request method"}, status=400)

    try:
        data = json.loads(request.body)
        stock_id = data.get("stock_id")
        product_status = data.get("product_status")
        assign_id = data.get("assign_id")
        if not stock_id or not product_status:
            return JsonResponse({"status": "error", "message": "Missing stock_id or product_status"}, status=400)

        print("--------------", stock_id, product_status,assign_id)

        if product_status == "RETURN":
            try:
                assign_product = AssignProduct.objects.get(stock_id=stock_id)
                return_stock_data, created = ReturnProductHistory.objects.get_or_create(return_stock_id=stock_id,assign_return_id=assign_id)
                return_stock_data.return_id = uuid.uuid4()
                return_stock_data.return_quantity = assign_product.assign_quantity
                return_stock_data.return_at = timezone.now()
                return_stock_data.reuturn_status = "Success"
                return_stock_data.save()

                product = StockEntry.objects.get(stock_id=stock_id)
                product.product_status = "INSTOCK"
                product.save()

                assign_product.delete()

                return JsonResponse({"status": "success", "message": "Product status updated and moved to return history successfully!"})
            except AssignProduct.DoesNotExist:
                print(traceback.format_exc())
                return JsonResponse({"status": "error", "message": "Product not found in AssignProduct"}, status=404)
            except StockEntry.DoesNotExist:
                print(traceback.format_exc())
                return JsonResponse({"status": "error", "message": "StockEntry not found"}, status=404)
        print(traceback.format_exc())
        return JsonResponse({"status": "success", "message": "Product status updated successfully!"})

    except json.JSONDecodeError:
        return JsonResponse({"status": "error", "message": "Invalid JSON data"}, status=400)
    except Exception as e:
        print(traceback.format_exc())
        return JsonResponse({"status": "error", "message": str(e)}, status=500)

def return_stock_history(request):
    return render(request, "stock/stock_return_history.html")
from django.db.models import Q


def get_return_history(request):
    # data = ReturnProductHistory.objects.filter(return_id__isnull=False)
    data = ReturnProductHistory.objects.filter(~Q(return_id__isnull=True) & ~Q(return_id=""))
    return_data = []
    for product in data:
        employee_name = product.return_employee.name if product.return_employee else ""
        if product.return_mode == "Exhibition":
            employee_name = (
                (product.return_employee.name if product.return_employee else "") +
                " - " +
                (product.return_exhibition.exhibition_name if product.return_exhibition else "")
            )
        elif product.return_mode == "Customers":
            employee_name = product.return_customer.customer_name if product.return_customer else None,
        elif product.return_mode == "Company":
            employee_name= product.return_company.company_name if product.return_company else None
        print("==========",employee_name)
        return_data.append({
            'return_id': product.return_id,
            'assign_return_id': product.assign_return_id,
            "barcode": product.return_stock_id.barcodes.first().barcode_text if product.return_stock_id.barcodes.exists() else " - ",
            'product_name': product.return_product.product_name,
            'employee_name': employee_name,
            'return_mode': product.return_mode,
            'reuturn_status': product.reuturn_status,
            'return_assign_status': product.return_product_status,
            'assign_return_id': product.assign_return_id,
            'return_product_status': product.return_product_status,
            'reuturn_assign_at': product.reuturn_assign_at if product.reuturn_assign_at else "",
            'return_at': product.return_at.strftime('%Y-%m-%d %H:%M:%S') if product.return_at else None,
        })
    
    return JsonResponse(return_data, safe=False)
import subprocess


@csrf_exempt
def receive_barcode(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            barcode = data.get('barcode')
            print(f"Received barcode: {barcode}")
            return JsonResponse({'status': 'success', 'barcode': barcode})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)





from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import cv2
import os
from django.core.files.storage import default_storage

@csrf_exempt
def scan_qr_image(request):
    if request.method == 'POST' and request.FILES.get('image'):
        # Save the uploaded file temporarily
        file = request.FILES['image']
        file_name = default_storage.save(file.name, file)
        file_path = default_storage.path(file_name)

        # Read the image using OpenCV
        img = cv2.imread(file_path)

        if img is not None:
            detector = cv2.QRCodeDetector()
            data, bbox, _ = detector.detectAndDecode(img)

            # Clean up the uploaded file
            default_storage.delete(file_name)

            if data:
                return JsonResponse({"qr_data": data})
            else:
                return JsonResponse({"error": "No QR code detected in the image!"}, status=400)
        else:
            return JsonResponse({"error": "Error reading the image!"}, status=400)
    else:
        return JsonResponse({"error": "No image provided!"}, status=400)

def scan_via_mobile(request):
    return render(request, "scanstock/scan_via_mobile.html")


latest_scanned_data = None

@csrf_exempt
def get_backend_data_scanned_data(request):
    global latest_scanned_data
    if request.method == 'POST':
        try:
            if request.content_type == 'application/json':
                data = json.loads(request.body)
            else:
                data = request.POST
            qr_data = data.get('qr_data')
            system_ip_data = data.get('system_ip')
            get_temporary_ipv6 = data.get('get_temporary_ipv6')
            print("QR DATA : ",qr_data,"SYSTEM IP : ",system_ip_data,"get_temporary_ipv6",get_temporary_ipv6)
            if not qr_data:
                return JsonResponse({"status": "error", "message": "Missing 'qr_data' in request."}, status=400)
            if isinstance(qr_data, str):
                barcode_text = qr_data.split(":")[1].strip().split("\n")[0]  # Extract "PROD-000001-ELEC-BO-01"
            elif isinstance(qr_data, dict) and 'Barcode' in qr_data:
                barcode_text = qr_data['Barcode']
            else:
                return JsonResponse({"status": "error", "message": "Invalid 'qr_data' format."}, status=400)
            try:
                assign_product = AssignProduct.objects.get(stock_id__barcodes__barcode_text=barcode_text)
                if assign_product.assign_mode == "Users":
                    emp_name = assign_product.assign_employee.name
                elif assign_product.assign_mode == "Customer":
                    emp_name = assign_product.assign_customer.customer_name
                elif assign_product.assign_mode == "Exhibition":
                    emp_name = f"{assign_product.assign_employee.name} - {assign_product.assign_exhibition.exhibition_name}"
                elif assign_product.assign_mode == "Company":
                    emp_name = assign_product.assign_company.company_name
                else:
                    emp_name = None
                response_data = {
                    "barcode": barcode_text,
                    "stock_id": assign_product.stock_id.stock_id,
                    "product_name": assign_product.assign_product.product_name,
                    "product_category": assign_product.assign_product.category.category_name,
                    "product_subcategory": assign_product.assign_product.subcategory.subcategory_name,
                    "assign_mode": assign_product.assign_mode,
                    "assign_at": assign_product.assign_at,
                    "emp_name": emp_name,
                    "box": assign_product.stock_id.box.box_name,
                    "rack": assign_product.stock_id.rack.rank_name,
                    "product_status": assign_product.stock_id.product_status,
                    "location": assign_product.stock_id.location.name,
                    "from": "assign_product",
                    "system_ip": system_ip_data,
                }
            except AssignProduct.DoesNotExist:
                stock_entry = StockEntry.objects.get(barcodes__barcode_text=barcode_text)
                response_data = {
                    "barcode": barcode_text,
                    "stock_id": stock_entry.stock_id,
                    "product_name": stock_entry.product.product_name,
                    "product_category": stock_entry.product.category.category_name,
                    "product_subcategory": stock_entry.product.subcategory.subcategory_name,
                    "box": stock_entry.box.box_name,
                    "rack": stock_entry.rack.rank_name,
                    "product_status": stock_entry.product_status,
                    "location": stock_entry.location.name,
                    "from": "stock_entry",
                    "system_ip": system_ip_data,
                }
            html = render_to_string('scanstock/scan_barcode_data_table.html', {"data": response_data})
            latest_scanned_data = html  
            return JsonResponse({
                "status": "success",
                "message": "Data retrieved successfully",
                "data": response_data,
                "html": html 
            })
        except Exception as e:
            print("Error:", traceback.format_exc())
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
    return JsonResponse({"status": "error", "message": "Invalid request method. Only POST is allowed."}, status=405)

@csrf_exempt
def get_latest_scanned_data(request):
    global latest_scanned_data
    if latest_scanned_data:
        return JsonResponse({
            "status": "success",
            "html": latest_scanned_data,
        })
    else:
        return JsonResponse({
            "status": "error",
            "message": "No data available.",
        })

def assign_exhibition(request):
    return render(request, "exhibition_operation/assign_exhibition.html")

def return_exhibition(request):
    return render(request, "exhibition_operation/return_exhibition.html")

def assign_operation_exhibition_list(request):
    return render(request, "exhibition_operation/assign_operation_exhibition_list.html")

def fetch_exhibition_list(request):
    exb_list = Exhibition.objects.all().values(
        'exhibition_id',
        'exhibition_name',
        'location',
        'city',
        'state',
        'start_date',
        'end_date'
    )
    return JsonResponse({
        "status": "success",
        "data": list(exb_list),
    })


def generate_employee_pdf(request):   
    # Get filter parameters from request
    employee_code = request.GET.get('employee_code', '').strip()
    employee_name = request.GET.get('employee_name', '').strip()
    work_location = request.GET.get('work_location', '')

    # Filter Employee data based on provided filters
    employees = Employee.objects.all()
    
    if employee_code:
        employees = employees.filter(employee_code__icontains=employee_code)
    if employee_name:
        employees = employees.filter(name__icontains=employee_name)
    if work_location:
        employees = employees.filter(work_location_id=work_location)
        
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="employee_details.pdf"'

    # Define page size
    page_width, page_height = A4

    # Create PDF Document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,  
        rightMargin=0.7 * inch,  
        topMargin=1.5 * inch,  
        bottomMargin=1.5 * inch  
    )

    elements = []
    styles = getSampleStyleSheet()

    # Define Custom Style for Paragraphs
    custom_normal = ParagraphStyle(
        'custom_normal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        leading=14,  # Adjusted line spacing
        spaceAfter=8
    )

    # Load images
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")
    footer_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_footer.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Define Header/Footer Function
    def on_page(canvas, doc):
        width, height = A4

        # Draw header
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)  

        # Draw footer
        if os.path.exists(footer_path):
            footer_img = Image(footer_path, width=width, height=1 * inch)
            footer_img.drawOn(canvas, 0, 0)  

        # Add Report Date/Time below header (Right-aligned)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")  

    # Define Page Template with Frames
    frame = Frame(
        x1=0.7 * inch, y1=1.5 * inch,  
        width=page_width - 1.4 * inch,  
        height=page_height - 3 * inch,  
        id='normal'
    )
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # **Header Section with Increased Font Size and Margin**
    elements.append(Spacer(1, 45))  

    # Report Title (Centered, Larger Font)
    elements.append(Paragraph('<para align="center"><b><font size=16>EMPLOYEE DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))  

    # Date (Left-aligned, Proper Formatting)
    elements.append(Paragraph('<font size=12><b>Date:</b> {}</font>'.format(report_date), custom_normal))
    
    # Department (Left-aligned)
    elements.append(Paragraph('<font size=12><b>Department:</b> HR/Management</font>', custom_normal))

    # Subject & Description
    elements.append(Paragraph('<font size=12><b>Subject:</b> Employee Details Summary</font>', custom_normal))
    
    elements.append(Paragraph( 
        '<font size=11>Please find below the detailed report of employees. '
        'This report contains vital employee information, including personal details, '
        'job roles, and employment status.</font>', 
        custom_normal
    ))

    elements.append(Spacer(1, 10))  

    # **Table Title**
    elements.append(Paragraph("<strong>Employee Information Summary</strong>", styles["Heading3"]))
    elements.append(Spacer(1, 10))

    # **Table Headers**
    data = [["Code", "Name", "Designation", "Mobile", "Work Location"]]

    # Fetch Employees and Append to Table Data (Use already filtered employees)
    if employees.exists():  
        for emp in employees:
            data.append([
                emp.employee_code, emp.name, emp.designation,
                emp.mobile_number, str(emp.work_location) if emp.work_location else "N/A"  
            ])
    else:
        data.append(["No Data", "-", "-", "-", "-"])  

    # **Reduced Table Width**
    col_widths = [0.8 * inch, 1.8 * inch, 1.4 * inch, 1.3 * inch, 1.6 * inch]  
    row_height = 22  

    table = Table(data, colWidths=col_widths, rowHeights=row_height)

    # **Table Styling**
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  
        ('FONTSIZE', (0, 0), (-1, -1), 9),  
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),  
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  
        ('WORDWRAP', (0, 0), (-1, -1)),  
    ])
    table.setStyle(style)

    # Add Table to PDF
    elements.append(table)

    # **Add Note at the End**
    elements.append(Spacer(1, 15))  
    note = Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_normal)
    elements.append(note)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response


def generate_employee_csv(request): 
    # Get filter parameters from request
    employee_code = request.GET.get('employee_code', '').strip()
    employee_name = request.GET.get('employee_name', '').strip()
    work_location = request.GET.get('work_location', '')

    # Filter Employee data based on provided filters
    employees = Employee.objects.all()
    
    if employee_code:
        employees = employees.filter(employee_code__icontains=employee_code)
    if employee_name:
        employees = employees.filter(name__icontains=employee_name)
    if work_location:
        employees = employees.filter(work_location_id=work_location)

    # Create CSV Response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="employee_details.csv"'

    writer = csv.writer(response)

    # Write CSV Header with additional fields
    writer.writerow([
        "Employee ID", "Employee Code", "Name", "Designation", "Address", "Location",
        "Work Location", "Date of Birth", "Mobile Number", "Email", 
        "Aadhaar Card", "PAN Card", "Created At", "Updated At"
    ])
    
    # Write Employee Data
    for emp in employees:
        writer.writerow([
            emp.employee_id,
            emp.employee_code,
            emp.name,
            emp.designation,
            emp.address,
            emp.location,
            emp.work_location.name if emp.work_location else "N/A",
            emp.date_of_birth.strftime('%Y-%m-%d'),  # Formatting Date of Birth
            emp.mobile_number,
            emp.email,
            emp.aadhaar_card,
            emp.pan_card if emp.pan_card else "N/A",
            emp.created_at.strftime('%Y-%m-%d %H:%M:%S'),  # Formatting DateTime
            emp.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        ])

    return response


def generate_employee_excel(request):
    # Get filter parameters from request
    employee_code = request.GET.get('employee_code', '').strip()
    employee_name = request.GET.get('employee_name', '').strip()
    work_location = request.GET.get('work_location', '')

    # Filter Employee data based on provided filters
    employees = Employee.objects.all()
    
    if employee_code:
        employees = employees.filter(employee_code__icontains=employee_code)
    if employee_name:
        employees = employees.filter(name__icontains=employee_name)
    if work_location:
        employees = employees.filter(work_location_id=work_location)

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Details"

    # Define Excel headers
    headers = [
        "Employee ID", "Employee Code", "Name", "Designation", "Address", "Location",
        "Work Location", "Date of Birth", "Mobile Number", "Email", 
        "Aadhaar Card", "PAN Card", "Created At", "Updated At"
    ]

    # Add headers to the first row with bold formatting
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Add Employee Data
    for row_num, emp in enumerate(employees, 2):
        ws.cell(row=row_num, column=1, value=emp.employee_id)
        ws.cell(row=row_num, column=2, value=emp.employee_code)
        ws.cell(row=row_num, column=3, value=emp.name)
        ws.cell(row=row_num, column=4, value=emp.designation)
        ws.cell(row=row_num, column=5, value=emp.address)
        ws.cell(row=row_num, column=6, value=emp.location)
        ws.cell(row=row_num, column=7, value=emp.work_location.name if emp.work_location else "N/A")
        ws.cell(row=row_num, column=8, value=emp.date_of_birth.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=9, value=emp.mobile_number)
        ws.cell(row=row_num, column=10, value=emp.email)
        ws.cell(row=row_num, column=11, value=emp.aadhaar_card)
        ws.cell(row=row_num, column=12, value=emp.pan_card if emp.pan_card else "N/A")
        ws.cell(row=row_num, column=13, value=emp.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_num, column=14, value=emp.updated_at.strftime('%Y-%m-%d %H:%M:%S'))

    # Create HTTP response for file download
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="employee_details.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


def generate_location_csv(request):
    # Get filter parameters from request
    location_name = request.GET.get('location_name', '').strip()
    shortcode = request.GET.get('shortcode', '').strip()
    status = request.GET.get('status', '')

    # Filter Location data based on provided filters
    locations = Location.objects.all()
    
    if location_name:
        locations = locations.filter(name__icontains=location_name)
    if shortcode:
        locations = locations.filter(shortcode__icontains=shortcode)
    if status:
        locations = locations.filter(status=status)

    # Create CSV Response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="location_details.csv"'

    writer = csv.writer(response)

    # Write CSV Header
    writer.writerow(["Location ID", "Name", "Shortcode", "Details", "Status", "Created At"])

    # Write Location Data
    for loc in locations:
        writer.writerow([
            loc.id, loc.name, loc.shortcode, loc.details, loc.status,
            loc.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])

    return response


def generate_location_excel(request):
    # Get filter parameters from request
    location_name = request.GET.get('location_name', '').strip()
    shortcode = request.GET.get('shortcode', '').strip()
    status = request.GET.get('status', '')

    # Filter Location data
    locations = Location.objects.all()
    
    if location_name:
        locations = locations.filter(name__icontains=location_name)
    if shortcode:
        locations = locations.filter(shortcode__icontains=shortcode)
    if status:
        locations = locations.filter(status=status)

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Location Details"

    # Define headers
    headers = ["Location ID", "Name", "Shortcode", "Details", "Status", "Created At"]
    
    # Add headers to the first row (bold formatting)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Add Location Data
    for row_num, loc in enumerate(locations, 2):
        ws.cell(row=row_num, column=1, value=loc.id)
        ws.cell(row=row_num, column=2, value=loc.name)
        ws.cell(row=row_num, column=3, value=loc.shortcode)
        ws.cell(row=row_num, column=4, value=loc.details)
        ws.cell(row=row_num, column=5, value=loc.status)
        ws.cell(row=row_num, column=6, value=loc.created_at.strftime('%Y-%m-%d %H:%M:%S'))

    # Create HTTP response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="location_details.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


def generate_location_pdf(request):
    # Get filter parameters
    name = request.GET.get('name', '').strip()
    shortcode = request.GET.get('shortcode', '').strip()
    status = request.GET.get('status', '')

    # Filter locations based on query parameters
    locations = Location.objects.all()
    if name:
        locations = locations.filter(name__icontains=name)
    if shortcode:
        locations = locations.filter(shortcode__icontains=shortcode)
    if status:
        locations = locations.filter(status=status)

    # Create response object
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="location_details.pdf"'

    # Define PDF document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()
    custom_normal = ParagraphStyle('custom_normal', parent=styles['Normal'], fontName='Helvetica', fontSize=11, leading=14, spaceAfter=8)

    # Load images (Header/Footer)
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get current date/time
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template with Frames
    frame = Frame(x1=0.7 * inch, y1=1.5 * inch, width=A4[0] - 1.4 * inch, height=A4[1] - 3 * inch, id='normal')
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>LOCATION DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Date
    elements.append(Paragraph('<font size=11>This report contains the details of locations, including their names, shortcodes, and statues.</font>', custom_normal))
    elements.append(Spacer(1, 10))

    # Table Title
    elements.append(Paragraph("<strong>Location Information Summary</strong>", styles["Heading3"]))
    elements.append(Spacer(1, 10))

    # Table Headers
    data = [["Name", "Shortcode", "Details", "Status"]]

    # Add location data to the table
    if locations.exists():
        for location in locations:
            data.append([location.name, location.shortcode, location.details, location.get_status_display()])
    else:
        data.append(["No Data", "-", "-", "-"])

    # Define Table
    col_widths = [1.8 * inch, 1.0 * inch, 2.5 * inch, 1.2 * inch]
    row_height = 22
    table = Table(data, colWidths=col_widths, rowHeights=row_height)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),
    ])
    table.setStyle(style)

    # Add Table to PDF
    elements.append(table)

    # Footer Note
    elements.append(Spacer(1, 15))
    note = Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_normal)
    elements.append(note)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response


def generate_company_csv(request):
    # Get filter parameters
    company_name = request.GET.get('company_name', '').strip()
    location = request.GET.get('location', '').strip()

    # Filter company data
    companies = Company.objects.all()
    
    if company_name:
        companies = companies.filter(company_name__icontains=company_name)  # Fixed field name
    if location:
        companies = companies.filter(location__icontains=location)

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="company_details.csv"'

    writer = csv.writer(response)

    # Write CSV Header
    writer.writerow(["Company ID", "Company Name", "CIN Number", "GST Number", "Location", "Email", "Phone Number"])

    # Write Company Data
    for company in companies:
        writer.writerow([
            company.company_id,  # Use correct field
            company.company_name,  # Use correct field
            company.CIN_number,
            company.GST_number,
            company.location,
            company.email,
            company.phone_number
        ])

    return response

def generate_company_excel(request):
    # Get filter parameters
    company_name = request.GET.get('company_name', '').strip()
    location = request.GET.get('location', '').strip()

    # Filter company data
    companies = Company.objects.all()
    
    if company_name:
        companies = companies.filter(company_name__icontains=company_name)  # Fixed field name
    if location:
        companies = companies.filter(location__icontains=location)

    # Create an Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Company Details"

    # Define headers
    headers = ["Company ID", "Company Name", "CIN Number", "GST Number", "Location", "Email", "Phone Number"]

    # Add headers to the first row (bold formatting)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Add Company Data
    for row_num, company in enumerate(companies, 2):
        ws.cell(row=row_num, column=1, value=company.company_id)
        ws.cell(row=row_num, column=2, value=company.company_name)  # Fixed field name
        ws.cell(row=row_num, column=3, value=company.CIN_number)
        ws.cell(row=row_num, column=4, value=company.GST_number)
        ws.cell(row=row_num, column=5, value=company.location)
        ws.cell(row=row_num, column=6, value=company.email)
        ws.cell(row=row_num, column=7, value=company.phone_number)

    # Create HTTP response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="company_details.xlsx"'

    # Save workbook to response
    wb.save(response)
    
    return response


def generate_company_pdf(request):
     # Get filter parameters
    print(f"Received GET parameters: {request.GET}")

    name = request.GET.get('company_name', '').strip()
    location = request.GET.get('location', '').strip()

    print(f"Filters received - Name: {name}, Location: {location}")

    companies = Company.objects.all()
    if name:
        print(f"Filtering by Name: {name}")
        companies = companies.filter(company_name__icontains=name)
    if location:
        print(f"Filtering by Location: {location}")
        companies = companies.filter(location__icontains=location)

    print(f"Filtered Companies Count: {companies.count()}")  
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="company_details.pdf"'

    # Define PDF document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()
    custom_normal = ParagraphStyle('custom_normal', parent=styles['Normal'], fontName='Helvetica', fontSize=11, leading=14, spaceAfter=8)

    # Load images (Header/Footer)
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get current date/time
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template with Frames
    frame = Frame(x1=0.7 * inch, y1=1.5 * inch, width=A4[0] - 1.4 * inch, height=A4[1] - 3 * inch, id='normal')
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>COMPANY DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains the details of companies, including their names, locations, emails, and phone numbers.</font>', custom_normal))
    elements.append(Spacer(1, 10))

    # Table Title
    elements.append(Paragraph("<strong>Company Information Summary</strong>", styles["Heading3"]))
    elements.append(Spacer(1, 10))

    # Table Headers
    data = [["ID", "Name", "Location", "Email", "Phone Number"]]

    # Add company data to the table
    if companies.exists():
        for company in companies:
            data.append([
                str(company.company_id), 
                company.company_name, 
                company.location, 
                company.email, 
                company.phone_number
            ])
    else:
        data.append(["No Data"] + ["-"] * 4)  # Adjusting for column count

    # Define Table
    col_widths = [0.8 * inch, 1.5 * inch, 1.5 * inch, 2.0 * inch, 1.5 * inch]
    row_height = 22
    table = Table(data, colWidths=col_widths, rowHeights=row_height)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),
    ])
    table.setStyle(style)

    # Add Table to PDF
    elements.append(table)

    # Footer Note
    elements.append(Spacer(1, 15))
    note = Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_normal)
    elements.append(note)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response


def generate_category_csv(request):
    # Get filter parameters from request
    category_name = request.GET.get('category_name', '').strip()
    shortcode = request.GET.get('shortcode', '').strip()
    
    # Filter Category data based on provided filters
    categories = Category.objects.all()
    
    if category_name:
        categories = categories.filter(category_name__icontains=category_name)
    if shortcode:
        categories = categories.filter(shortcode__icontains=shortcode)
    
    # Create CSV Response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="category_details.csv"'
    
    writer = csv.writer(response)
    
    # Write CSV Header
    writer.writerow(["Category ID", "Category Name", "Details", "Shortcode", "Created On"])
    
    # Write Category Data
    for cat in categories:
        writer.writerow([
            cat.category_id, cat.category_name, cat.details, cat.shortcode,
            cat.created_on.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    return response

def generate_category_excel(request):
    # Get filter parameters from request
    category_name = request.GET.get('category_name', '').strip()
    shortcode = request.GET.get('shortcode', '').strip()
    
    # Filter Category data
    categories = Category.objects.all()
    
    if category_name:
        categories = categories.filter(category_name__icontains=category_name)
    if shortcode:
        categories = categories.filter(shortcode__icontains=shortcode)
    
    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Category Details"
    
    # Define headers
    headers = ["Category ID", "Category Name", "Details", "Shortcode", "Created On"]
    
    # Add headers to the first row (bold formatting)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Add Category Data
    for row_num, cat in enumerate(categories, 2):
        ws.cell(row=row_num, column=1, value=cat.category_id)
        ws.cell(row=row_num, column=2, value=cat.category_name)
        ws.cell(row=row_num, column=3, value=cat.details)
        ws.cell(row=row_num, column=4, value=cat.shortcode)
        ws.cell(row=row_num, column=5, value=cat.created_on.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Create HTTP response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="category_details.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response

def generate_category_pdf(request):
    # Get filter parameters
    name = request.GET.get('name', '').strip()
    shortcode = request.GET.get('shortcode', '').strip()

    # Filter categories based on query parameters
    categories = Category.objects.all()
    if name:
        categories = categories.filter(category_name__icontains=name)
    if shortcode:
        categories = categories.filter(shortcode__icontains=shortcode)

    # Create response object
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="category_details.pdf"'

    # Define PDF document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()
    custom_normal = ParagraphStyle('custom_normal', parent=styles['Normal'], fontName='Helvetica', fontSize=11, leading=14, spaceAfter=8)

    # Load header image (if available)
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get current date/time
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(x1=0.7 * inch, y1=1.5 * inch, width=A4[0] - 1.4 * inch, height=A4[1] - 3 * inch, id='normal')
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>CATEGORY DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains the details of categories, including their names and shortcodes.</font>', custom_normal))
    elements.append(Spacer(1, 10))

    # Table Title
    elements.append(Paragraph('<para align="center"><strong>Category Information Summary</strong></para>', styles["Heading3"]))
    elements.append(Spacer(1, 10))

    # Table Headers
    data = [["Category Name", "Shortcode"]]

    # Add category data to the table
    if categories.exists():
        for category in categories:
            data.append([category.category_name, category.shortcode])
    else:
        data.append(["No Data", "-", "-"])

    # Define Table
    col_widths = [2.5 * inch, 1.5 * inch, 3.0 * inch]
    row_height = 22
    table = Table(data, colWidths=col_widths, rowHeights=row_height)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),
    ])
    table.setStyle(style)

    # Add Table to PDF
    elements.append(table)

    # Footer Note
    elements.append(Spacer(1, 15))
    note = Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_normal)
    elements.append(note)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response



def generate_subcategory_csv(request):
    # Get filter parameters from request
    subcategory_name = request.GET.get('subcategory_name', '').strip()
    shortcode = request.GET.get('shortcode', '').strip()
    category_id = request.GET.get('category_id', '').strip()

    # Filter SubCategory data based on provided filters
    subcategories = SubCategory.objects.all()
    
    if subcategory_name:
        subcategories = subcategories.filter(subcategory_name__icontains=subcategory_name)
    if shortcode:
        subcategories = subcategories.filter(shortcode__icontains=shortcode)
    if category_id:
        subcategories = subcategories.filter(category_id=category_id)

    # Create CSV Response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="subcategory_details.csv"'

    writer = csv.writer(response)

    # Write CSV Header
    writer.writerow(["SubCategory ID", "SubCategory Name", "Details", "Shortcode", "Category Name", "Created On"])

    # Write SubCategory Data
    for sub in subcategories:
        writer.writerow([
            sub.subcategory_id, sub.subcategory_name, sub.details, sub.shortcode,
            sub.category.category_name, sub.created_on.strftime('%Y-%m-%d %H:%M:%S')
        ])

    return response


def generate_subcategory_excel(request):
    # Get filter parameters from request
    subcategory_name = request.GET.get('subcategory_name', '').strip()
    shortcode = request.GET.get('shortcode', '').strip()
    category_id = request.GET.get('category_id', '').strip()

    # Filter SubCategory data
    subcategories = SubCategory.objects.all()

    if subcategory_name:
        subcategories = subcategories.filter(subcategory_name__icontains=subcategory_name)
    if shortcode:
        subcategories = subcategories.filter(shortcode__icontains=shortcode)
    if category_id:
        subcategories = subcategories.filter(category_id=category_id)

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SubCategory Details"

    # Define headers
    headers = ["SubCategory ID", "SubCategory Name", "Details", "Shortcode", "Category Name", "Created On"]

    # Add headers to the first row (bold formatting)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Add SubCategory Data
    for row_num, sub in enumerate(subcategories, 2):
        ws.cell(row=row_num, column=1, value=sub.subcategory_id)
        ws.cell(row=row_num, column=2, value=sub.subcategory_name)
        ws.cell(row=row_num, column=3, value=sub.details)
        ws.cell(row=row_num, column=4, value=sub.shortcode)
        ws.cell(row=row_num, column=5, value=sub.category.category_name)
        ws.cell(row=row_num, column=6, value=sub.created_on.strftime('%Y-%m-%d %H:%M:%S'))

    # Create HTTP response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="subcategory_details.xlsx"'

    # Save workbook to response
    wb.save(response)

    return response


def generate_subcategory_pdf(request):
     # Debugging: Print received parameters
    print("Received Parameters:", request.GET)

    # Get filter parameters
    name = request.GET.get('name', '').strip()
    shortcode = request.GET.get('shortcode', '').strip()
    category_id = request.GET.get('category_id', '').strip()

    # Filter SubCategories
    subcategories = SubCategory.objects.all()

    if name:
        subcategories = subcategories.filter(subcategory_name__icontains=name)
    if shortcode:
        subcategories = subcategories.filter(shortcode__icontains=shortcode)
    if category_id:
        subcategories = subcategories.filter(category__id=category_id)

    # Debugging: Print the filtered queryset
    print("Filtered SubCategories:", subcategories.query)

    # If no data found, return an empty PDF
    if not subcategories.exists():
        return HttpResponse("No data found matching filters.", content_type="text/plain")

    # Create response object
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="subcategory_details.pdf"'

    # Define PDF document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()
    custom_normal = ParagraphStyle('custom_normal', parent=styles['Normal'], fontName='Helvetica', fontSize=11, leading=14, spaceAfter=8)

    # Load header image (if available)
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get current date/time
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(x1=0.7 * inch, y1=1.5 * inch, width=A4[0] - 1.4 * inch, height=A4[1] - 3 * inch, id='normal')
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>SUBCATEGORY DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains details of subcategories, including their names, shortcodes, and categories.</font>', custom_normal))
    elements.append(Spacer(1, 10))

    # Table Title
    elements.append(Paragraph('<para align="center"><strong>SubCategory Information Summary</strong></para>', styles["Heading3"]))
    elements.append(Spacer(1, 10))

    # Table Headers
    data = [["SubCategory Name", "Shortcode", "Category Name"]]

    # Add subcategory data to the table
    if subcategories.exists():
        for subcategory in subcategories:
            data.append([subcategory.subcategory_name, subcategory.shortcode, subcategory.category.category_name])
    else:
        data.append(["No Data", "-", "-"])

    # Define Table
    col_widths = [2.5 * inch, 1.5 * inch, 3.0 * inch]
    row_height = 22
    table = Table(data, colWidths=col_widths, rowHeights=row_height)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),
    ])
    table.setStyle(style)

    # Add Table to PDF
    elements.append(table)

    # Footer Note
    elements.append(Spacer(1, 15))
    note = Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_normal)
    elements.append(note)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response


def generate_box_csv(request):
    # Get filter parameters from request
    box_name = request.GET.get('box_name', '').strip()
    
    # Filter Box data
    boxes = Box.objects.all()
    if box_name:
        boxes = boxes.filter(box_name__icontains=box_name)
    
    # Create CSV Response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="box_details.csv"'
    
    writer = csv.writer(response)
    
    # Write CSV Header
    writer.writerow(["Box ID", "Box Name", "Details", "Created On"])
    
    # Write Box Data
    for box in boxes:
        writer.writerow([
            box.box_id, box.box_name, box.details, box.created_on.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    return response


def generate_box_excel(request):
    # Get filter parameters from request
    box_name = request.GET.get('box_name', '').strip()
    
    # Filter Box data
    boxes = Box.objects.all()
    if box_name:
        boxes = boxes.filter(box_name__icontains=box_name)
    
    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Box Details"
    
    # Define headers
    headers = ["Box ID", "Box Name", "Details", "Created On"]
    
    # Add headers to the first row (bold formatting)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Add Box Data
    for row_num, box in enumerate(boxes, 2):
        ws.cell(row=row_num, column=1, value=box.box_id)
        ws.cell(row=row_num, column=2, value=box.box_name)
        ws.cell(row=row_num, column=3, value=box.details)
        ws.cell(row=row_num, column=4, value=box.created_on.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Create HTTP response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="box_details.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response



def generate_box_pdf(request):
    # Get filter parameters
    box_id = request.GET.get('box_id', '').strip()
    box_name = request.GET.get('box_name', '').strip()

    # Filter Box records based on query parameters
    boxes = Box.objects.all()
    if box_id:
        boxes = boxes.filter(box_id__icontains=box_id)
    if box_name:
        boxes = boxes.filter(box_name__icontains=box_name)

    # Create response object
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="box_details.pdf"'

    # Define PDF document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()
    custom_normal = ParagraphStyle('custom_normal', parent=styles['Normal'], fontName='Helvetica', fontSize=11, leading=14, spaceAfter=8)

    # Load header image (if available)
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get current date/time
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(x1=0.7 * inch, y1=1.5 * inch, width=A4[0] - 1.4 * inch, height=A4[1] - 3 * inch, id='normal')
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>BOX DETAILS</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains details of boxes, including their IDs and names.</font>', custom_normal))
    elements.append(Spacer(1, 10))

    # Table Title
    elements.append(Paragraph('<para align="center"><strong>Box Information Summary</strong></para>', styles["Heading3"]))
    elements.append(Spacer(1, 10))

    # Table Headers
    data = [["Box ID", "Box Name"]]

    # Add Box data to the table
    if boxes.exists():
        for box in boxes:
            data.append([box.box_id, box.box_name])
    else:
        data.append(["No Data", "-"])

    # Define Table
    col_widths = [2.5 * inch, 3.0 * inch]
    row_height = 22
    table = Table(data, colWidths=col_widths, rowHeights=row_height)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),
    ])
    table.setStyle(style)

    # Add Table to PDF
    elements.append(table)

    # Footer Note
    elements.append(Spacer(1, 15))
    note = Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_normal)
    elements.append(note)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response



def generate_rank_csv(request):
    # Get filter parameters from request
    rank_id = request.GET.get('rank_id', '').strip()
    rank_name = request.GET.get('rank_name', '').strip()

    # Filter Rank data
    ranks = Rank.objects.all()
    if rank_id:
        ranks = ranks.filter(rank_id=int(rank_id)) 
    if rank_name:
        ranks = ranks.filter(rank_name__icontains=rank_name)

    # Create CSV Response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="rank_details.csv"'

    writer = csv.writer(response)

    # Write CSV Header
    writer.writerow(["Rank ID", "Rank Name", "Details", "Created On"])

    # Write Rank Data
    for rank in ranks:
        writer.writerow([
            rank.rank_id, rank.rank_name, rank.details, rank.created_on.strftime('%Y-%m-%d %H:%M:%S')
        ])

    return response


def generate_rank_excel(request):
    # Get filter parameters from request
    rank_id = request.GET.get('rank_id', '').strip()
    rank_name = request.GET.get('rank_name', '').strip()

    # Filter Rank data
    ranks = Rank.objects.all()
    if rank_id:
        ranks = ranks.filter(rank_id=int(rank_id)) 
    if rank_name:
        ranks = ranks.filter(rank_name__icontains=rank_name)

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rank Details"

    # Define headers
    headers = ["Rank ID", "Rank Name", "Details", "Created On"]

    # Add headers to the first row (bold formatting)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Add Rank Data
    for row_num, rank in enumerate(ranks, 2):
        ws.cell(row=row_num, column=1, value=rank.rank_id)
        ws.cell(row=row_num, column=2, value=rank.rank_name)
        ws.cell(row=row_num, column=3, value=rank.details)
        ws.cell(row=row_num, column=4, value=rank.created_on.strftime('%Y-%m-%d %H:%M:%S'))

    # Create HTTP response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="rank_details.xlsx"'

    # Save workbook to response
    wb.save(response)

    return response


def generate_rank_pdf(request):
    # Get filter parameters from request
    rank_id = request.GET.get('rank_id', '').strip()
    rank_name = request.GET.get('rank_name', '').strip()

    # Filter Rank data
    ranks = Rank.objects.all()

    if rank_id:
        ranks = ranks.filter(rank_id=rank_id)

    if rank_name:
        ranks = ranks.filter(rank_name__icontains=rank_name)

    # Debugging: Print filtered results
    print("Filtered Ranks:", ranks)

    # Create response object
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="rank_details.pdf"'

    # Define PDF document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    custom_style = ParagraphStyle('custom_style', parent=styles['Normal'], fontName='Helvetica', fontSize=11, leading=14, spaceAfter=8)

    # Header Image (Optional)
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get current date/time
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Report Title
    elements.append(Spacer(1, 50))
    elements.append(Paragraph('<para align="center"><b><font size=16>RANK DETAILS</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains details of ranks, including their IDs and names.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Table Headers
    data = [["Rank ID", "Rank Name", "Details", "Created On"]]

    # Add Rank data to the table
    for rank in ranks:
        data.append([
            rank.rank_id, rank.rank_name, rank.details, rank.created_on.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # Define Table
    col_widths = [1.5 * inch, 2.0 * inch, 2.5 * inch, 2.0 * inch]
    table = Table(data, colWidths=col_widths)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),
    ])
    table.setStyle(style)

    # Add Table to PDF
    elements.append(table)

    # Footer Note
    elements.append(Spacer(1, 15))
    note = Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_style)
    elements.append(note)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response



# Generate Customer CSV
def generate_customer_csv(request):
    customer_name = request.GET.get('customer_name', '').strip()
    company_id = request.GET.get('company_name', '')

    customers = Customer.objects.all()
    
    if customer_name:
        customers = customers.filter(customer_name__icontains=customer_name)
    if company_id:
        customers = customers.filter(company_id=company_id)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="customer_details.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Customer ID", "Customer Name", "Email", "Company", "Phone Number",
        "Address", "City", "State", "ZIP Code", "Country", "Details"
    ])

    for customer in customers:
        writer.writerow([
            customer.customer_id, customer.customer_name, customer.email,
            customer.company.company_name if customer.company else "N/A",
            customer.phone_number, customer.address, customer.city,
            customer.state, customer.zip_code, customer.country,
            customer.details
        ])

    return response

# Generate Customer Excel
def generate_customer_excel(request):
    customer_name = request.GET.get('customer_name', '').strip()
    company_id = request.GET.get('company_name', '')

    customers = Customer.objects.all()
    
    if customer_name:
        customers = customers.filter(customer_name__icontains=customer_name)
    if company_id:
        customers = customers.filter(company_id=company_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Details"

    headers = [
        "Customer ID", "Customer Name", "Email", "Company", "Phone Number",
        "Address", "City", "State", "ZIP Code", "Country", "Details"
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header).font = openpyxl.styles.Font(bold=True)

    for row_num, customer in enumerate(customers, 2):
        ws.cell(row=row_num, column=1, value=customer.customer_id)
        ws.cell(row=row_num, column=2, value=customer.customer_name)
        ws.cell(row=row_num, column=3, value=customer.email)
        ws.cell(row=row_num, column=4, value=customer.company.company_name if customer.company else "N/A")
        ws.cell(row=row_num, column=5, value=customer.phone_number)
        ws.cell(row=row_num, column=6, value=customer.address)
        ws.cell(row=row_num, column=7, value=customer.city)
        ws.cell(row=row_num, column=8, value=customer.state)
        ws.cell(row=row_num, column=9, value=customer.zip_code)
        ws.cell(row=row_num, column=10, value=customer.country)
        ws.cell(row=row_num, column=11, value=customer.details)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="customer_details.xlsx"'

    wb.save(response)
    return response

def generate_customer_pdf(request):
    customer_name = request.GET.get('customer_name', '').strip()
    company_name = request.GET.get('company_name', '').strip()

    print(f"Received company_name: {company_name}")  # Debugging

    # Get all customers
    customers = Customer.objects.all()

    if customer_name:
        customers = customers.filter(customer_name__icontains=customer_name)

    if company_name:
        if company_name.isdigit():  # If input is an ID
            customers = customers.filter(company__company_id=int(company_name))
        else:
            customers = customers.filter(company__company_name__icontains=company_name)

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="customer_details.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    custom_style = ParagraphStyle('custom_style', parent=styles['Normal'], fontName='Helvetica', fontSize=11, leading=14, spaceAfter=8)

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Report Title
    elements.append(Spacer(1, 80))
    elements.append(Paragraph('<para align="center"><b><font size=16>CUSTOMER DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains details of customers, including their names, emails, and companies.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Table Headers
    data = [["Customer Name", "Email", "Company", "Phone Number"]]
    
    # Fetch Customers and Append to Table Data
    if customers.exists():
        for customer in customers:
            data.append([customer.customer_name, customer.email, customer.company.company_name, customer.phone_number])
    else:
        data.append(["No Data", "-", "-", "-"])  

    # Table Column Widths
    col_widths = [2 * inch, 2.5 * inch, 2 * inch, 1.5 * inch]  # Approx 72 points per inch
    table = Table(data, colWidths=col_widths)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)

    elements.append(table)
    elements.append(Spacer(1, 15))

    # Footer Note
    elements.append(Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_style))
    
    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response



# Generate Supplier CSV
def generate_supplier_csv(request):
    supplier_name = request.GET.get('supplier_name', '').strip()
    company_id = request.GET.get('company_name', '')

    suppliers = Supplier.objects.all()
    
    if supplier_name:
        suppliers = suppliers.filter(supplier_name__icontains=supplier_name)
    if company_id:
        suppliers = suppliers.filter(company_id=company_id)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="supplier_details.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Supplier ID", "Supplier Name", "Email", "Company", "Phone Number",
        "Address", "City", "State", "ZIP Code", "Country", "Details"
    ])

    for supplier in suppliers:
        writer.writerow([
            supplier.supplier_id, supplier.supplier_name, supplier.email,
            supplier.company.company_name if supplier.company else "N/A",
            supplier.phone_number, supplier.address, supplier.city,
            supplier.state, supplier.zip_code, supplier.country,
            supplier.details
        ])

    return response


# Generate Supplier Excel
def generate_supplier_excel(request):
    supplier_name = request.GET.get('supplier_name', '').strip()
    company_id = request.GET.get('company_name', '')

    suppliers = Supplier.objects.all()
    
    if supplier_name:
        suppliers = suppliers.filter(supplier_name__icontains=supplier_name)
    if company_id:
        suppliers = suppliers.filter(company_id=company_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supplier Details"

    headers = [
        "Supplier ID", "Supplier Name", "Email", "Company", "Phone Number",
        "Address", "City", "State", "ZIP Code", "Country", "Details"
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header).font = openpyxl.styles.Font(bold=True)

    for row_num, supplier in enumerate(suppliers, 2):
        ws.cell(row=row_num, column=1, value=supplier.supplier_id)
        ws.cell(row=row_num, column=2, value=supplier.supplier_name)
        ws.cell(row=row_num, column=3, value=supplier.email)
        ws.cell(row=row_num, column=4, value=supplier.company.company_name if supplier.company else "N/A")
        ws.cell(row=row_num, column=5, value=supplier.phone_number)
        ws.cell(row=row_num, column=6, value=supplier.address)
        ws.cell(row=row_num, column=7, value=supplier.city)
        ws.cell(row=row_num, column=8, value=supplier.state)
        ws.cell(row=row_num, column=9, value=supplier.zip_code)
        ws.cell(row=row_num, column=10, value=supplier.country)
        ws.cell(row=row_num, column=11, value=supplier.details)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="supplier_details.xlsx"'

    wb.save(response)
    return response


def generate_supplier_pdf(request):
    supplier_name = request.GET.get('supplier_name', '').strip()
    company_name = request.GET.get('company_name', '').strip()

    print(f"Received company_name: {company_name}")  # Debugging

    # Get all suppliers
    suppliers = Supplier.objects.all()

    if supplier_name:
        suppliers = suppliers.filter(supplier_name__icontains=supplier_name)

    if company_name:
        if company_name.isdigit():  # If input is an ID
            suppliers = suppliers.filter(company__company_id=int(company_name))
        else:
            suppliers = suppliers.filter(company__company_name__icontains=company_name)

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="supplier_details.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    custom_style = ParagraphStyle('custom_style', parent=styles['Normal'], fontName='Helvetica', fontSize=11, leading=14, spaceAfter=8)

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Report Title
    elements.append(Spacer(1, 80))
    elements.append(Paragraph('<para align="center"><b><font size=16>SUPPLIER DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains details of suppliers, including their names, emails, and companies.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Table Headers
    data = [["Supplier Name", "Email", "Company", "Phone Number"]]
    
    # Fetch Suppliers and Append to Table Data
    if suppliers.exists():
        for supplier in suppliers:
            data.append([supplier.supplier_name, supplier.email, supplier.company.company_name, supplier.phone_number])
    else:
        data.append(["No Data", "-", "-", "-"])  

    # Table Column Widths
    col_widths = [2 * inch, 2.5 * inch, 2 * inch, 1.5 * inch]  # Approx 72 points per inch
    table = Table(data, colWidths=col_widths)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)

    elements.append(table)
    elements.append(Spacer(1, 15))

    # Footer Note
    elements.append(Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_style))
    
    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response


def generate_exhibition_csv(request):
    exhibition_name = request.GET.get('exhibition_name', '').strip()
    location = request.GET.get('location', '').strip()

    exhibitions = Exhibition.objects.all()

    if exhibition_name:
        exhibitions = exhibitions.filter(exhibition_name__icontains=exhibition_name)
    if location:
        exhibitions = exhibitions.filter(location__icontains=location)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="exhibition_details.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Exhibition ID", "Exhibition Name", "Location", "Address", "City",
        "State", "Pincode", "Start Date", "End Date", "Details"
    ])

    for exhibition in exhibitions:
        writer.writerow([
            exhibition.exhibition_id, exhibition.exhibition_name, exhibition.location,
            exhibition.address, exhibition.city, exhibition.state, exhibition.pincode,
            exhibition.start_date.strftime('%Y-%m-%d'), exhibition.end_date.strftime('%Y-%m-%d'),
            exhibition.details
        ])

    return response


def generate_exhibition_excel(request):
    exhibition_name = request.GET.get('exhibition_name', '').strip()
    location = request.GET.get('location', '').strip()

    exhibitions = Exhibition.objects.all()

    if exhibition_name:
        exhibitions = exhibitions.filter(exhibition_name__icontains=exhibition_name)
    if location:
        exhibitions = exhibitions.filter(location__icontains=location)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exhibition Details"

    headers = [
        "Exhibition ID", "Exhibition Name", "Location", "Address", "City",
        "State", "Pincode", "Start Date", "End Date", "Details"
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header).font = openpyxl.styles.Font(bold=True)

    for row_num, exhibition in enumerate(exhibitions, 2):
        ws.cell(row=row_num, column=1, value=exhibition.exhibition_id)
        ws.cell(row=row_num, column=2, value=exhibition.exhibition_name)
        ws.cell(row=row_num, column=3, value=exhibition.location)
        ws.cell(row=row_num, column=4, value=exhibition.address)
        ws.cell(row=row_num, column=5, value=exhibition.city)
        ws.cell(row=row_num, column=6, value=exhibition.state)
        ws.cell(row=row_num, column=7, value=exhibition.pincode)
        ws.cell(row=row_num, column=8, value=exhibition.start_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=9, value=exhibition.end_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=10, value=exhibition.details)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="exhibition_details.xlsx"'

    wb.save(response)
    return response



def generate_exhibition_pdf(request): 
    exhibition_name = request.GET.get('exhibition_name', '').strip()
    location = request.GET.get('location', '').strip()

    # Fetch exhibitions from the database
    exhibitions = Exhibition.objects.all()

    if exhibition_name:
        exhibitions = exhibitions.filter(exhibition_name__icontains=exhibition_name)
    
    if location:
        exhibitions = exhibitions.filter(location__icontains=location)

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="exhibition_details.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    custom_style = ParagraphStyle('custom_style', parent=styles['Normal'], fontName='Helvetica', fontSize=11, leading=14, spaceAfter=8)

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Report Title
    elements.append(Spacer(1, 80))
    elements.append(Paragraph('<para align="center"><b><font size=16>EXHIBITION DETAILS</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains details of exhibitions, including their names, cities, states, locations, start, and end dates.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Table Headers (Added Start Date and End Date)
    data = [["Exhibition Name", "City", "State", "Location", "Start Date", "End Date"]]
    
    # Fetch Exhibitions and Append to Table Data
    if exhibitions.exists():
        for exhibition in exhibitions:
            data.append([
                exhibition.exhibition_name, 
                exhibition.city, 
                exhibition.state, 
                exhibition.location,
                exhibition.start_date.strftime("%Y-%m-%d"),  # Convert to string format
                exhibition.end_date.strftime("%Y-%m-%d")  # Convert to string format
            ])
    else:
        data.append(["No Data", "-", "-", "-", "-", "-"])  

    # Adjusted Table Column Widths
    col_widths = [1.9 * inch, 1.3 * inch, 1.2 * inch, 1.5 * inch, 1.0 * inch, 1.0 * inch]  
    table = Table(data, colWidths=col_widths)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)

    elements.append(table)
    elements.append(Spacer(1, 15))

    # Footer Note
    elements.append(Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_style))
    
    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response  


def generate_product_csv(request):  
    # Get filter parameters
    name = request.GET.get('name', '')
    pcode = request.GET.get('pcode', '')
    category = request.GET.get('category', '')
    subcategory = request.GET.get('subcategory', '')

    # Apply filters
    products = Product.objects.all()
    if name:
        products = products.filter(product_name__icontains=name)  
    if pcode:
        products = products.filter(product_code__icontains=pcode) 
    if category and category.isdigit():
        products = products.filter(category=category)
    if subcategory and subcategory.isdigit():
        products = products.filter(subcategory=subcategory)

    # Create CSV Response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="product_details.csv"'

    writer = csv.writer(response)

    # Write CSV Header
    writer.writerow([
        "Product ID", "Product Code", "Product Name", "Category", "Subcategory", 
        "Date of Entry", "Size (LxBxH)", "Weight", "Manufacturer", "Description", 
        "Vendor"
    ])

    # Write Product Data or Show Blank Row If No Data
    if products.exists():
        for product in products:
            writer.writerow([
                product.id,  
                product.product_code,
                product.product_name,
                product.category.category_name if product.category else "",
                product.subcategory.subcategory_name if product.subcategory else "",
                product.date_of_entry.strftime('%Y-%m-%d') if product.date_of_entry else "",
                f"{product.product_size_length} x {product.product_size_breadth} x {product.product_size_height}",
                product.product_weight,
                product.manufacture_name,
                product.description,
                product.vendor
            ])
    else:
        writer.writerow([""] * 11)  # ✅ **Blank row with 11 empty columns**

    return response


def generate_product_excel(request):
    # Get filter parameters
    name = request.GET.get('name', '')
    pcode = request.GET.get('pcode', '')
    category = request.GET.get('category', '')
    subcategory = request.GET.get('subcategory', '')

    # Apply filters
    products = Product.objects.all()
    products = Product.objects.all()
    if name:
        products = products.filter(product_name__icontains=name)  
    if pcode:
        products = products.filter(product_code__icontains=pcode) 
    if category and category.isdigit():
        products = products.filter(category=category)
    if subcategory and subcategory.isdigit():
        products = products.filter(subcategory=subcategory)


    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Details"

    # Define Excel headers (Removed Barcode, Purchase Date, Purchase Amount)
    headers = [
        "Product ID", "Product Code", "Product Name", "Category", "Subcategory",
        "Date of Entry", "Size (LxBxH)", "Weight", "Manufacturer", "Description",
        "Vendor"
    ]

    # Add headers to the first row with bold formatting
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header).font = openpyxl.styles.Font(bold=True)

    # Add Product Data
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=product.product_id)
        ws.cell(row=row_num, column=2, value=product.product_code)
        ws.cell(row=row_num, column=3, value=product.product_name)
        ws.cell(row=row_num, column=4, value=product.category.category_name if product.category else "N/A")
        ws.cell(row=row_num, column=5, value=product.subcategory.subcategory_name if product.subcategory else "N/A")
        ws.cell(row=row_num, column=6, value=product.date_of_entry.strftime('%Y-%m-%d') if product.date_of_entry else "N/A")
        ws.cell(row=row_num, column=7, value=f"{product.product_size_length} x {product.product_size_breadth} x {product.product_size_height}")
        ws.cell(row=row_num, column=8, value=product.product_weight)
        ws.cell(row=row_num, column=9, value=product.manufacture_name)
        ws.cell(row=row_num, column=10, value=product.description)
        ws.cell(row=row_num, column=11, value=product.vendor)

    # Create HTTP response for file download
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="product_details.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response 


def generate_product_pdf(request):
    # Get filter parameters from request
   # Get filter parameters
    name = request.GET.get('name', '')
    pcode = request.GET.get('pcode', '')
    category = request.GET.get('category', '')
    subcategory = request.GET.get('subcategory', '')

    # Apply filters
    products = Product.objects.all()
    if name:
        products = products.filter(product_name__icontains=name)  
    if pcode:
        products = products.filter(product_code__icontains=pcode) 
    if category and category.isdigit():
        products = products.filter(category=category)
    if subcategory and subcategory.isdigit():
        products = products.filter(subcategory=subcategory)



    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="product_details.pdf"'

    # Create PDF Document
    doc = SimpleDocTemplate(
        response, pagesize=A4,
        leftMargin=0.7 * inch, rightMargin=0.7 * inch,
        topMargin=1.5 * inch, bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()
    custom_normal = ParagraphStyle('custom_normal', parent=styles['Normal'], fontName='Helvetica', fontSize=11, leading=14, spaceAfter=8)

    # Load Header Image
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(x1=0.7 * inch, y1=1.5 * inch, width=A4[0] - 1.4 * inch, height=A4[1] - 3 * inch, id='normal')
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>PRODUCT DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains the details of products, including their codes, names, categories, and images.</font>', custom_normal))
    elements.append(Spacer(1, 10))

    # Table Title
    elements.append(Paragraph("<strong>Product Information Summary</strong>", styles["Heading3"]))
    elements.append(Spacer(1, 10))

    # Table Headers
    data = [["Product Code", "Product Image", "Product Name", "Category", "Subcategory"]]

    # Fetch Products and Append to Table Data
    if products.exists():
        for prod in products:
            product_image = "No Image"
            if prod.product_images.exists():
                img_path = prod.product_images.first().image.path
                if os.path.exists(img_path):
                    product_image = Image(img_path, width=50, height=50)
            
            data.append([
                prod.product_code,
                product_image,
                prod.product_name,
                prod.category.category_name if prod.category else "N/A",
                prod.subcategory.subcategory_name if prod.subcategory else "N/A",
            ])
    else:
        data.append(["No Data", "-", "-", "-", "-"])  

    # Define Table
    col_widths = [1.3 * inch, 1.3 * inch, 2.4 * inch, 1.3 * inch, 1.3 * inch]
    table = Table(data, colWidths=col_widths)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), 
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)

    # Add Table to PDF
    elements.append(table)
    elements.append(Spacer(1, 15))

    # Footer Note
    note = Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_normal)
    elements.append(note)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=lambda canvas, doc: None)

    return response


def print_barcode_page(request):
    return render(request, "printbarcode/print_barcode.html")



def generate_barcode_pdf(request):
    # Extract filter parameters
    category_id = request.GET.get("category", None)
    subcategory_id = request.GET.get("subcategory", None)
    product_id = request.GET.get("product", None)
    location_id = request.GET.get("location", None)
    from_date = request.GET.get("from_date", None)
    to_date = request.GET.get("to_date", None)

    # Fetch stock entries and apply filters if provided
    stock_entries = StockEntry.objects.prefetch_related("barcodes").all()

    if category_id:
        stock_entries = stock_entries.filter(category_id=category_id)
    if subcategory_id:
        stock_entries = stock_entries.filter(subcategory_id=subcategory_id)
    if product_id:
        stock_entries = stock_entries.filter(product_id=product_id)
    if location_id:
        stock_entries = stock_entries.filter(location_id=location_id)
    if from_date:
        from_date = parse_date(from_date)
        stock_entries = stock_entries.filter(created_on__gte=from_date)
    if to_date:
        to_date = parse_date(to_date)
        stock_entries = stock_entries.filter(created_on__lte=to_date)

    # Get page size and grid parameters
    page_size_param = request.GET.get("page_size", "A4")  
    grid_param = request.GET.get("grid", "3x8")  

    # Define available page sizes and grids
    page_size_map = {"A4": A4, "A3": A3, "A5": A5}
    grid_map = {
        "3x8": (3, 8), "4x10": (4, 10), "5x12": (5, 12), 
        "6x15": (6, 15), "2x5": (2, 5), "3x7": (3, 7)
    }

    # Get selected page size and grid dimensions
    page_size = page_size_map.get(page_size_param, A4)
    cols, rows = grid_map.get(grid_param, (3, 8))

    # PDF settings
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="barcodes_{page_size_param}_{grid_param}.pdf"'

    page_width, page_height = page_size
    margin = 10  
    barcode_width = (page_width - 2 * margin) / cols
    barcode_height = (page_height - 2 * margin) / rows

    pdf = canvas.Canvas(response, pagesize=page_size)
    
    x_offset = margin
    y_offset = page_height - margin - barcode_height  

    count = 0  
    for stock in stock_entries:
        for barcode in stock.barcodes.all():
            if barcode.barcode_image:
                img = ImageReader(barcode.barcode_image.path)
                
                # Draw box border
                pdf.rect(x_offset, y_offset, barcode_width, barcode_height, stroke=1, fill=0)
                
                # Draw barcode image inside the box with padding
                padding = 5
                pdf.drawImage(
                    img, 
                    x_offset + padding, 
                    y_offset + padding, 
                    width=barcode_width - 2 * padding, 
                    height=barcode_height - 2 * padding, 
                    preserveAspectRatio=True
                )

                count += 1
                if count % cols == 0:  
                    x_offset = margin
                    y_offset -= barcode_height
                else:
                    x_offset += barcode_width

                if count % (cols * rows) == 0:  
                    pdf.showPage()
                    x_offset = margin
                    y_offset = page_height - margin - barcode_height

    pdf.save()
    return response


import datetime


def generate_barcode_details_pdf(request):
    # Get filter parameters
    category_id = request.GET.get("category")
    subcategory_id = request.GET.get("subcategory")
    product_id = request.GET.get("product")
    location_id = request.GET.get("location")
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    # Filter stock data
    stocks = StockEntry.objects.all()
    if category_id:
        stocks = stocks.filter(category_id=category_id)
    if subcategory_id:
        stocks = stocks.filter(subcategory_id=subcategory_id)
    if product_id:
        stocks = stocks.filter(product_id=product_id)
    if location_id:
        stocks = stocks.filter(location_id=location_id)
    if from_date and to_date:
        stocks = stocks.filter(created_on__range=[from_date, to_date])

    # Create PDF response
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=barcode_details.pdf"

    # Define page size
    page_size = A4  # Change to A3, A5 dynamically if needed
    doc = SimpleDocTemplate(response, pagesize=page_size)
    elements = []

    # Load styles
    styles = getSampleStyleSheet()
    styleN = styles["Normal"]
    styleN.wordWrap = "CJK"

    # Load header image (Only for first page)
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def on_first_page(canvas, doc):
        width, height = page_size
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    def on_later_pages(canvas, doc):
        # No header image, just continue content
        pass
    
    
    # **Add space above the title**
    elements.append(Spacer(1, 70))  # Adjust space as needed

    # **Heading (Centered)**
    heading_style = ParagraphStyle(
        "HeadingStyle",
        fontSize=16,  # Reduce the font size
        fontName="Helvetica-Bold",
        alignment=TA_CENTER,
        spaceAfter=20,
    )

    elements.append(Paragraph("Product Label & Barcode Details", heading_style))


    # **Table Headers**
    data = [["Barcode Image", "Barcode Label", "Product", "Category", "Subcategory", "Rack", "Box", "Location","SN & MFG No"]]

    # **Barcode Data**
    for stock in stocks:
        barcode = stock.barcodes.first() if hasattr(stock, "barcodes") and stock.barcodes.exists() else None
        barcode_img = Image(barcode.barcode_image.path, width=80, height=40) if barcode and barcode.barcode_image else "N/A"

        row = [
            barcode_img, 
            Paragraph(barcode.barcode_text if barcode else "N/A", styleN),
            Paragraph(getattr(stock.product, "product_name", "N/A"), styleN),
            Paragraph(getattr(stock.category, "category_name", "N/A"), styleN),
            Paragraph(getattr(stock.subcategory, "subcategory_name", "N/A"), styleN),
            Paragraph(getattr(stock.rack, "rank_name", "N/A"), styleN),
            Paragraph(getattr(stock.box, "box_name", "N/A"), styleN),
            Paragraph(getattr(stock.location, "name", "N/A"), styleN),
            Paragraph(stock.serial_number if stock.serial_number else "N/A", styleN),
        ]
        data.append(row)

    # **3x8 Grid Layout for A4 (Can be modified dynamically)**
    table = Table(data, colWidths=[90, 80, 80, 60, 60, 45, 45, 60, 60], repeatRows=1)
    # Apply styling
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),  # Header background color
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),  # Header text color
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),  # Align all text to center
        ("GRID", (0, 0), (-1, -1), 1, colors.black),  # Table border
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),  # Vertical alignment middle
        ("FONTSIZE", (0, 0), (-1, 0), 8),  # Increase font size for headers
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),  # Bold headers
        ("ROWHEIGHT", (0, 0), (-1, 0), 50),  # Increase header row height
        ("FONTSIZE", (0, 1), (-1, -1), 6),  # Normal font size for data rows
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),  # Alternating row colors
    ]))

    elements.append(table)
    
     # **Add space above the table**
    elements.append(Spacer(1, 20))  # Adjust space as needed
    
  # **Footer Note**
    custom_style = styles["Normal"]
    custom_style.fontSize = 10  # Increase font size for better readability
    custom_style.alignment = 0  # Align to the left (use 1 for center, 2 for right)

    note = Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_style)
    elements.append(note)
    
    doc.build(elements, onFirstPage=on_first_page, onLaterPages=on_later_pages)

    return response


def create_user_page(request):
    return render(request, "users/add_user.html")

def user_list_page(request):
    return render(request, "users/list_user.html")


def add_user(request):
    if request.method == 'POST':
        try:
            print("POST Data:", request.POST)  # Debugging step

            # Extract data from request.POST instead of request.body
            username = request.POST.get('username')
            password = request.POST.get('password')
            email = request.POST.get('email')
            name = request.POST.get('name')
            mobile_no = request.POST.get('mobile_no')
            company = request.POST.get('company')
            location = request.POST.get('location')
            role_id = request.POST.get('role')

            errors = {}

            # Email validation
            email_validator = EmailValidator()
            try:
                email_validator(email)
            except ValidationError:
                errors['email'] = 'Enter a valid email address.'

            # Phone number validation (10-digit format)
            phone_validator = RegexValidator(r'^\d{10}$', 'Enter a valid 10-digit phone number.')
            try:
                phone_validator(mobile_no)
            except ValidationError:
                errors['mobile_no'] = 'Mobile number must be a 10-digit number.'

            # Length validation for fields
            length_validator = [MinLengthValidator(5), MaxLengthValidator(30)]
            for field, value in [('name', name), ('username', username), ('password', password), ('company', company), ('location', location)]:
                for validator in length_validator:
                    try:
                        validator(value)
                    except ValidationError:
                        errors[field] = f'{field.capitalize()} must be between 5 and 30 characters long.'

            # Check if role is provided
            if not role_id:
                errors['role'] = 'Role is required.'

            if errors:
                return JsonResponse({'success': False, 'errors': errors})

            # Create user
            uobj = User(username=username, email=email, password=make_password(password))
            uobj.save()

            # Create or update profile
            profile_obj = Profile(user=uobj, name=name, mobile_no=mobile_no, company=company, location=location)

            if role_id:
                try:
                    user_role = Role.objects.get(id=role_id)
                    profile_obj.role = user_role
                except Role.DoesNotExist:
                    return JsonResponse({'success': False, 'errors': {'role': 'Invalid role selected.'}})

            profile_obj.save()

            # Create UserRole instance
            UserRole.objects.create(user=profile_obj.user, role=profile_obj.role)

            # Log the action
            UserAuditLog.objects.create(
                user=request.user,
                action='Create User',
                details=f"User {uobj.username} created.",
                username=uobj.username,
                last_login=uobj.last_login
            )

            # Send email notification
            connection = CustomEmailBackend(use_secondary=False)
            connection.open()

            subject = 'Welcome to TTSPL IMS!'
            html_message = f"""
            <!DOCTYPE html>
            <html>
            <body>
                <p>Dear {name},<br><br>
                Your account has been successfully created on TTSPL IMS! 🎉</p>
                <p><strong>Account Details:</strong><br>
                - Username: {username}<br>
                - Email Address: {email}</p>
                <p>Thank you for joining us!</p>
            </body>
            </html>
            """

            try:
                send_mail(subject, '', 'noreply@example.com', [email], connection=connection, html_message=html_message)
                connection.close()
            except Exception as e:
                return JsonResponse({'success': False, 'errors': {'email': f'Error sending email: {str(e)}'}})

            return JsonResponse({'success': True, 'message': 'User created successfully!'})

        except IntegrityError as e:
            if 'unique constraint' in str(e):
                return JsonResponse({'success': False, 'errors': {'email': 'Email already exists.'}})

        except Exception as e:
            return JsonResponse({'success': False, 'errors': {'error': str(e)}})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})





def create_rol_permission_page(request):
    return render(request, "rol_permission/add_rol_permission.html")

def rol_permission_list_page(request):
    return render(request, "rol_permission/list_rol_permission.html")

def add_role(request):
    if request.method == 'POST':
        try:
            form = RoleForm(request.POST, created_by=request.user)
            if form.is_valid():
                role = form.save(commit=True)
                created_permissions = form.created_permissions

                # Log the creation of the role and associated permissions
                RolePermissionAuditLog.objects.create(
                    user=request.user,
                    action='Role created',
                    details=f'Role "{role}" created with permissions: {", ".join(created_permissions)}',
                    role=str(role),
                    permission=', '.join(created_permissions),
                )

                return JsonResponse({'success': True, 'message': 'Role added successfully!'})

            return JsonResponse({'success': False, 'errors': form.errors})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def get_roles(request):
    # Query all Role objects and order by id in ascending order
    roles = Role.objects.order_by('id').all()

    # Process roles as needed (convert to JSON, render in template, etc.)
    role_list = [{'id': role.id, 'role_name': role.role_name}
                 for role in roles]

    # Example: Return as JSON response
    return JsonResponse({'roles': role_list})




def others_settings_page(request):
    return render(request, "settings/general_settings.html")


@csrf_exempt  # Disable CSRF for simplicity, but use carefully!
def save_secondary_email_config(request):
    if request.method == 'POST':
        # Get values from the request
        otp_host = request.POST.get('otp_host')
        otp_port = request.POST.get('otp_port')
        otp_use_tls = request.POST.get(
            'otp_use_tls') == 'true'  # Convert to boolean
        otp_user = request.POST.get('otp_user')
        otp_password = request.POST.get('otp_password')
        otp_from_email = request.POST.get('otp_from_email')

        # Save or update the existing configuration
        secondary_email_config, created = SecondaryEmailConfig.objects.update_or_create(
            id=1,  # Assuming you only have one configuration
            defaults={
                'host': otp_host,
                'port': otp_port,
                'use_tls': otp_use_tls,
                'host_user': otp_user,
                'host_password': otp_password,
                'default_from_email': otp_from_email,
            }
        )

        return JsonResponse({'message': 'Secondary email configuration saved successfully.'}, status=200)

    elif request.method == 'GET':
        # Retrieve the existing configuration, if any
        try:
            config = SecondaryEmailConfig.objects.get(id=1)
            data = {
                'otp_host': config.host,
                'otp_port': config.port,
                'otp_use_tls': config.use_tls,
                'otp_user': config.host_user,
                'otp_password': config.host_password,
                'otp_from_email': config.default_from_email,
            }
            return JsonResponse(data, status=200)
        except SecondaryEmailConfig.DoesNotExist:
            return JsonResponse({'error': 'Configuration not found.'}, status=404)

    return JsonResponse({'error': 'Invalid request method.'}, status=400)




def fetch_role_permission_datatable_data(request): 
    role_id = request.GET.get('role_id', '')
    role_name = request.GET.get('role_name', '')  
    status = request.GET.get('status', '')

    roles = Role.objects.all()

    if role_id:
        roles = roles.filter(id=role_id) 
    if role_name and role_name != "All Roles":
        roles = roles.filter(role_name__icontains=role_name) 

    if status:
        roles = roles.filter(is_active=(status == 'Active'))

    role_table = [
        {
            'id': role.id,
            'role_name': role.role_name,
            'role_description': role.role_description,
            'is_active': role.is_active,
            'permissions': list(role.defaultrolepermission_set.values_list('permission__permission_name', flat=True))
        }
        for role in roles
    ]

    return JsonResponse({'role_table': role_table}, status=200)




def get_role_details(request, role_id):
    role = get_object_or_404(Role, id=role_id)
    permissions = Permission.objects.all().values('id', 'permission_name')
    assigned_permissions = role.defaultrolepermission_set.values_list('permission_id', flat=True)

    # Get Default Permissions
    default_permissions = DefaultRolePermission.objects.filter(role=role).values_list('permission_id', flat=True)

    data = {
        'id': role.id,
        'role_name': role.role_name,
        'role_description': role.role_description,
        'is_active': role.is_active,
        'permissions': list(permissions),
        'assigned_permissions': list(assigned_permissions),
        'default_permissions': list(default_permissions),
    }
    return JsonResponse(data)


def get_role_update_details(request, role_id):
    role = get_object_or_404(Role, id=role_id)
    permissions = Permission.objects.all().values('id', 'permission_name')
    assigned_permissions = role.defaultrolepermission_set.values_list(
        'permission_id', flat=True)

    data = {
        'id': role.id,
        'role_name': role.role_name,
        'role_description': role.role_description,
        'is_active': role.is_active,
        'permissions': list(permissions),
        'assigned_permissions': list(assigned_permissions),
    }
    return JsonResponse(data)


def update_role(request, role_id):
    # Retrieve role object based on role_id
    try:
        role = Role.objects.get(id=role_id)
    except Role.DoesNotExist:
        return JsonResponse({'error': 'Role not found.'}, status=404)

    # Retrieve current permissions for logging
    current_permissions = role.defaultrolepermission_set.all(
    ).values_list('permission__permission_name', flat=True)

    # Process form data
    role_name = request.POST.get('role_name')
    role_description = request.POST.get('role_description')
    is_active = request.POST.get('is_active') == '1'  # Convert to boolean
    # Assuming permissions are passed as an array
    permissions = request.POST.getlist('permissions[]')

    # Update role fields
    role.role_name = role_name
    role.role_description = role_description
    role.is_active = is_active
    role.save()

    # Log the previous permissions
    RolePermissionAuditLog.objects.create(
        user=request.user,
        action='Role permissions before update',
        details=f'Previous permissions for role "{role}": {", ".join(current_permissions)}',
        role=str(role),
        permission=", ".join(current_permissions)  # Store previous permissions
    )

    # Update role permissions
    updated_permissions = []

    # Remove existing permissions not in the updated list
    for existing_perm in role.defaultrolepermission_set.all():
        if str(existing_perm.permission.id) not in permissions:
            existing_perm.delete()

    # Add new permissions or reactivate existing ones
    for permission_id in permissions:
        permission_id = int(permission_id)
        permission, created = Permission.objects.get_or_create(
            id=permission_id)
        obj, created = DefaultRolePermission.objects.get_or_create(
            role=role, permission=permission)
        updated_permissions.append(permission.permission_name)

    # Log the update action
    RolePermissionAuditLog.objects.create(
        user=request.user,
        action='Role updated',
        details=f'Role "{role}" updated with permissions: {", ".join(updated_permissions)}',
        role=str(role),
        permission=", ".join(updated_permissions)  # Store updated permissions
    )

    # Return success response
    return JsonResponse({'message': 'Role updated successfully.'})


@login_required
def delete_role(request):
    if request.method == 'POST':
        role_id = request.POST.get('role_id')

        try:
            # Retrieve the Role object by ID
            role = Role.objects.get(id=role_id)

            # Get all profiles assigned to this role
            assigned_profiles = role.profiles.all()  # Use 'profiles' related name here

            # Check if any profiles are assigned to this role
            if assigned_profiles.exists():
                # Create a list of usernames assigned to this role
                assigned_usernames = [
                    profile.user.username for profile in assigned_profiles]

                # Return a response indicating the role is assigned to users and cannot be deleted
                return JsonResponse({
                    'success': False,
                    'message': 'Role is assigned to users. Cannot delete.',
                    'assigned_users': assigned_usernames
                })

            # Save the role details before deleting
            role_name = role.role_name  # Assuming role has a role_name field

            # Retrieve role permissions
            role_permissions = DefaultRolePermission.objects.filter(
                role_id=role_id).values_list('permission__permission_name', flat=True)

            # Create a list of permissions associated with the role
            permissions_list = list(role_permissions)

            # Delete the role if no profiles are assigned to it
            role.delete()

            # Create audit log entry with role and permissions details
            RolePermissionAuditLog.objects.create(
                user=request.user,
                action='Role Deleted',
                details=f'Role "{role_name}"  and associated permissions were deleted.',
                role=role_name,
                permission=", ".join(
                    permissions_list) if permissions_list else None
            )

            return JsonResponse({'success': True, 'message': 'Role deleted successfully.'})

        except Role.DoesNotExist:
            # Return a response indicating the role was not found
            return JsonResponse({'success': False, 'message': 'Role not found.'})

    # Return a response indicating an invalid request method
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})



@csrf_exempt
def user_list(request):
    if request.method == "GET":
        user_id = request.GET.get('user_id', '').strip()
        user_name = request.GET.get('user_name', '').strip()
        role = request.GET.get('role', '').strip()

        # Fetch users along with related Profile and Role models
        users = User.objects.select_related('profile', 'profile__role').all()

        if user_id.isdigit():
            users = users.filter(id=int(user_id))

        if user_name:
            users = users.filter(profile__name__icontains=user_name)

        if role.isdigit():
            users = users.filter(profile__role__id=int(role)) 

        user_list = [
            {
                'id': user.id,
                'name': getattr(user.profile, 'name', ''),  # Correct reference
                'role': getattr(user.profile.role, 'role_name', '') if user.profile and user.profile.role else '',
                'email': user.email,
                'mobile': getattr(user.profile, 'mobile_no', ''),
                'company': getattr(user.profile, 'company', ''),
                'location': getattr(user.profile, 'location', ''),
                'status': 'Active' if user.is_active else 'Inactive',
            }
            for user in users if hasattr(user, 'profile')  # Ensure profile exists
        ]

        # 🔹 Print user_list in the console
        # print("User List:", user_list)

        return JsonResponse({'users': user_list}, status=200)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def get_user_details(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    profile = user.profile
    user_role = profile.role  # Retrieve the role directly from the profile

    # Fetch all roles available in the database
    all_roles = Role.objects.all()

    roles_data = []
    if user_role:
        roles_data.append({
            'id': user_role.id,
            'role_name': user_role.role_name
        })

    all_roles_data = []
    for role in all_roles:
        all_roles_data.append({
            'id': role.id,
            'role_name': role.role_name
        })

    user_data = {
        'id': user.id,
        'name': profile.name,
        'email': user.email,
        'mobile_no': profile.mobile_no,
        'company': profile.company,
        'location': profile.location,
        'role_id': user_role.id if user_role else None,
        'role_name': user_role.role_name if user_role else None,
        'all_roles': all_roles_data,
        'is_active': user.is_active
    }

    return JsonResponse(user_data)


def get_user_data(request):
    user_id = request.GET.get('user_id')
    if user_id:
        user = get_object_or_404(User, id=user_id)
        user_data = {
            'id': user.id,
            'full_name': user.name,
            'email': user.email_address,
            'mobile': user.mobile_no,
            'company': user.company,
            'location': user.location,
            'is_active': user.is_active
        }
        print('User Data:', user_data)
        return JsonResponse(user_data)
    return JsonResponse({'error': 'User ID not provided'}, status=400)


import datetime

def update_user(request):
    if request.method == 'POST':
        user_id = request.POST.get('userId')
        user_name = request.POST.get('userName')
        user_email = request.POST.get('userEmail')
        user_mobile = request.POST.get('userMobile')
        user_company = request.POST.get('userCompany')
        user_location = request.POST.get('userLocation')
        user_role_id = request.POST.get('userRole')
        user_is_active = request.POST.get(
            'userIsActive') == 'true'  # Convert to boolean

        user = get_object_or_404(User, id=user_id)

        # Save the original user details for the audit log
        original_username = user.username
        original_last_login = user.last_login

        # Update user fields
        user.username = user_name
        user.email = user_email
        user.is_active = user_is_active  # Update is_active status
        user.save()

        # Update profile fields if available
        if hasattr(user, 'profile'):
            profile = user.profile
            profile.name = user_name
            profile.mobile_no = user_mobile
            profile.company = user_company
            profile.location = user_location

            # Update user role if provided
            if user_role_id:
                new_role = get_object_or_404(Role, id=user_role_id)
                profile.role = new_role
                profile.save()

                # Update or create UserRole entry
                user_role, created = UserRole.objects.update_or_create(
                    user=profile.user,
                    defaults={'role': new_role}
                )

                # Determine active status display
            active_status = "Active" if user_is_active else "Inactive"

        # Create an audit log entry
        UserAuditLog.objects.create(
            user=request.user,
            action='Update User',
            details=f"User {original_username} updated. Changes: username={user_name}, email={user_email}, is_active={active_status}",
            username=original_username,
            last_login=original_last_login,
            timestamp=datetime.datetime.now()
        )

        return JsonResponse({'success': True, 'message': 'User updated successfully.'})
    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def delete_user(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        try:
            user = get_object_or_404(User, id=user_id)

            # Save user details for audit log before deletion
            user_details = {
                'username': user.username,
                'email': user.email,
                'last_login': user.last_login,
            }

            # Delete the user
            user.delete()

            # Create an audit log entry
            UserAuditLog.objects.create(
                user=request.user,
                action='Delete User',
                details=f"User {user_details['username']} deleted. Details: {user_details}",
                username=user_details['username'],
                last_login=user_details['last_login'],
                timestamp=datetime.datetime.now()
            )

            return JsonResponse({'success': True})
        except Exception as e:
            print(f'Error deleting user: {e}')
            return JsonResponse({'success': False, 'message': str(e)})
    else:
        return JsonResponse({'success': False})


def audit_logs(request):
    return render(request, "audit_log/audit_log.html")


def get_filter_options_users(request):
    if request.method == 'GET':
        usernames = UserAuditLog.objects.values_list(
            'username', flat=True).distinct()
        actions = UserAuditLog.objects.values_list(
            'action', flat=True).distinct()
        users = User.objects.values_list('id', 'username').distinct()

        return JsonResponse({
            'usernames': list(usernames),
            'actions': list(actions),
            'users': list(users)
        })


def get_user_audit_logs(request):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':  # Detect AJAX request
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        username = request.GET.get('username')
        action = request.GET.get('action', '')

        filters = Q()
        if start_date and end_date:
            filters &= Q(date__range=[start_date, end_date])
        if username:
            filters &= Q(username=username)
        if action:
            filters &= Q(action=action)

        audit_logs = UserAuditLog.objects.filter(filters).values(
            'user__username', 'action', 'timestamp', 'details', 'username', 'date', 'last_login'
        )

        return JsonResponse({'data': list(audit_logs)})  

    return render(request, "audit_log/audit_log.html")  


def generate_audit_log_csv(request):
    # Get filter parameters from the request
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    username = request.GET.get('username', '').strip()
    action = request.GET.get('action', '').strip()

    # Filter logs based on input criteria
    audit_logs = UserAuditLog.objects.all()

    if start_date:
        audit_logs = audit_logs.filter(date__gte=start_date)
    if end_date:
        audit_logs = audit_logs.filter(date__lte=end_date)
    if username:
        audit_logs = audit_logs.filter(username__icontains=username)
    if action:
        audit_logs = audit_logs.filter(action__icontains=action)

    # Create CSV Response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="user_audit_log.csv"'

    writer = csv.writer(response)
    writer.writerow(["ID", "Username", "Action", "Details", "Timestamp", "Last Login", "Created At"])

    for log in audit_logs:
        writer.writerow([
            log.id, log.username, log.action, log.details, 
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else "",
            log.last_login.strftime('%Y-%m-%d %H:%M:%S') if log.last_login else "",
            log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else ""
        ])

    return response



def generate_audit_log_excel(request):
    # Get filter parameters
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    username = request.GET.get('username', '').strip()
    action = request.GET.get('action', '').strip()

    # Filter logs based on input criteria
    audit_logs = UserAuditLog.objects.all()

    if start_date:
        audit_logs = audit_logs.filter(date__gte=start_date)
    if end_date:
        audit_logs = audit_logs.filter(date__lte=end_date)
    if username:
        audit_logs = audit_logs.filter(username__icontains=username)
    if action:
        audit_logs = audit_logs.filter(action__icontains=action)

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Audit Log"

    # Define headers
    headers = ["ID", "Username", "Action", "Details", "Timestamp", "Last Login", "Created At"]
    ws.append(headers)

    # Add data rows
    for log in audit_logs:
        ws.append([
            log.id, log.username, log.action, log.details, 
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else "",
            log.last_login.strftime('%Y-%m-%d %H:%M:%S') if log.last_login else "",
            log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else ""
        ])

    # Create HTTP response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="user_audit_log.xlsx"'

    # Save workbook to response
    wb.save(response)

    return response


def generate_user_audit_pdf(request):
    # Get filter parameters
    start_date = request.GET.get('start_date_user', '').strip()
    end_date = request.GET.get('end_date_user', '').strip()
    username = request.GET.get('username_user', '').strip()
    action = request.GET.get('action_user', '').strip()

    # Filter User Audit Log records
    logs = UserAuditLog.objects.all()
    if start_date:
        logs = logs.filter(timestamp__gte=start_date)
    if end_date:
        logs = logs.filter(timestamp__lte=end_date)
    if username:
        logs = logs.filter(username__icontains=username)
    if action:
        logs = logs.filter(action__icontains=action)

    # Create response object
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="user_audit_log.pdf"'

    # Define PDF document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()
    custom_normal = ParagraphStyle('custom_normal', parent=styles['Normal'], fontName='Helvetica', fontSize=11, leading=14, spaceAfter=8)

    # Load header image (if available)
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get current date/time
    report_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(x1=0.7 * inch, y1=1.5 * inch, width=A4[0] - 1.4 * inch, height=A4[1] - 3 * inch, id='normal')
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>USER AUDIT LOG</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains user audit logs including their actions.</font>', custom_normal))
    elements.append(Spacer(1, 10))

    # Table Headers
    data = [["Timestamp", "Username", "Action", "Details"]]

    # Add User Audit Log data to the table
    if logs.exists():
        for log in logs:
            data.append([log.timestamp.strftime('%Y-%m-%d %H:%M:%S'), log.username, log.action, log.details])
    else:
        data.append(["No Data", "-", "-", "-"])

    # Define Table
    col_widths = [2.0 * inch, 1.5 * inch, 2.0 * inch, 2.5 * inch]
    row_height = 22
    table = Table(data, colWidths=col_widths, rowHeights=row_height)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),
    ])
    table.setStyle(style)

    # Add Table to PDF
    elements.append(table)

    # Footer Note
    elements.append(Spacer(1, 15))
    note = Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_normal)
    elements.append(note)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response




# Fetch distinct filter options
def get_filter_options_role_permissions(request):
    if request.method == 'GET':
        actions = RolePermissionAuditLog.objects.values_list('action', flat=True).distinct()
        roles = RolePermissionAuditLog.objects.values_list('role', flat=True).distinct()
        permissions = RolePermissionAuditLog.objects.values_list('permission', flat=True).distinct()
        created_by = RolePermissionAuditLog.objects.values_list('user__username', flat=True).distinct()

        return JsonResponse({
            'actions': list(actions),
            'roles': list(roles),
            'permissions': list(permissions),
            'created_by': list(created_by),
        })

# Fetch role-permission audit logs with filters
def get_role_permission_audit_logs(request):
    if request.method == 'GET':
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        action = request.GET.get('action')
        role = request.GET.get('role')
        permission = request.GET.get('permission')

        filters = Q()
        if start_date and end_date:
            filters &= Q(date__range=[start_date, end_date])
        if action:
            filters &= Q(action=action)
        if role:
            filters &= Q(role=role)
        if permission:
            filters &= Q(permission=permission)

        role_permission_logs = RolePermissionAuditLog.objects.filter(filters).values(
            'date', 'role', 'action', 'details', 'permission'
        )

        return JsonResponse({'data': list(role_permission_logs)})
    
    return render(request, 'audit_log/audit_log.html')

# Export Role Permission Audit Logs as CSV
def generate_role_permission_audit_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="role_permission_audit_log.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Role', 'Action', 'Details', 'Permission'])

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    action = request.GET.get('action')
    role = request.GET.get('role')
    permission = request.GET.get('permission')

    filters = Q()
    if start_date and end_date:
        filters &= Q(date__range=[start_date, end_date])
    if action:
        filters &= Q(action=action)
    if role:
        filters &= Q(role=role)
    if permission:
        filters &= Q(permission=permission)

    role_permission_logs = RolePermissionAuditLog.objects.filter(filters)

    for log in role_permission_logs:
        writer.writerow([log.date, log.role, log.action, log.details, log.permission])

    return response

def generate_role_permission_audit_excel(request):
    # Get filter parameters
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    action = request.GET.get('action', '').strip()
    role = request.GET.get('role', '').strip()
    permission = request.GET.get('permission', '').strip()

    # Filter logs based on input criteria
    filters = Q()
    if start_date:
        filters &= Q(date__gte=start_date)
    if end_date:
        filters &= Q(date__lte=end_date)
    if action:
        filters &= Q(action__icontains=action)
    if role:
        filters &= Q(role__icontains=role)
    if permission:
        filters &= Q(permission__icontains=permission)

    role_permission_logs = RolePermissionAuditLog.objects.filter(filters)

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Role Permission Audit Log"

    # Define headers
    headers = ["ID", "Role", "Action", "Details", "Permission", "Date"]
    ws.append(headers)

    # Add data rows
    for log in role_permission_logs:
        ws.append([
            log.id, log.role, log.action, log.details, log.permission,
            log.date.strftime('%Y-%m-%d %H:%M:%S') if log.date else ""
        ])

    # Create HTTP response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="role_permission_audit_log.xlsx"'

    # Save workbook to response
    wb.save(response)

    return response




def generate_role_permission_audit_pdf(request): 
    # Get filter parameters
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    action = request.GET.get('action', '').strip()
    role = request.GET.get('role', '').strip()
    permission = request.GET.get('permission', '').strip()

    # Filter Role Permission Audit Log records
    filters = Q()
    if start_date:
        filters &= Q(date__gte=start_date)
    if end_date:
        filters &= Q(date__lte=end_date)
    if action:
        filters &= Q(action__icontains=action)
    if role:
        filters &= Q(role__icontains=role)
    if permission:
        filters &= Q(permission__icontains=permission)
    
    logs = RolePermissionAuditLog.objects.filter(filters)
    
    # Create response object
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="role_permission_audit_log.pdf"'

    # Define PDF document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()
    custom_normal = ParagraphStyle('custom_normal', parent=styles['Normal'], fontName='Helvetica', fontSize=10, leading=12, spaceAfter=8, wordWrap='CJK')

    # Load header image (if available)
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")
    report_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(x1=0.7 * inch, y1=1.5 * inch, width=A4[0] - 1.4 * inch, height=A4[1] - 3 * inch, id='normal')
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>ROLE PERMISSION AUDIT LOG</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains role permission audit logs including actions performed on role permissions.</font>', custom_normal))
    elements.append(Spacer(1, 10))

   # Define custom paragraph style for text wrapping
    styles = getSampleStyleSheet()
    custom_style = ParagraphStyle(
        'CustomStyle',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,  # Line height for better readability
        wordWrap='CJK'  # Enables text wrapping
    )

   # Table Headers
    data = [["Date", "Role", "Action", "Permission", "Details"]]

    # Add Role Permission Audit Log data to the table
    if logs.exists():
        for log in logs:
            data.append([
                Paragraph(log.date.strftime('%Y-%m-%d %H:%M:%S'), custom_style),
                Paragraph(log.role, custom_style),
                Paragraph(log.action, custom_style),
                Paragraph(log.permission, custom_style),
                Paragraph(log.details, custom_style)
            ])
    else:
        data.append([
            Paragraph("No Data", custom_style), "-", "-", "-", "-"
        ])

    # Define Table with Auto Row Height and Header Repeat on New Pages
    col_widths = [1.0 * inch, 1.0 * inch, 1.2 * inch, 2.5 * inch, 2.5 * inch]
    table = LongTable(data, colWidths=col_widths, repeatRows=1)  # Header repeats on every page

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Left-align for better readability
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Align text to top
    ])

    table.setStyle(style)

    # Add Table to PDF
    elements.append(table)

    # Footer Note
    elements.append(Spacer(1, 15))
    note = Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_normal)
    elements.append(note)

    # Ensure the table does not break across pages improperly
    elements.append(KeepTogether(table))

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response



@csrf_exempt
def generate_user_csv(request):
    user_id = request.GET.get('user_id', '').strip()
    user_name = request.GET.get('user_name', '').strip()
    role = request.GET.get('role', '').strip()

    users = User.objects.all().prefetch_related('profile', 'profile__role')

    if user_id.isdigit():
        users = users.filter(id=int(user_id))

    if user_name:
        users = users.filter(profile__name__icontains=user_name)

    if role.isdigit():
        users = users.filter(profile__role__id=int(role))

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="user_list.csv"'

    writer = csv.writer(response)
    writer.writerow(["User ID", "Name", "Role", "Email", "Mobile", "Company", "Location", "Status"])

    for user in users:
        profile = getattr(user, 'profile', None)  # Check if the user has a profile
        role_obj = getattr(profile, 'role', None) if profile else None

        writer.writerow([
            user.id,
            getattr(profile, 'name', '') if profile else '',  # Safe access
            getattr(role_obj, 'role_name', '') if role_obj else '',  # Safe access
            user.email,
            getattr(profile, 'mobile_no', '') if profile else '',
            getattr(profile, 'company', '') if profile else '',
            getattr(profile, 'location', '') if profile else '',
            "Active" if user.is_active else "Inactive",
        ])

    return response


@csrf_exempt
def generate_user_excel(request):
    user_id = request.GET.get('user_id', '').strip()
    user_name = request.GET.get('user_name', '').strip()
    role = request.GET.get('role', '').strip()

    users = User.objects.all().prefetch_related('profile', 'profile__role')

    if user_id.isdigit():
        users = users.filter(id=int(user_id))

    if user_name:
        users = users.filter(profile__name__icontains=user_name)

    if role.isdigit():
        users = users.filter(profile__role__id=int(role))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User List"

    headers = ["User ID", "Name", "Role", "Email", "Mobile", "Company", "Location", "Status"]

    # Write headers in bold
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for row_num, user in enumerate(users, 2):
        profile = getattr(user, 'profile', None)  # Check if user has a profile
        role_obj = getattr(profile, 'role', None) if profile else None

        ws.cell(row=row_num, column=1, value=user.id)
        ws.cell(row=row_num, column=2, value=getattr(profile, 'name', '') if profile else '')
        ws.cell(row=row_num, column=3, value=getattr(role_obj, 'role_name', '') if role_obj else '')
        ws.cell(row=row_num, column=4, value=user.email)
        ws.cell(row=row_num, column=5, value=getattr(profile, 'mobile_no', '') if profile else '')
        ws.cell(row=row_num, column=6, value=getattr(profile, 'company', '') if profile else '')
        ws.cell(row=row_num, column=7, value=getattr(profile, 'location', '') if profile else '')
        ws.cell(row=row_num, column=8, value="Active" if user.is_active else "Inactive")

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="user_list.xlsx"'

    wb.save(response)

    return response



@csrf_exempt
def generate_user_pdf(request):
    user_id = request.GET.get('user_id', '').strip()
    user_name = request.GET.get('user_name', '').strip()
    role = request.GET.get('role', '').strip()

    users = User.objects.all().select_related('profile', 'profile__role')

    if user_id.isdigit():
        users = users.filter(id=int(user_id))

    if user_name:
        users = users.filter(profile__name__icontains=user_name)

    if role.isdigit():
        users = users.filter(profile__role__id=int(role))

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="user_list.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, leftMargin=0.7 * inch, rightMargin=0.7 * inch, topMargin=1.5 * inch, bottomMargin=1.5 * inch)
    elements = []
    styles = getSampleStyleSheet()
    custom_normal = ParagraphStyle('custom_normal', parent=styles['Normal'], fontName='Helvetica', fontSize=10, leading=14, spaceAfter=5)

    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            img = Image(header_path, width=width, height=1.5 * inch)
            img.wrapOn(canvas, width, height)
            img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    frame = Frame(0.7 * inch, 1.5 * inch, A4[0] - 1.4 * inch, A4[1] - 3 * inch, id='normal')
    template = PageTemplate(id='custom', frames=[frame], onPage=on_page)
    doc.addPageTemplates([template])

    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>USER DETAILS</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))
    elements.append(Paragraph('<font size=11>This report contains details of users, including their IDs, names, roles, and contact details.</font>', custom_normal))
    elements.append(Spacer(1, 10))

    data = [["User ID", "Name", "Role", "Email", "Mobile", "Company", "Location", "Status"]]

    if users.exists():
        for user in users:
            profile = getattr(user, 'profile', None)
            data.append([
                str(user.id),
                Paragraph(getattr(profile, 'name', 'N/A'), custom_normal),
                Paragraph(getattr(profile.role, 'role_name', 'N/A') if profile and profile.role else 'N/A', custom_normal),
                Paragraph(user.email, custom_normal),
                Paragraph(getattr(profile, 'mobile_no', 'N/A'), custom_normal),
                Paragraph(getattr(profile, 'company', 'N/A'), custom_normal),
                Paragraph(getattr(profile, 'location', 'N/A'), custom_normal),
                Paragraph("Active" if user.is_active else "Inactive", custom_normal)
            ])
    else:
        data.append(["No Data", "-", "-", "-", "-", "-", "-", "-"])

    col_widths = [0.6 * inch, 1.2 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch]
    
    table = Table(data, colWidths=col_widths)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ])
    
    table.setStyle(style)

    elements.append(table)
    elements.append(Spacer(1, 15))
    
    note = Paragraph("<strong>Note:</strong> This is a system-generated report. No manual modifications have been made.", custom_normal)
    elements.append(note)

    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)
    return response




@csrf_exempt
def generate_role_permission_csv(request):
    """Generate CSV file for Role & Permissions with applied filters"""
    role_id = request.GET.get('role_id', '').strip()
    role_name = request.GET.get('role_name', '').strip()
    status = request.GET.get('status', '').strip()

    roles = Role.objects.all()

    # Apply filters
    if role_id.isdigit():
        roles = roles.filter(id=int(role_id))

    if role_name:
        roles = roles.filter(role_name__icontains=role_name)

    if status:
        roles = roles.filter(is_active=(status == 'Active'))

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="role_permission_list.csv"'

    writer = csv.writer(response)
    writer.writerow(["Role ID", "Role Name", "Description", "Permissions", "Status"])

    for role in roles:
        permissions = ", ".join(role.defaultrolepermission_set.values_list('permission__permission_name', flat=True))
        writer.writerow([
            role.id,
            role.role_name,
            role.role_description,
            permissions,
            "Active" if role.is_active else "Inactive",
        ])

    return response


@csrf_exempt
def generate_role_permission_excel(request):
    """Generate Excel file for Role & Permissions with applied filters"""
    role_id = request.GET.get('role_id', '').strip()
    role_name = request.GET.get('role_name', '').strip()
    status = request.GET.get('status', '').strip()

    roles = Role.objects.all()

    # Apply filters
    if role_id.isdigit():
        roles = roles.filter(id=int(role_id))

    if role_name:
        roles = roles.filter(role_name__icontains=role_name)

    if status:
        roles = roles.filter(is_active=(status == 'Active'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Role & Permissions"

    headers = ["Role ID", "Role Name", "Description", "Permissions", "Status"]

    # Write headers in bold
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for row_num, role in enumerate(roles, 2):
        permissions = ", ".join(role.defaultrolepermission_set.values_list('permission__permission_name', flat=True))
        
        ws.cell(row=row_num, column=1, value=role.id)
        ws.cell(row=row_num, column=2, value=role.role_name)
        ws.cell(row=row_num, column=3, value=role.role_description)
        ws.cell(row=row_num, column=4, value=permissions)
        ws.cell(row=row_num, column=5, value="Active" if role.is_active else "Inactive")

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="role_permission_list.xlsx"'

    wb.save(response)

    return response



@csrf_exempt
def generate_role_permission_pdf(request):
    """Generate PDF file for Role & Permissions with auto-adjusting text and header image."""

    # Fetch GET parameters with safe defaults
    role_id = request.GET.get('role_id', '').strip()
    role_name = request.GET.get('role_name', '').strip()
    status = request.GET.get('status', '').strip()

    # Query all roles initially
    roles = Role.objects.all()

    # Apply filters dynamically
    if role_id.isdigit():
        roles = roles.filter(id=int(role_id))
    if role_name:
        roles = roles.filter(role_name__icontains=role_name)
    if status:
        is_active = status.lower() == 'active'
        roles = roles.filter(is_active=is_active)

    # Prepare HTTP Response for PDF download
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="role_permission_report.pdf"'

    # Initialize PDF document
    doc = SimpleDocTemplate(response, pagesize=A4, leftMargin=0.7 * inch, rightMargin=0.7 * inch, topMargin=1.5 * inch, bottomMargin=1.5 * inch)
    elements = []

    # Load header image
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")
    report_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(x1=0.7 * inch, y1=1.5 * inch, width=A4[0] - 1.4 * inch, height=A4[1] - 3 * inch, id='normal')
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Title & Date
    styles = getSampleStyleSheet()
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>ROLE PERMISSION REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Define text wrapping style
    custom_style = ParagraphStyle(
        'CustomStyle',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        wordWrap='CJK'
    )

    # Table Header
    data = [['Role ID', 'Role Name', 'Description', 'Permissions', 'Status']]

    # Add roles to the table
    if roles.exists():
        for role in roles:
            permissions = ", ".join(role.defaultrolepermission_set.values_list('permission__permission_name', flat=True)) or "No Permissions"
            data.append([
                Paragraph(str(role.id), custom_style),
                Paragraph(role.role_name, custom_style),
                Paragraph(role.role_description or "N/A", custom_style),
                Paragraph(permissions, custom_style),
                Paragraph("Active" if role.is_active else "Inactive", custom_style)
            ])
    else:
        data.append([Paragraph("No Data", custom_style), "-", "-", "-", "-"])

    # Define column widths
    col_widths = [0.8 * inch, 1.2 * inch, 2.0 * inch, 2.5 * inch, 1.0 * inch]

    # Create table
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))

    # Add table and note
    elements.append(KeepTogether(table))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("<b>Note:</b> This is a system-generated report.", custom_style))

    # Build and return PDF response
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response


def backup_recovery_page(request):
    return render(request, "settings/backup_recovery.html")


def get_tables():
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT tablename 
                FROM pg_catalog.pg_tables 
                WHERE schemaname = 'public';
            """)
            tables = [row[0] for row in cursor.fetchall()]
        return tables
    except Exception as e:
        print("Error fetching tables:", e)
        return []


def get_tables_ajax(request):
    tables = get_tables()
    
    # Fetch existing schedule data for the logged-in user, if available
    existing_schedules = ScheduledBackup.objects.filter(user=request.user)
    schedule_data = {
        'days': existing_schedules.first().backup_interval_days if existing_schedules.exists() else "",
        'operation': existing_schedules.first().operation if existing_schedules.exists() else "",
        'tables': [schedule.table_name for schedule in existing_schedules]  # List of selected tables
    }

    return JsonResponse({'tables': tables, 'schedule_data': schedule_data})


def fetch_backup_details(request):
    backup_data = ScheduledBackupDetails.objects.all().values('id', 'backup_name', 'sql_file_name', 'txt_file_name', 'from_date', 'to_date', 'scheduled_operation', 'file_size')
    return JsonResponse({'backup_data': list(backup_data)})



@csrf_exempt
def import_db(request):
    if request.method == 'POST':
        # Step 1: Handle file upload and get table names
        if 'sql_file' in request.FILES:
            sql_file = request.FILES['sql_file']
            sql_content = sql_file.read().decode('utf-8')

            # Extract table names from the SQL file
            insert_pattern = r"INSERT INTO (\w+) \("
            matches = re.findall(insert_pattern, sql_content)
            table_names = list(set(matches))  # Remove duplicates

            if request.POST.get('selected_tables'):
                selected_tables = json.loads(request.POST['selected_tables'])
                if not selected_tables:
                    return JsonResponse({"error": "No tables selected."}, status=400)

                # Import selected tables into the database
                try:
                    db_settings = settings.DATABASES['default']
                    conn = psycopg2.connect(
                        dbname=db_settings['NAME'],
                        user=db_settings['USER'],
                        password=db_settings['PASSWORD'],
                        host=db_settings['HOST'],
                        port=db_settings['PORT']
                    )
                    cursor = conn.cursor()

                    for table_name in selected_tables:
                        # Extract and execute SQL for each selected table
                        pattern = f"INSERT INTO {table_name} \((.*?)\) VALUES\s*\((.*?)\);"
                        table_matches = re.findall(
                            pattern, sql_content, re.DOTALL)

                        for match in table_matches:
                            columns = match[0].split(', ')
                            values_str = match[1]
                            values_str = values_str.strip()
                            rows = re.split(r'\),\s*\(', values_str)

                            for row in rows:
                                row = row.strip("()")
                                values = re.split(r"',\s*'", row)
                                values = [v.strip("'") for v in values]
                                values = [v.replace("''", "'") for v in values]
                                values = [None if v =='NULL' else v for v in values]
                                insert_query = f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({', '.join(['%s'] * len(values))})"
                                cursor.execute(insert_query, values)

                    conn.commit()
                    cursor.close()
                    conn.close()

                    return JsonResponse({"messages": ["Database imported successfully."]})

                except Exception as e:
                    return JsonResponse({"error": str(e)}, status=500)

            return JsonResponse({"tables": table_names})

        return JsonResponse({"error": "No SQL file uploaded."}, status=400)

    return JsonResponse({"error": "Invalid request method."}, status=405)


@csrf_exempt
def export_db(request):
    if request.method == 'POST':
        try:
            table_names = request.POST.getlist('table_name')
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')

            if not table_names:
                return HttpResponse("At least one table name is required.", status=400)

            # Process start date
            if start_date:
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d')
                except ValueError:
                    return HttpResponse("Invalid start date format. Use YYYY-MM-DD.", status=400)

            # Process end date
            if end_date:
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d')
                    end_date = end_date.replace(hour=23, minute=59, second=59)
                except ValueError:
                    return HttpResponse("Invalid end date format. Use YYYY-MM-DD.", status=400)

            # Check if end_date is earlier than start_date
            if end_date and start_date and end_date < start_date:
                return HttpResponse("End date cannot be earlier than start date.", status=400)

            db_settings = settings.DATABASES['default']

            def get_table_names_from_db():
                conn = psycopg2.connect(
                    dbname=db_settings['NAME'],
                    user=db_settings['USER'],
                    password=db_settings['PASSWORD'],
                    host=db_settings['HOST'],
                    port=db_settings['PORT']
                )
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT table_name FROM information_schema.tables WHERE table_schema = 'public';")
                all_table_names = [row[0] for row in cursor.fetchall()]
                cursor.close()
                conn.close()
                return all_table_names

            # If "all" is selected, fetch all table names
            if "all" in table_names:
                table_names = get_table_names_from_db()

            queries = []

            for table_name in table_names:
                conn = psycopg2.connect(
                    dbname=db_settings['NAME'],
                    user=db_settings['USER'],
                    password=db_settings['PASSWORD'],
                    host=db_settings['HOST'],
                    port=db_settings['PORT']
                )
                cursor = conn.cursor()

                # Check if table has either `created_at` or `timestamp` column
                cursor.execute(f"""
                    SELECT column_name FROM information_schema.columns 
                    WHERE table_name = '{table_name}' AND column_name IN ('created_at', 'timestamp')
                """)
                date_column = cursor.fetchone()

                query = f"SELECT * FROM {table_name}"

                if date_column:
                    date_column = date_column[0]  # Get the column name (either 'created_at' or 'timestamp')

                    # Append date filters if they are present
                    if start_date and end_date:
                        query += f" WHERE {date_column} BETWEEN '{start_date}' AND '{end_date}'"
                    elif start_date:
                        query += f" WHERE {date_column} >= '{start_date}'"
                    elif end_date:
                        query += f" WHERE {date_column} <= '{end_date}'"

                # If neither `created_at` nor `timestamp` is present, export all rows without filters
                queries.append((table_name, query))

                cursor.close()
                conn.close()

            if not queries:
                return HttpResponse("No valid queries generated.", status=400)

            def export_data(query, table_name):
                conn = psycopg2.connect(
                    dbname=db_settings['NAME'],
                    user=db_settings['USER'],
                    password=db_settings['PASSWORD'],
                    host=db_settings['HOST'],
                    port=db_settings['PORT']
                )
                cursor = conn.cursor()

                # Fetch column names
                cursor.execute(
                    f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}' ORDER BY ordinal_position;")
                columns = [row[0] for row in cursor.fetchall()]

                cursor.execute(query)
                rows = cursor.fetchall()

                output = io.StringIO()

                # Write column names
                columns_str = ', '.join(columns)
                output.write(f'-- Data from table {table_name}\n')
                output.write(f'INSERT INTO {table_name} ({columns_str}) VALUES\n')

                # Write data rows
                row_strs = []
                for row in rows:
                    values_str = ', '.join(map(lambda x: f"'{x}'" if x is not None else 'NULL', row))
                    row_strs.append(f'({values_str})')

                output.write(',\n'.join(row_strs) + ';\n')

                cursor.close()
                conn.close()
                return output.getvalue()

            output = io.StringIO()
            for table_name, query in queries:
                data = export_data(query, table_name)
                output.write(data)

            response = HttpResponse(content_type='application/octet-stream')
            response['Content-Disposition'] = 'attachment; filename=database_backup.sql'
            response.write(output.getvalue())
            return response

        except Exception as e:
            return HttpResponse(f"Error: {e}", status=500)

    elif request.method == 'GET' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        def get_table_names():
            db_settings = settings.DATABASES['default']
            conn = psycopg2.connect(
                dbname=db_settings['NAME'],
                user=db_settings['USER'],
                password=db_settings['PASSWORD'],
                host=db_settings['HOST'],
                port=db_settings['PORT']
            )
            cursor = conn.cursor()
            cursor.execute(
                "SELECT table_name FROM information_schema.tables WHERE table_schema = 'public';")
            table_names = [row[0] for row in cursor.fetchall()]
            cursor.close()
            conn.close()
            return table_names

        table_names = get_table_names()
        return JsonResponse({'table_names': table_names})

    else:
        return render(request, 'settings/backup_recovery.html')
    


def schedule_backup(request):
    if request.method == "POST":
        days = int(request.POST.get('days'))
        selected_tables = request.POST.getlist('tables')
        operation = request.POST.get('operations')


        # Your existing code to update schedules
        existing_schedules = ScheduledBackup.objects.filter(user=request.user)
        existing_schedules.exclude(table_name__in=selected_tables).delete()

        for table in selected_tables:
            ScheduledBackup.objects.update_or_create(
                user=request.user,
                table_name=table,
                defaults={
                    'backup_interval_days': days,
                    'operation': operation,
                    'next_backup': timezone.now() + timedelta(days=days)
                }
            )

        # Return JSON response with success message
        return JsonResponse({'success_message': "Backup schedule updated successfully!"})

    # Prepare data for the GET request
    existing_schedules = ScheduledBackup.objects.filter(user=request.user)
    schedule_data = {
        'days': existing_schedules.first().backup_interval_days if existing_schedules.exists() else "",
        'tables': [schedule.table_name for schedule in existing_schedules]
    }

    return render(request, 'settings/backup_recovery.html', {
        'schedule_data': schedule_data
    })



def get_scheduled_backups(request):
    backups = ScheduledBackup.objects.all()
    data = [
        {
            "table_name": backup.table_name,
            "backup_interval_days": backup.backup_interval_days,
            "next_backup": backup.next_backup,
            "user": backup.user.id
        }
        for backup in backups
    ]
    return JsonResponse(data, safe=False)


backup_dir = os.path.join(os.path.dirname(__file__), 'backups')

def download_sql_file(request, backup_id):
    try:
        backup = ScheduledBackupDetails.objects.get(id=backup_id)
        file_path = os.path.join(backup_dir, backup.sql_file_name)
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=backup.sql_file_name)
    except (ScheduledBackupDetails.DoesNotExist, FileNotFoundError):
        raise Http404("SQL file not found.")

def download_txt_file(request, backup_id):
    try:
        backup = ScheduledBackupDetails.objects.get(id=backup_id)
        file_path = os.path.join(backup_dir, backup.txt_file_name)
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=backup.txt_file_name)
    except (ScheduledBackupDetails.DoesNotExist, FileNotFoundError):
        raise Http404("Text file not found.")
    


def master_reports_page(request):
    return render(request, "reports/master_reports.html")


def report_generate_employee_csv(request):   
    employee_name = request.GET.get('employee_name', '').strip()
    work_location = request.GET.get('work_location', '').strip()
    start_date = request.GET.get('from_date', '')
    end_date = request.GET.get('to_date', '')
    selected_columns = request.GET.get('columns', '').split(',')

    print(f"Filters received: Employee Name: {employee_name}, Work Location: {work_location}, Dates: {start_date} - {end_date}")

    employees = Employee.objects.all()

    # Print all employees before filtering
    print(f"Total employees in DB: {employees.count()}")

    # Filter by Employee Name
    if employee_name:
        if employee_name.isdigit():  
            employees = employees.filter(employee_id=int(employee_name))
        else:
            employees = employees.filter(name__icontains=employee_name)

    # Filter by Work Location
    if work_location:
        employees = employees.filter(work_location__id=work_location)

    # Filter by Date Range
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to proper date object
        end_date = parse_date(end_date)
        if start_date and end_date:
            employees = employees.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

    print(f"Filtered employees count: {employees.count()}")

    # Filter by Employee Name
    if employee_name:
        if employee_name.isdigit():  
            employees = employees.filter(employee_id=int(employee_name))  
        else:  
            employees = employees.filter(name__icontains=employee_name)

    # Filter by Work Location (ID or Name)
    if work_location:
        try:
            if work_location.isdigit():  
                employees = employees.filter(work_location_id=int(work_location))  # Filtering by ID
            else:  
                employees = employees.filter(work_location__name__icontains=work_location)  # Filtering by Name
        except Exception as e:
            print(f"Error filtering work location: {e}")

    print(f"After filtering by work location: {employees.count()}")

    # Filter by Date Range
    if start_date and end_date:
        employees = employees.filter(created_at__range=[start_date, end_date])
        print(f"After filtering by date range: {employees.count()}")

    # Print filtered results before returning response
    print("Filtered employees:", list(employees.values()))

    if not employees.exists():
        print("No data found after filtering.")
        return HttpResponse("No data found", content_type="text/plain")

    # Generate CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="employees.csv"'
    writer = csv.writer(response)

    if not selected_columns or selected_columns == ['']:
        selected_columns = [
            "employee_code", "name", "designation",
            "location", "work_location", "mobile_number", "email"
        ]

    writer.writerow(selected_columns)

    for employee in employees:
        writer.writerow([getattr(employee, col, '') for col in selected_columns])

    return response

from django.utils.timezone import localtime
from datetime import datetime, date

def report_generate_employee_excel(request):
    employee_id = request.GET.get('employee_name', '').strip()
    work_location = request.GET.get('work_location', '').strip()
    start_date = request.GET.get('from_date', '')
    end_date = request.GET.get('to_date', '')
    selected_columns = request.GET.get('columns', '').split(',')

    # Fetch data based on filters
    employees = Employee.objects.all()
    
    # Filter by Employee ID
    if employee_id:
        employees = employees.filter(employee_id=employee_id)

    # Filter by Work Location
    if work_location:
        employees = employees.filter(work_location__id=work_location)

    # Filter by Date Range
    if start_date and end_date:
        employees = employees.filter(created_at__range=[start_date, end_date])

    # Define column mappings
    column_map = {
        "employee_id": "ID",
        "employee_code": "Code",
        "name": "Name",
        "designation": "Designation",
        "location": "Location",
        "work_location": "Work Location",
        "date_of_birth": "Date of Birth",
        "mobile_number": "Mobile",
        "email": "Email",
        "aadhaar_card": "Aadhaar",
        "pan_card": "PAN",
        "created_at": "Created At",
        "updated_at": "Updated At"
    }

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Report"

    # Set the header row
    header_row = [column_map[col] for col in selected_columns if col in column_map]
    ws.append(header_row)

    # Populate data rows
    for emp in employees:
        row_data = []
        for col in selected_columns:
            if hasattr(emp, col):
                value = getattr(emp, col)

                # Convert ForeignKey (Location) to string
                if isinstance(value, Location):
                    value = str(value.name)  # Assuming Location model has a `name` field

                # Handle datetime and date fields separately
                elif isinstance(value, datetime):  # Check for datetime.datetime
                    value = localtime(value).strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(value, date):  # Check for datetime.date
                    value = value.strftime("%Y-%m-%d")  # No time part

                row_data.append(value)

        ws.append(row_data)

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="employee_report.xlsx"'
    wb.save(response)
    
    return response


def calculate_column_widths(data):
    # Calculate the maximum length of content in each column
    max_lengths = [max(len(str(row[i])) for row in data for i in range(len(data[0])))]
    # Normalize widths to fit the page
    total_width = sum(max_lengths)
    column_widths = [((A4[0] - 1.4 * inch) * (length / total_width)) for length in max_lengths]
    return column_widths
    

def report_generate_employee_pdf(request):
    employee_id = request.GET.get('employee_name', '').strip()
    work_location = request.GET.get('work_location', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').strip().split(',')

    # Fetch all employees
    employees = Employee.objects.all()

    # Filter by Employee ID
    if employee_id:
        employees = employees.filter(employee_id=employee_id)

    # Filter by Work Location
    if work_location:
        if work_location.isdigit():
            employees = employees.filter(work_location_id=int(work_location))
        else:
            employees = employees.filter(work_location__name__icontains=work_location)

    # Filter by Date Range
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            employees = employees.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="employee_details.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom Style
    custom_style = ParagraphStyle(
        'custom_style',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        leading=14,
        spaceAfter=8
    )

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(
        x1=0.7 * inch,
        y1=1.5 * inch,
        width=A4[0] - 1.4 * inch,
        height=A4[1] - 3 * inch,
        id='normal'
    )
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>EMPLOYEE DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains employee details including their personal and work-related information.</font>', custom_style))
    elements.append(Spacer(1, 10))

   # Define column headers and corresponding data fields
    column_mapping = {
        'employee_id': 'Employee ID',
        'employee_code': 'Employee Code',
        'name': 'Name',
        'designation': 'Designation',
        'location': 'Location',
        'work_location': 'Work Location',
        'date_of_birth': 'Date of Birth',
        'mobile_number': 'Mobile Number',
        'email': 'Email',
        'aadhaar_card': 'Aadhaar Card',
        'pan_card': 'PAN Card',
        'created_at': 'Created At',
        'updated_at': 'Updated At'
    }

    # Filter column headers based on selected columns
    headers = [column_mapping[col] for col in selected_columns if col in column_mapping]

    # Table Data
    data = [headers]

    # Add employee data to the table
    if employees.exists():
        for emp in employees:
            row = []
            for col in selected_columns:
                if col == 'employee_id':
                    row.append(str(emp.employee_id))
                elif col == 'employee_code':
                    row.append(str(emp.employee_code))
                elif col == 'name':
                    row.append(str(emp.name))
                elif col == 'designation':
                    row.append(str(emp.designation))
                elif col == 'location':
                    row.append(str(emp.location))
                elif col == 'work_location':
                    row.append(str(emp.work_location.name if emp.work_location else ""))
                elif col == 'date_of_birth':
                    row.append(str(emp.date_of_birth.strftime('%Y-%m-%d') if emp.date_of_birth else ""))
                elif col == 'mobile_number':
                    row.append(str(emp.mobile_number))
                elif col == 'email':
                    row.append(str(emp.email))
                elif col == 'aadhaar_card':
                    row.append(str(emp.aadhaar_card))
                elif col == 'pan_card':
                    row.append(str(emp.pan_card))
                elif col == 'created_at':
                    row.append(str(emp.created_at.strftime('%Y-%m-%d %H:%M:%S') if emp.created_at else ""))
                elif col == 'updated_at':
                    row.append(str(emp.updated_at.strftime('%Y-%m-%d %H:%M:%S') if emp.updated_at else ""))
            data.append(row)
    else:
        # If no data is found, add a "No data found" row
        data.append(["No data found for the selected filters."])

    # Calculate column widths proportionally
    total_width = A4[0] - 1.4 * inch  # Total available width (page width minus margins)
    num_columns = len(headers)
    column_widths = [total_width / num_columns] * num_columns

    # Create table
    table = Table(data, colWidths=column_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Enable text wrapping
    ]))

    # Add Table to PDF
    elements.append(table)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response


def get_locations_filter(request):
    locations = Location.objects.all().values('id', 'name', 'shortcode')
    return JsonResponse({'locations': list(locations)})


def report_generate_location_csv(request):
    location_name = request.GET.get('location_name', '').strip()
    status = request.GET.get('status', '').strip()
    start_date = request.GET.get('from_date', '')
    end_date = request.GET.get('to_date', '')
    selected_columns = request.GET.get('columns', '').split(',')

    print(f"Filters received: Location Name: {location_name}, Status: {status}, Dates: {start_date} - {end_date}")

    locations = Location.objects.all()

    # Print all locations before filtering
    print(f"Total locations in DB: {locations.count()}")

    # Filter by Location Name
    if location_name:
        if location_name.isdigit():  
            locations = locations.filter(id=int(location_name))
        else:
            locations = locations.filter(name__icontains=location_name)

    # Filter by Status
    if status:
        locations = locations.filter(status=status)

    # Filter by Date Range
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to proper date object
        end_date = parse_date(end_date)
        if start_date and end_date:
            locations = locations.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

    print(f"Filtered locations count: {locations.count()}")

    # Print filtered results before returning response
    print("Filtered locations:", list(locations.values()))

    if not locations.exists():
        print("No data found after filtering.")
        return HttpResponse("No data found", content_type="text/plain")

    # Generate CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="locations.csv"'
    writer = csv.writer(response)

    # Define default columns if none are selected
    if not selected_columns or selected_columns == ['']:
        selected_columns = [
            "id", "name", "details", "shortcode", "status", "created_at"
        ]

    # Write headers
    writer.writerow(selected_columns)

    # Write data rows
    for location in locations:
        writer.writerow([getattr(location, col, '') for col in selected_columns])

    return response


def report_generate_location_excel(request):
    location_name = request.GET.get('location_name', '').strip()
    status = request.GET.get('status', '').strip()
    start_date = request.GET.get('from_date', '')
    end_date = request.GET.get('to_date', '')
    selected_columns = request.GET.get('columns', '').split(',')

    print(f"Filters received: Location Name: {location_name}, Status: {status}, Dates: {start_date} - {end_date}")
    print(f"Selected columns: {selected_columns}")

    # Fetch data based on filters
    locations = Location.objects.all()

    # Print all locations before filtering
    print(f"Total locations in DB: {locations.count()}")

    # Filter by Location Name
    if location_name:
        if location_name.isdigit():
            locations = locations.filter(id=int(location_name))
        else:
            locations = locations.filter(name__icontains=location_name)
        print(f"After filtering by location name: {locations.count()}")

    # Filter by Status
    if status:
        locations = locations.filter(status=status)
        print(f"After filtering by status: {locations.count()}")

    # Filter by Date Range
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            locations = locations.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            print(f"After filtering by date range: {locations.count()}")

    print(f"Filtered locations count: {locations.count()}")

    # Print filtered results before generating Excel
    print("Filtered locations:", list(locations.values()))

    if not locations.exists():
        print("No data found after filtering.")
        return HttpResponse("No data found", content_type="text/plain")

    # Define column mappings
    column_map = {
        "id": "ID",
        "name": "Name",
        "details": "Details",
        "shortcode": "Shortcode",
        "status": "Status",
        "created_at": "Created At",
        "updated_at": "Updated At"
    }

    # Validate selected columns
    valid_columns = [col for col in selected_columns if col in column_map]
    if not valid_columns:
        print("No valid columns selected. Using default columns.")
        valid_columns = ["id", "name", "details", "shortcode", "status", "created_at"]

    print(f"Valid columns: {valid_columns}")

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Location Report"

    # Set the header row
    header_row = [column_map[col] for col in valid_columns]
    ws.append(header_row)

    # Populate data rows
    for location in locations:
        row_data = []
        for col in valid_columns:
            if hasattr(location, col):
                value = getattr(location, col)

                # Handle datetime and date fields separately
                if isinstance(value, datetime):  # Check for datetime.datetime
                    value = localtime(value).strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(value, date):  # Check for datetime.date
                    value = value.strftime("%Y-%m-%d")  # No time part

                row_data.append(value)
            else:
                # If the column does not exist, append an empty string
                row_data.append('')
                print(f"Warning: Column '{col}' does not exist for location ID {location.id}")

        print(f"Row data: {row_data}")  # Debugging: Print each row's data
        ws.append(row_data)

    # Debugging: Print the entire worksheet data
    print("Worksheet data:")
    for row in ws.iter_rows(values_only=True):
        print(row)

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="location_report.xlsx"'
    wb.save(response)
    
    return response


def report_generate_location_pdf(request):
    location_name = request.GET.get('location_name', '').strip()
    status = request.GET.get('status', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').strip().split(',')

    # Fetch all locations
    locations = Location.objects.all()

    # Filter by Location Name
    if location_name:
        if location_name.isdigit():
            locations = locations.filter(id=int(location_name))
        else:
            locations = locations.filter(name__icontains=location_name)

    # Filter by Status
    if status:
        locations = locations.filter(status=status)

    # Filter by Date Range
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            locations = locations.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="location_details.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom Style
    custom_style = ParagraphStyle(
        'custom_style',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        leading=14,
        spaceAfter=8
    )

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(
        x1=0.7 * inch,
        y1=1.5 * inch,
        width=A4[0] - 1.4 * inch,
        height=A4[1] - 3 * inch,
        id='normal'
    )
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>LOCATION DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains location details including their name, shortcode, status, and other information.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Define column headers and corresponding data fields
    column_mapping = {
        'id': 'ID',
        'name': 'Name',
        'details': 'Details',
        'shortcode': 'Shortcode',
        'status': 'Status',
        'created_at': 'Created At',
        'updated_at': 'Updated At'
    }

    # Filter column headers based on selected columns
    headers = [column_mapping[col] for col in selected_columns if col in column_mapping]

    # Table Data
    data = [headers]

    # Add location data to the table
    if locations.exists():
        for location in locations:
            row = []
            for col in selected_columns:
                if col == 'id':
                    row.append(str(location.id))
                elif col == 'name':
                    row.append(str(location.name))
                elif col == 'details':
                    row.append(str(location.details))
                elif col == 'shortcode':
                    row.append(str(location.shortcode))
                elif col == 'status':
                    row.append(str(location.status))
                elif col == 'created_at':
                    row.append(str(location.created_at.strftime('%Y-%m-%d %H:%M:%S') if location.created_at else ""))
                elif col == 'updated_at':
                    row.append(str(location.updated_at.strftime('%Y-%m-%d %H:%M:%S') if location.updated_at else ""))
            data.append(row)
    else:
        # If no data is found, add a "No data found" row
        data.append(["No data found for the selected filters."])

    # Calculate column widths proportionally
    total_width = A4[0] - 1.4 * inch  # Total available width (page width minus margins)
    num_columns = len(headers)
    column_widths = [total_width / num_columns] * num_columns

    # Create table
    table = Table(data, colWidths=column_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Enable text wrapping
    ]))

    # Add Table to PDF
    elements.append(table)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response


def get_companies_filter(request):
    # Fetch all companies from the database
    companies = Company.objects.all().values('company_id', 'company_name')

    # Prepare the data to be returned as JSON
    data = {
        'companies': list(companies)  # Convert QuerySet to a list of dictionaries
    }

    # Return the data as a JSON response
    return JsonResponse(data)


@require_GET
def report_generate_company_csv(request):
    """
    Generates a CSV report for companies based on filters and selected columns.
    """
    # Extract filters from the request
    company_name = request.GET.get('company_name', '').strip()
    location = request.GET.get('location', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').split(',')

    print(f"Filters received: Company Name: {company_name}, Location: {location}, Dates: {start_date} - {end_date}")

    # Fetch all companies
    companies = Company.objects.all()

    # Filter by Company Name
    if company_name:
        if company_name.isdigit():  # If company_name is an ID
            companies = companies.filter(company_id=int(company_name))
        else:
            companies = companies.filter(company_name__icontains=company_name)

    # Filter by Location
    if location:
        companies = companies.filter(location__icontains=location)

    # Filter by Date Range (if applicable)
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            companies = companies.filter(created_on__gte=start_date, created_on__lte=end_date)

    print(f"Filtered companies count: {companies.count()}")

    if not companies.exists():
        print("No data found after filtering.")
        return HttpResponse("No data found", content_type="text/plain")

    # Generate CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="companies.csv"'
    writer = csv.writer(response)

    # Define default columns if none are selected
    if not selected_columns or selected_columns == ['']:
        selected_columns = [
            "company_id", "company_name", "CIN_number", "GST_number", "location", "address", "email", "phone_number", "created_on"
        ]

    # Write headers
    writer.writerow(selected_columns)

    # Write data rows
    for company in companies:
        writer.writerow([getattr(company, col, '') for col in selected_columns])

    return response


@require_GET
def report_generate_company_excel(request):
    """
    Generates an Excel report for companies based on filters and selected columns.
    """
    # Extract filters from the request
    company_name = request.GET.get('company_name', '').strip()
    location = request.GET.get('location', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').split(',')

    print(f"Filters received: Company Name: {company_name}, Location: {location}, Dates: {start_date} - {end_date}")

    # Fetch all companies
    companies = Company.objects.all()

    # Filter by Company Name
    if company_name:
        if company_name.isdigit():  # If company_name is an ID
            companies = companies.filter(company_id=int(company_name))
        else:
            companies = companies.filter(company_name__icontains=company_name)

    # Filter by Location
    if location:
        companies = companies.filter(location__icontains=location)

    # Filter by Date Range (if applicable)
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            companies = companies.filter(created_on__gte=start_date, created_on__lte=end_date)

    print(f"Filtered companies count: {companies.count()}")

    if not companies.exists():
        print("No data found after filtering.")
        return HttpResponse("No data found", content_type="text/plain")

    # Define column mappings
    column_map = {
        "company_id": "Company ID",
        "company_name": "Company Name",
        "CIN_number": "CIN Number",
        "GST_number": "GST Number",
        "location": "Location",
        "address": "Address",
        "email": "Email",
        "phone_number": "Phone Number",
        "created_on": "Created On"
    }

    # Validate selected columns
    valid_columns = [col for col in selected_columns if col in column_map]
    if not valid_columns:
        print("No valid columns selected. Using default columns.")
        valid_columns = [
            "company_id", "company_name", "CIN_number", "GST_number", "location", "address", "email", "phone_number", "created_on"
        ]

    print(f"Valid columns: {valid_columns}")

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Company Report"

    # Set the header row
    header_row = [column_map[col] for col in valid_columns]
    ws.append(header_row)

    # Populate data rows
    for company in companies:
        row_data = []
        for col in valid_columns:
            if hasattr(company, col):
                value = getattr(company, col)
                if isinstance(value, (datetime, date)):  # Handle date/datetime fields
                    value = value.strftime('%Y-%m-%d %H:%M:%S') if isinstance(value, datetime) else value.strftime('%Y-%m-%d')
                row_data.append(str(value))
            else:
                row_data.append('')
        ws.append(row_data)

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="company_report.xlsx"'
    wb.save(response)
    
    return response


def report_generate_company_pdf(request):
    """
    Generates a PDF report for companies based on filters and selected columns.
    """
    # Extract filters from the request
    company_name = request.GET.get('company_name', '').strip()
    location = request.GET.get('location', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').strip().split(',')

    # Fetch all companies
    companies = Company.objects.all()

    # Filter by Company Name
    if company_name:
        if company_name.isdigit():  # If company_name is an ID
            companies = companies.filter(company_id=int(company_name))
        else:
            companies = companies.filter(company_name__icontains=company_name)

    # Filter by Location
    if location:
        companies = companies.filter(location__icontains=location)

    # Filter by Date Range (assuming created_on field exists)
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            companies = companies.filter(created_on__gte=start_date, created_on__lte=end_date)

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="company_details.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom Style
    custom_style = ParagraphStyle(
        'custom_style',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        leading=14,
        spaceAfter=8
    )

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(
        x1=0.7 * inch,
        y1=1.5 * inch,
        width=A4[0] - 1.4 * inch,
        height=A4[1] - 3 * inch,
        id='normal'
    )
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>COMPANY DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains company details including their name, CIN number, GST number, location, and other information.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Define column headers and corresponding data fields
    column_mapping = {
        'company_id': 'Company ID',
        'company_name': 'Company Name',
        'CIN_number': 'CIN Number',
        'GST_number': 'GST Number',
        'location': 'Location',
        'address': 'Address',
        'email': 'Email',
        'phone_number': 'Phone Number',
        'created_on': 'Created On'
    }

    # Filter column headers based on selected columns
    headers = [column_mapping[col] for col in selected_columns if col in column_mapping]

    # Table Data
    data = [headers]

    # Add company data to the table
    if companies.exists():
        for company in companies:
            row = []
            for col in selected_columns:
                if hasattr(company, col):
                    value = getattr(company, col)
                    if isinstance(value, (datetime, date)):  # Handle date/datetime fields
                        value = value.strftime('%Y-%m-%d %H:%M:%S') if isinstance(value, datetime) else value.strftime('%Y-%m-%d')
                    row.append(str(value))
                else:
                    row.append('')
            data.append(row)
    else:
        # If no data is found, add a "No data found" row
        data.append(["No data found for the selected filters."])

    # Calculate column widths proportionally
    total_width = A4[0] - 1.4 * inch  # Total available width (page width minus margins)
    num_columns = len(headers)
    column_widths = [total_width / num_columns] * num_columns

    # Create table
    table = Table(data, colWidths=column_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Enable text wrapping
    ]))

    # Add Table to PDF
    elements.append(table)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response

def get_exhibitions_filter(request):
    # Fetch all exhibitions from the database
    exhibitions = Exhibition.objects.all().values('exhibition_id', 'exhibition_name')

    # Prepare the data to be returned as JSON
    data = {
        'exhibitions': list(exhibitions)  # Convert QuerySet to a list of dictionaries
    }

    # Return the data as a JSON response
    return JsonResponse(data)


def get_locations_exhibition_filter(request):
    # Fetch unique locations from the Exhibition model
    locations = Exhibition.objects.values_list('location', flat=True).distinct()

    # Prepare the data to be returned as JSON
    data = {
        'locations': [{'id': idx, 'name': location} for idx, location in enumerate(locations, start=1)]
    }

    # Return the data as a JSON response
    return JsonResponse(data)


def report_generate_exhibition_csv(request):
    # Extract filters from the request
    exhibition_name = request.GET.get('exhibition_name', '').strip()
    location = request.GET.get('location', '').strip()
    start_date = request.GET.get('from_date', '')
    end_date = request.GET.get('to_date', '')
    selected_columns = request.GET.get('columns', '').split(',')

    print(f"Filters received: Exhibition Name: {exhibition_name}, Location: {location}, Dates: {start_date} - {end_date}")

    # Fetch all exhibitions
    exhibitions = Exhibition.objects.all()

    # Print all exhibitions before filtering
    print(f"Total exhibitions in DB: {exhibitions.count()}")

    # Filter by Exhibition Name
    if exhibition_name:
        if exhibition_name.isdigit():  # If exhibition_name is an ID
            exhibitions = exhibitions.filter(exhibition_id=int(exhibition_name))
        else:
            exhibitions = exhibitions.filter(exhibition_name__icontains=exhibition_name)

    # Filter by Location
    if location:
        exhibitions = exhibitions.filter(location__icontains=location)

    # Filter by Date Range
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            exhibitions = exhibitions.filter(start_date__gte=start_date, end_date__lte=end_date)

    print(f"Filtered exhibitions count: {exhibitions.count()}")

    # Print filtered results before returning response
    print("Filtered exhibitions:", list(exhibitions.values()))

    if not exhibitions.exists():
        print("No data found after filtering.")
        return HttpResponse("No data found", content_type="text/plain")

    # Generate CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="exhibitions.csv"'
    writer = csv.writer(response)

    # Define default columns if none are selected
    if not selected_columns or selected_columns == ['']:
        selected_columns = [
            "exhibition_id", "exhibition_name", "location", "address", "city", "state", "pincode", "start_date", "end_date", "details"
        ]

    # Write headers
    writer.writerow(selected_columns)

    # Write data rows
    for exhibition in exhibitions:
        writer.writerow([getattr(exhibition, col, '') for col in selected_columns])

    return response



def report_generate_exhibition_excel(request):
    # Extract filters from the request
    exhibition_name = request.GET.get('exhibition_name', '').strip()
    location = request.GET.get('location', '').strip()
    start_date = request.GET.get('from_date', '')
    end_date = request.GET.get('to_date', '')
    selected_columns = request.GET.get('columns', '').split(',')

    print(f"Filters received: Exhibition Name: {exhibition_name}, Location: {location}, Dates: {start_date} - {end_date}")
    print(f"Selected columns: {selected_columns}")

    # Fetch all exhibitions
    exhibitions = Exhibition.objects.all()

    # Print all exhibitions before filtering
    print(f"Total exhibitions in DB: {exhibitions.count()}")

    # Filter by Exhibition Name
    if exhibition_name:
        if exhibition_name.isdigit():  # If exhibition_name is an ID
            exhibitions = exhibitions.filter(exhibition_id=int(exhibition_name))
        else:
            exhibitions = exhibitions.filter(exhibition_name__icontains=exhibition_name)

    # Filter by Location
    if location:
        exhibitions = exhibitions.filter(location__icontains=location)

    # Filter by Date Range
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            exhibitions = exhibitions.filter(start_date__gte=start_date, end_date__lte=end_date)

    print(f"Filtered exhibitions count: {exhibitions.count()}")

    # Print filtered results before generating Excel
    print("Filtered exhibitions:", list(exhibitions.values()))

    if not exhibitions.exists():
        print("No data found after filtering.")
        return HttpResponse("No data found", content_type="text/plain")

    # Define column mappings
    column_map = {
        "exhibition_id": "Exhibition ID",
        "exhibition_name": "Exhibition Name",
        "location": "Location",
        "address": "Address",
        "city": "City",
        "state": "State",
        "pincode": "Pincode",
        "start_date": "Start Date",
        "end_date": "End Date",
        "details": "Details",
        "created_on": "Created On"
    }

    # Validate selected columns
    valid_columns = [col for col in selected_columns if col in column_map]
    if not valid_columns:
        print("No valid columns selected. Using default columns.")
        valid_columns = ["exhibition_id", "exhibition_name", "location", "address", "city", "state", "pincode", "start_date", "end_date", "details"]

    print(f"Valid columns: {valid_columns}")

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exhibition Report"

    # Set the header row
    header_row = [column_map[col] for col in valid_columns]
    ws.append(header_row)

    # Populate data rows
    for exhibition in exhibitions:
        row_data = []
        for col in valid_columns:
            if hasattr(exhibition, col):
                value = getattr(exhibition, col)

                # Handle datetime and date fields separately
                if isinstance(value, datetime):  # Check for datetime.datetime
                    value = localtime(value).strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(value, date):  # Check for datetime.date
                    value = value.strftime("%Y-%m-%d")  # No time part

                row_data.append(value)
            else:
                # If the column does not exist, append an empty string
                row_data.append('')
                print(f"Warning: Column '{col}' does not exist for exhibition ID {exhibition.exhibition_id}")

        print(f"Row data: {row_data}")  # Debugging: Print each row's data
        ws.append(row_data)

    # Debugging: Print the entire worksheet data
    print("Worksheet data:")
    for row in ws.iter_rows(values_only=True):
        print(row)

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="exhibition_report.xlsx"'
    wb.save(response)
    
    return response


def report_generate_exhibition_pdf(request):
    # Extract filters from the request
    exhibition_name = request.GET.get('exhibition_name', '').strip()
    location = request.GET.get('location', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').strip().split(',')

    # Fetch all exhibitions
    exhibitions = Exhibition.objects.all()

    # Filter by Exhibition Name
    if exhibition_name:
        if exhibition_name.isdigit():  # If exhibition_name is an ID
            exhibitions = exhibitions.filter(exhibition_id=int(exhibition_name))
        else:
            exhibitions = exhibitions.filter(exhibition_name__icontains=exhibition_name)

    # Filter by Location
    if location:
        exhibitions = exhibitions.filter(location__icontains=location)

    # Filter by Date Range
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            exhibitions = exhibitions.filter(start_date__gte=start_date, end_date__lte=end_date)

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="exhibition_details.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom Style
    custom_style = ParagraphStyle(
        'custom_style',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        leading=14,
        spaceAfter=8
    )

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(
        x1=0.7 * inch,
        y1=1.5 * inch,
        width=A4[0] - 1.4 * inch,
        height=A4[1] - 3 * inch,
        id='normal'
    )
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>EXHIBITION DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains exhibition details including their name, location, address, and other information.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Define column headers and corresponding data fields
    column_mapping = {
        'exhibition_id': 'Exhibition ID',
        'exhibition_name': 'Exhibition Name',
        'location': 'Location',
        'address': 'Address',
        'city': 'City',
        'state': 'State',
        'pincode': 'Pincode',
        'start_date': 'Start Date',
        'end_date': 'End Date',
        'details': 'Details',
        'created_on': 'Created On'
    }

    # Filter column headers based on selected columns
    headers = [column_mapping[col] for col in selected_columns if col in column_mapping]

    # Table Data
    data = [headers]

    # Add exhibition data to the table
    if exhibitions.exists():
        for exhibition in exhibitions:
            row = []
            for col in selected_columns:
                if col == 'exhibition_id':
                    row.append(str(exhibition.exhibition_id))
                elif col == 'exhibition_name':
                    row.append(str(exhibition.exhibition_name))
                elif col == 'location':
                    row.append(str(exhibition.location))
                elif col == 'address':
                    row.append(str(exhibition.address))
                elif col == 'city':
                    row.append(str(exhibition.city))
                elif col == 'state':
                    row.append(str(exhibition.state))
                elif col == 'pincode':
                    row.append(str(exhibition.pincode))
                elif col == 'start_date':
                    row.append(str(exhibition.start_date.strftime('%Y-%m-%d') if exhibition.start_date else ""))
                elif col == 'end_date':
                    row.append(str(exhibition.end_date.strftime('%Y-%m-%d') if exhibition.end_date else ""))
                elif col == 'details':
                    row.append(str(exhibition.details))
                elif col == 'created_on':
                    row.append(str(exhibition.created_on.strftime('%Y-%m-%d %H:%M:%S') if exhibition.created_on else ""))
            data.append(row)
    else:
        # If no data is found, add a "No data found" row
        data.append(["No data found for the selected filters."])

    # Calculate column widths proportionally
    total_width = A4[0] - 1.4 * inch  # Total available width (page width minus margins)
    num_columns = len(headers)
    column_widths = [total_width / num_columns] * num_columns

    # Create table
    table = Table(data, colWidths=column_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Enable text wrapping
    ]))

    # Add Table to PDF
    elements.append(table)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response


def get_customers_filter(request):
    """
    Returns a JSON response containing a list of customers for filtering.
    """
    try:
        # Fetch all customers from the database
        customers = Customer.objects.all().values('customer_id', 'customer_name')
        
        # Convert the queryset to a list of dictionaries
        customers_list = list(customers)
        
        # Return the JSON response
        return JsonResponse({
            'customers': customers_list,
            'status': 'success',
            'message': 'Customers fetched successfully.'
        }, status=200)
    
    except Exception as e:
        # Handle any exceptions and return an error response
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)
        
        

@require_GET
def report_generate_customer_csv(request):
    """
    Generates a CSV report for customers based on filters and selected columns.
    """
    # Extract filters from the request
    customer_name = request.GET.get('customer_name', '').strip()
    company_name = request.GET.get('company_name', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').split(',')

    print(f"Filters received: Customer Name: {customer_name}, Company Name: {company_name}, Dates: {start_date} - {end_date}")

    # Fetch all customers
    customers = Customer.objects.all()

    # Filter by Customer Name
    if customer_name:
        if customer_name.isdigit():  # If customer_name is an ID
            customers = customers.filter(customer_id=int(customer_name))
        else:
            customers = customers.filter(customer_name__icontains=customer_name)

    # Filter by Company Name
    if company_name:
        customers = customers.filter(company__company_name__icontains=company_name)

    # Filter by Date Range (if applicable)
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            customers = customers.filter(created_on__gte=start_date, created_on__lte=end_date)

    print(f"Filtered customers count: {customers.count()}")

    if not customers.exists():
        print("No data found after filtering.")
        return HttpResponse("No data found", content_type="text/plain")

    # Generate CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="customers.csv"'
    writer = csv.writer(response)

    # Define default columns if none are selected
    if not selected_columns or selected_columns == ['']:
        selected_columns = [
            "customer_id", "customer_name", "email", "company", "phone_number", "address", "city", "state", "zip_code", "country", "details", "created_on"
        ]

    # Write headers
    writer.writerow(selected_columns)

    # Write data rows
    for customer in customers:
        writer.writerow([getattr(customer, col, '') for col in selected_columns])

    return response

@require_GET
def report_generate_customer_excel(request):
    """
    Generates an Excel report for customers based on filters and selected columns.
    """
    # Extract filters from the request
    customer_name = request.GET.get('customer_name', '').strip()
    company_name = request.GET.get('company_name', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').split(',')

    print(f"Filters received: Customer Name: {customer_name}, Company Name: {company_name}, Dates: {start_date} - {end_date}")

    # Fetch all customers
    customers = Customer.objects.all()

    # Filter by Customer Name
    if customer_name:
        if customer_name.isdigit():  # If customer_name is an ID
            customers = customers.filter(customer_id=int(customer_name))
        else:
            customers = customers.filter(customer_name__icontains=customer_name)

    # Filter by Company Name
    if company_name:
        customers = customers.filter(company__company_name__icontains=company_name)

    # Filter by Date Range (if applicable)
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            customers = customers.filter(created_on__gte=start_date, created_on__lte=end_date)

    print(f"Filtered customers count: {customers.count()}")

    if not customers.exists():
        print("No data found after filtering.")
        return HttpResponse("No data found", content_type="text/plain")

    # Define column mappings
    column_map = {
        "customer_id": "Customer ID",
        "customer_name": "Customer Name",
        "email": "Email",
        "company": "Company",
        "phone_number": "Phone Number",
        "address": "Address",
        "city": "City",
        "state": "State",
        "zip_code": "Zip Code",
        "country": "Country",
        "details": "Details",
        "created_on": "Created On"  # Add "Created On" to the column map
    }

    # Validate selected columns
    valid_columns = [col for col in selected_columns if col in column_map]
    if not valid_columns:
        print("No valid columns selected. Using default columns.")
        valid_columns = ["customer_id", "customer_name", "email", "company", "phone_number", "address", "city", "state", "zip_code", "country", "details", "created_on"]

    print(f"Valid columns: {valid_columns}")

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Report"

    # Set the header row
    header_row = [column_map[col] for col in valid_columns]
    ws.append(header_row)

    # Populate data rows
    for customer in customers:
        row_data = []
        for col in valid_columns:
            if hasattr(customer, col):
                value = getattr(customer, col)
                row_data.append(str(value))
            else:
                row_data.append('')
        ws.append(row_data)

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="customer_report.xlsx"'
    wb.save(response)
    
    return response


@require_GET
def report_generate_customer_pdf(request):
    """
    Generates a PDF report for customers based on filters and selected columns.
    """
    # Extract filters from the request
    customer_name = request.GET.get('customer_name', '').strip()
    company_name = request.GET.get('company_name', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').strip().split(',')

    # Fetch all customers
    customers = Customer.objects.all()

    # Filter by Customer Name
    if customer_name:
        if customer_name.isdigit():  # If customer_name is an ID
            customers = customers.filter(customer_id=int(customer_name))
        else:
            customers = customers.filter(customer_name__icontains=customer_name)

    # Filter by Company Name
    if company_name:
        customers = customers.filter(company__company_name__icontains=company_name)

    # Filter by Date Range (if applicable)
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            customers = customers.filter(created_on__gte=start_date, created_on__lte=end_date)

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="customer_details.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom Style
    custom_style = ParagraphStyle(
        'custom_style',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        leading=14,
        spaceAfter=8
    )

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(
        x1=0.7 * inch,
        y1=1.5 * inch,
        width=A4[0] - 1.4 * inch,
        height=A4[1] - 3 * inch,
        id='normal'
    )
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>CUSTOMER DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains customer details including their name, email, company, and other information.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Define column headers and corresponding data fields
    column_mapping = {
        'customer_id': 'Customer ID',
        'customer_name': 'Customer Name',
        'email': 'Email',
        'company': 'Company',
        'phone_number': 'Phone Number',
        'address': 'Address',
        'city': 'City',
        'state': 'State',
        'zip_code': 'Zip Code',
        'country': 'Country',
        'details': 'Details',
        'created_on': 'Created On'  # Add "Created On" to the column mapping
    }

    # Filter column headers based on selected columns
    headers = [column_mapping[col] for col in selected_columns if col in column_mapping]

    # Table Data
    data = [headers]

    # Add customer data to the table
    if customers.exists():
        for customer in customers:
            row = []
            for col in selected_columns:
                if hasattr(customer, col):
                    value = getattr(customer, col)
                    if col == 'created_on':  # Format the "created_on" field
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    row.append(str(value))
                else:
                    row.append('')
            data.append(row)
    else:
        # If no data is found, add a "No data found" row
        data.append(["No data found for the selected filters."])

    # Calculate column widths proportionally
    total_width = A4[0] - 1.4 * inch  # Total available width (page width minus margins)
    num_columns = len(headers)
    column_widths = [total_width / num_columns] * num_columns

    # Create table
    table = Table(data, colWidths=column_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Enable text wrapping
    ]))

    # Add Table to PDF
    elements.append(table)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response



def get_suppliers_filter(request):
    """
    Returns a JSON response containing a list of suppliers for filtering.
    """
    try:
        # Fetch all suppliers from the database
        suppliers = Supplier.objects.all().values('supplier_id', 'supplier_name')
        
        # Convert the queryset to a list of dictionaries
        suppliers_list = list(suppliers)
        
        # Return the JSON response
        return JsonResponse({
            'suppliers': suppliers_list,
            'status': 'success',
            'message': 'Suppliers fetched successfully.'
        }, status=200)
    
    except Exception as e:
        # Handle any exceptions and return an error response
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)
        

@require_GET
def report_generate_supplier_csv(request):
    """
    Generates a CSV report for suppliers based on filters and selected columns.
    """
    # Extract filters from the request
    supplier_name = request.GET.get('supplier_name', '').strip()
    company_name = request.GET.get('company_name', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').split(',')

    print(f"Filters received: Supplier Name: {supplier_name}, Company Name: {company_name}, Dates: {start_date} - {end_date}")

    # Fetch all suppliers
    suppliers = Supplier.objects.all()

    # Filter by Supplier Name
    if supplier_name:
        if supplier_name.isdigit():  # If supplier_name is an ID
            suppliers = suppliers.filter(supplier_id=int(supplier_name))
        else:
            suppliers = suppliers.filter(supplier_name__icontains=supplier_name)

    # Filter by Company Name
    if company_name:
        suppliers = suppliers.filter(company__company_name__icontains=company_name)

    # Filter by Date Range (if applicable)
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            suppliers = suppliers.filter(created_on__gte=start_date, created_on__lte=end_date)

    print(f"Filtered suppliers count: {suppliers.count()}")

    if not suppliers.exists():
        print("No data found after filtering.")
        return HttpResponse("No data found", content_type="text/plain")

    # Generate CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="suppliers.csv"'
    writer = csv.writer(response)

    # Define default columns if none are selected
    if not selected_columns or selected_columns == ['']:
        selected_columns = [
            "supplier_id", "supplier_name", "email", "company", "phone_number", "address", "city", "state", "zip_code", "country", "details", "created_on"
        ]

    # Write headers
    writer.writerow(selected_columns)

    # Write data rows
    for supplier in suppliers:
        row = []
        for col in selected_columns:
            if hasattr(supplier, col):
                value = getattr(supplier, col)
                if col == 'created_on':  # Format the "created_on" field
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                row.append(str(value))
            else:
                row.append('')
        writer.writerow(row)

    return response


@require_GET
def report_generate_supplier_excel(request):
    """
    Generates an Excel report for suppliers based on filters and selected columns.
    """
    # Extract filters from the request
    supplier_name = request.GET.get('supplier_name', '').strip()
    company_name = request.GET.get('company_name', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').split(',')

    print(f"Filters received: Supplier Name: {supplier_name}, Company Name: {company_name}, Dates: {start_date} - {end_date}")

    # Fetch all suppliers
    suppliers = Supplier.objects.all()

    # Filter by Supplier Name
    if supplier_name:
        if supplier_name.isdigit():  # If supplier_name is an ID
            suppliers = suppliers.filter(supplier_id=int(supplier_name))
        else:
            suppliers = suppliers.filter(supplier_name__icontains=supplier_name)

    # Filter by Company Name
    if company_name:
        suppliers = suppliers.filter(company__company_name__icontains=company_name)

    # Filter by Date Range (if applicable)
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            suppliers = suppliers.filter(created_on__gte=start_date, created_on__lte=end_date)

    print(f"Filtered suppliers count: {suppliers.count()}")

    if not suppliers.exists():
        print("No data found after filtering.")
        return HttpResponse("No data found", content_type="text/plain")

    # Define column mappings
    column_map = {
        "supplier_id": "Supplier ID",
        "supplier_name": "Supplier Name",
        "email": "Email",
        "company": "Company",
        "phone_number": "Phone Number",
        "address": "Address",
        "city": "City",
        "state": "State",
        "zip_code": "Zip Code",
        "country": "Country",
        "details": "Details",
        "created_on": "Created On"  # Add "Created On" to the column map
    }

    # Validate selected columns
    valid_columns = [col for col in selected_columns if col in column_map]
    if not valid_columns:
        print("No valid columns selected. Using default columns.")
        valid_columns = ["supplier_id", "supplier_name", "email", "company", "phone_number", "address", "city", "state", "zip_code", "country", "details", "created_on"]

    print(f"Valid columns: {valid_columns}")

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supplier Report"

    # Set the header row
    header_row = [column_map[col] for col in valid_columns]
    ws.append(header_row)

    # Populate data rows
    for supplier in suppliers:
        row_data = []
        for col in valid_columns:
            if hasattr(supplier, col):
                value = getattr(supplier, col)
                if col == 'created_on':  # Format the "created_on" field
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                row_data.append(str(value))
            else:
                row_data.append('')
        ws.append(row_data)

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="supplier_report.xlsx"'
    wb.save(response)
    
    return response


def report_generate_supplier_pdf(request):
    """
    Generates a PDF report for suppliers based on filters and selected columns.
    """
    # Extract filters from the request
    supplier_name = request.GET.get('supplier_name', '').strip()
    company_name = request.GET.get('company_name', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').strip().split(',')

    # Fetch all suppliers
    suppliers = Supplier.objects.all()

    # Filter by Supplier Name
    if supplier_name:
        if supplier_name.isdigit():  # If supplier_name is an ID
            suppliers = suppliers.filter(supplier_id=int(supplier_name))
        else:
            suppliers = suppliers.filter(supplier_name__icontains=supplier_name)

    # Filter by Company Name
    if company_name:
        suppliers = suppliers.filter(company__company_name__icontains=company_name)

    # Filter by Date Range (if applicable)
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            suppliers = suppliers.filter(created_on__gte=start_date, created_on__lte=end_date)

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="supplier_details.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom Style
    custom_style = ParagraphStyle(
        'custom_style',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        leading=14,
        spaceAfter=8
    )

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(
        x1=0.7 * inch,
        y1=1.5 * inch,
        width=A4[0] - 1.4 * inch,
        height=A4[1] - 3 * inch,
        id='normal'
    )
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>SUPPLIER DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains supplier details including their name, email, company, and other information.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Define column headers and corresponding data fields
    column_mapping = {
        'supplier_id': 'Supplier ID',
        'supplier_name': 'Supplier Name',
        'email': 'Email',
        'company': 'Company',
        'phone_number': 'Phone Number',
        'address': 'Address',
        'city': 'City',
        'state': 'State',
        'zip_code': 'Zip Code',
        'country': 'Country',
        'details': 'Details'
    }

    # Filter column headers based on selected columns
    headers = [column_mapping[col] for col in selected_columns if col in column_mapping]

    # Table Data
    data = [headers]

    # Add supplier data to the table
    if suppliers.exists():
        for supplier in suppliers:
            row = []
            for col in selected_columns:
                if hasattr(supplier, col):
                    value = getattr(supplier, col)
                    row.append(str(value))
                else:
                    row.append('')
            data.append(row)
    else:
        # If no data is found, add a "No data found" row
        data.append(["No data found for the selected filters."])

    # Calculate column widths proportionally
    total_width = A4[0] - 1.4 * inch  # Total available width (page width minus margins)
    num_columns = len(headers)
    column_widths = [total_width / num_columns] * num_columns

    # Create table
    table = Table(data, colWidths=column_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Enable text wrapping
    ]))

    # Add Table to PDF
    elements.append(table)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response


@require_GET
def report_generate_supplier_pdf(request):
    """
    Generates a PDF report for suppliers based on filters and selected columns.
    """
    # Extract filters from the request
    supplier_name = request.GET.get('supplier_name', '').strip()
    company_name = request.GET.get('company_name', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').strip().split(',')

    # Fetch all suppliers
    suppliers = Supplier.objects.all()

    # Filter by Supplier Name
    if supplier_name:
        if supplier_name.isdigit():  # If supplier_name is an ID
            suppliers = suppliers.filter(supplier_id=int(supplier_name))
        else:
            suppliers = suppliers.filter(supplier_name__icontains=supplier_name)

    # Filter by Company Name
    if company_name:
        suppliers = suppliers.filter(company__company_name__icontains=company_name)

    # Filter by Date Range (if applicable)
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            suppliers = suppliers.filter(created_on__gte=start_date, created_on__lte=end_date)

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="supplier_details.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom Style
    custom_style = ParagraphStyle(
        'custom_style',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        leading=14,
        spaceAfter=8
    )

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(
        x1=0.7 * inch,
        y1=1.5 * inch,
        width=A4[0] - 1.4 * inch,
        height=A4[1] - 3 * inch,
        id='normal'
    )
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>SUPPLIER DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains supplier details including their name, email, company, and other information.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Define column headers and corresponding data fields
    column_mapping = {
        'supplier_id': 'Supplier ID',
        'supplier_name': 'Supplier Name',
        'email': 'Email',
        'company': 'Company',
        'phone_number': 'Phone Number',
        'address': 'Address',
        'city': 'City',
        'state': 'State',
        'zip_code': 'Zip Code',
        'country': 'Country',
        'details': 'Details',
        'created_on': 'Created On'  # Add "Created On" to the column mapping
    }

    # Filter column headers based on selected columns
    headers = [column_mapping[col] for col in selected_columns if col in column_mapping]

    # Table Data
    data = [headers]

    # Add supplier data to the table
    if suppliers.exists():
        for supplier in suppliers:
            row = []
            for col in selected_columns:
                if hasattr(supplier, col):
                    value = getattr(supplier, col)
                    if col == 'created_on':  # Format the "created_on" field
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    row.append(str(value))
                else:
                    row.append('')
            data.append(row)
    else:
        # If no data is found, add a "No data found" row
        data.append(["No data found for the selected filters."])

    # Calculate column widths proportionally
    total_width = A4[0] - 1.4 * inch  # Total available width (page width minus margins)
    num_columns = len(headers)
    column_widths = [total_width / num_columns] * num_columns

    # Create table
    table = Table(data, colWidths=column_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Enable text wrapping
    ]))

    # Add Table to PDF
    elements.append(table)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response



@require_GET
def get_products_filter(request):
    """
    Returns a JSON response containing a list of products for filtering.
    """
    try:
        # Fetch all products from the database
        products = Product.objects.all().values('product_id', 'product_name')
        
        # Convert the queryset to a list of dictionaries
        products_list = list(products)
        
        # Return the JSON response
        return JsonResponse({
            'products': products_list,
            'status': 'success',
            'message': 'Products fetched successfully.'
        }, status=200)
    
    except Exception as e:
        # Handle any exceptions and return an error response
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)
        
        
@require_GET
def get_categories_filter(request):
    """
    Returns a JSON response containing a list of categories for filtering.
    """
    try:
        # Fetch all categories from the database
        categories = Category.objects.all().values('category_id', 'category_name')
        
        # Convert the queryset to a list of dictionaries
        categories_list = list(categories)
        
        # Return the JSON response
        return JsonResponse({
            'categories': categories_list,
            'status': 'success',
            'message': 'Categories fetched successfully.'
        }, status=200)
    
    except Exception as e:
        # Handle any exceptions and return an error response
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)
        
        
def get_subcategories_filter(request):
    """
    Returns a JSON response containing a list of subcategories for a given category.
    """
    try:
        category_id = request.GET.get('category_id')
        if not category_id:
            return JsonResponse({
                'status': 'error',
                'message': 'Category ID is required.'
            }, status=400)
        
        # Fetch subcategories for the given category
        subcategories = SubCategory.objects.filter(category_id=category_id).values('subcategory_id', 'subcategory_name')
        
        # Convert the queryset to a list of dictionaries
        subcategories_list = list(subcategories)
        
        # Return the JSON response
        return JsonResponse({
            'subcategories': subcategories_list,
            'status': 'success',
            'message': 'Subcategories fetched successfully.'
        }, status=200)
    
    except Exception as e:
        # Handle any exceptions and return an error response
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


@require_GET
def report_generate_product_csv(request):
    """
    Generates a CSV report for products based on filters and selected columns.
    """
    # Extract filters from the request
    product_name = request.GET.get('product_name', '').strip()
    category = request.GET.get('category', '').strip()
    subcategory = request.GET.get('subcategory', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').split(',')

    print(f"Filters received: Product Name: {product_name}, Category: {category}, Subcategory: {subcategory}, Dates: {start_date} - {end_date}")

    # Fetch all products
    products = Product.objects.all()

    # Filter by Product Name
    if product_name:
        if product_name.isdigit():  # If product_name is an ID
            products = products.filter(product_id=int(product_name))
        else:
            products = products.filter(product_name__icontains=product_name)

    # Filter by Category
    if category:
        products = products.filter(category_id=int(category))

    # Filter by Subcategory
    if subcategory:
        products = products.filter(subcategory_id=int(subcategory))

    # Filter by Date Range (if applicable)
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            products = products.filter(created_on__gte=start_date, created_on__lte=end_date)

    print(f"Filtered products count: {products.count()}")

    if not products.exists():
        print("No data found after filtering.")
        return HttpResponse("No data found", content_type="text/plain")

    # Generate CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products.csv"'
    writer = csv.writer(response)

    # Define default columns if none are selected
    if not selected_columns or selected_columns == ['']:
        selected_columns = [
            "product_id", "product_code", "product_name", "category", "subcategory", "date_of_entry",
            "product_size_length", "product_size_breadth", "product_size_height", "product_weight",
            "manufacture_name", "description", "vendor", "barcode", "purchase_date", "purchase_amount", "created_on"
        ]

    # Write headers
    writer.writerow(selected_columns)

    # Write data rows
    for product in products:
        row = []
        for col in selected_columns:
            if hasattr(product, col):
                value = getattr(product, col)
                if isinstance(value, (datetime, date)):  # Handle date/datetime fields
                    value = value.strftime('%Y-%m-%d %H:%M:%S') if isinstance(value, datetime) else value.strftime('%Y-%m-%d')
                row.append(str(value))
            else:
                row.append('')
        writer.writerow(row)

    return response


@require_GET
def report_generate_product_excel(request):
    """
    Generates an Excel report for products based on filters and selected columns.
    """
    # Extract filters from the request
    product_name = request.GET.get('product_name', '').strip()
    category = request.GET.get('category', '').strip()
    subcategory = request.GET.get('subcategory', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').split(',')

    print(f"Filters received: Product Name: {product_name}, Category: {category}, Subcategory: {subcategory}, Dates: {start_date} - {end_date}")

    # Fetch all products
    products = Product.objects.all()

    # Filter by Product Name
    if product_name:
        if product_name.isdigit():  # If product_name is an ID
            products = products.filter(product_id=int(product_name))
        else:
            products = products.filter(product_name__icontains=product_name)

    # Filter by Category
    if category:
        products = products.filter(category_id=int(category))

    # Filter by Subcategory
    if subcategory:
        products = products.filter(subcategory_id=int(subcategory))

    # Filter by Date Range (if applicable)
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            products = products.filter(created_on__gte=start_date, created_on__lte=end_date)

    print(f"Filtered products count: {products.count()}")

    if not products.exists():
        print("No data found after filtering.")
        return HttpResponse("No data found", content_type="text/plain")

    # Define column mappings
    column_map = {
        "product_id": "Product ID",
        "product_code": "Product Code",
        "product_name": "Product Name",
        "category": "Category",
        "subcategory": "Subcategory",
        "date_of_entry": "Date of Entry",
        "product_size_length": "Size (Length)",
        "product_size_breadth": "Size (Breadth)",
        "product_size_height": "Size (Height)",
        "product_weight": "Weight",
        "manufacture_name": "Manufacture Name",
        "description": "Description",
        "vendor": "Vendor",
        "barcode": "Barcode",
        "purchase_date": "Purchase Date",
        "purchase_amount": "Purchase Amount",
        "created_on": "Created On"
    }

    # Validate selected columns
    valid_columns = [col for col in selected_columns if col in column_map]
    if not valid_columns:
        print("No valid columns selected. Using default columns.")
        valid_columns = [
            "product_id", "product_code", "product_name", "category", "subcategory", "date_of_entry",
            "product_size_length", "product_size_breadth", "product_size_height", "product_weight",
            "manufacture_name", "description", "vendor", "barcode", "purchase_date", "purchase_amount", "created_on"
        ]

    print(f"Valid columns: {valid_columns}")

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Report"

    # Set the header row
    header_row = [column_map[col] for col in valid_columns]
    ws.append(header_row)

    # Populate data rows
    for product in products:
        row_data = []
        for col in valid_columns:
            if hasattr(product, col):
                value = getattr(product, col)
                if isinstance(value, (datetime, date)):  # Handle date/datetime fields
                    value = value.strftime('%Y-%m-%d %H:%M:%S') if isinstance(value, datetime) else value.strftime('%Y-%m-%d')
                row_data.append(str(value))
            else:
                row_data.append('')
        ws.append(row_data)

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="product_report.xlsx"'
    wb.save(response)
    
    return response


def report_generate_product_pdf(request):
    """
    Generates a PDF report for products based on filters and selected columns.
    """
    # Extract filters from the request
    product_name = request.GET.get('product_name', '').strip()
    category = request.GET.get('category', '').strip()
    subcategory = request.GET.get('subcategory', '').strip()
    start_date = request.GET.get('from_date', '').strip()
    end_date = request.GET.get('to_date', '').strip()
    selected_columns = request.GET.get('columns', '').strip().split(',')

    # Fetch all products
    products = Product.objects.all()

    # Filter by Product Name
    if product_name:
        if product_name.isdigit():  # If product_name is an ID
            products = products.filter(product_id=int(product_name))
        else:
            products = products.filter(product_name__icontains=product_name)

    # Filter by Category
    if category:
        products = products.filter(category_id=int(category))

    # Filter by Subcategory
    if subcategory:
        products = products.filter(subcategory_id=int(subcategory))

    # Filter by Date Range (if applicable)
    if start_date and end_date:
        start_date = parse_date(start_date)  # Convert to date object
        end_date = parse_date(end_date)      # Convert to date object
        if start_date and end_date:
            products = products.filter(created_on__gte=start_date, created_on__lte=end_date)

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="product_details.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=1.5 * inch,
        bottomMargin=1.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom Style
    custom_style = ParagraphStyle(
        'custom_style',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        leading=14,
        spaceAfter=8
    )

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Define Page Template
    frame = Frame(
        x1=0.7 * inch,
        y1=1.5 * inch,
        width=A4[0] - 1.4 * inch,
        height=A4[1] - 3 * inch,
        id='normal'
    )
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    doc.addPageTemplates([template])

    # Report Title
    elements.append(Spacer(1, 45))
    elements.append(Paragraph('<para align="center"><b><font size=16>PRODUCT DETAILS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 30))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains product details including their name, category, subcategory, and other information.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Define column headers and corresponding data fields
    column_mapping = {
        'product_id': 'Product ID',
        'product_code': 'Product Code',
        'product_name': 'Product Name',
        'category': 'Category',
        'subcategory': 'Subcategory',
        'date_of_entry': 'Date of Entry',
        'product_size_length': 'Size (Length)',
        'product_size_breadth': 'Size (Breadth)',
        'product_size_height': 'Size (Height)',
        'product_weight': 'Weight',
        'manufacture_name': 'Manufacture Name',
        'description': 'Description',
        'vendor': 'Vendor',
        'barcode': 'Barcode',
        'purchase_date': 'Purchase Date',
        'purchase_amount': 'Purchase Amount',
        'created_on': 'Created On'
    }

    # Filter column headers based on selected columns
    headers = [column_mapping[col] for col in selected_columns if col in column_mapping]

    # Table Data
    data = [headers]

    # Add product data to the table
    if products.exists():
        for product in products:
            row = []
            for col in selected_columns:
                if hasattr(product, col):
                    value = getattr(product, col)
                    if isinstance(value, (datetime, date)):  # Handle date/datetime fields
                        value = value.strftime('%Y-%m-%d %H:%M:%S') if isinstance(value, datetime) else value.strftime('%Y-%m-%d')
                    row.append(str(value))
                else:
                    row.append('')
            data.append(row)
    else:
        # If no data is found, add a "No data found" row
        data.append(["No data found for the selected filters."])

    # Calculate column widths proportionally
    total_width = A4[0] - 1.4 * inch  # Total available width (page width minus margins)
    num_columns = len(headers)
    column_widths = [total_width / num_columns] * num_columns

    # Create table
    table = Table(data, colWidths=column_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1)),  # Enable text wrapping
    ]))

    # Add Table to PDF
    elements.append(table)

    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response




def user_profile_page(request):
    return render(request, "users/user_profile.html")


def get_user_data(request):
    user = request.user  # Get the currently logged-in user

    # Handle the case where the user might not have a profile
    try:
        profile = user.profile
    except Profile.DoesNotExist:
        profile = None

    # Handle the case where the user might not have an address
    try:
        address = user.addresses
    except UserAddress.DoesNotExist:
        address = None

    data = {
        "username": user.username,
        "email": user.email,
        "bio": profile.bio if profile else "",
        "name": profile.name if profile else "",
        "mobile": profile.mobile_no if profile else "",
        "company": profile.company if profile else "",
        # Corrected to role_name
        "role": profile.role.role_name if profile and profile.role else "",
        "address1": address.address1 if address else "",
        "address2": address.address2 if address else "",
        "location": profile.location if profile else "",
        "city": address.city if address else "",
        "state": address.state if address else "",
        "zip_code": address.zip_code if address else "",
        "profile_picture_url": profile.profile_picture.url if profile and profile.profile_picture else "",
    }
    return JsonResponse(data)


@csrf_exempt
def update_user_profile(request, user_id):
    if request.method == 'POST':
        data = request.POST
        files = request.FILES

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return JsonResponse({"status": "error", "message": "User not found"}, status=404)

        # Handle profile update
        try:
            profile = user.profile
        except Profile.DoesNotExist:
            profile = Profile(user=user)

        profile.bio = data.get('bio', profile.bio)
        profile.name = data.get('name', profile.name)
        profile.mobile_no = data.get('mobile', profile.mobile_no)
        profile.company = data.get('company', profile.company)
        profile.location = data.get('location', profile.location)

        # Handle address update
        try:
            address = user.addresses
        except UserAddress.DoesNotExist:
            address = UserAddress(user=user)

        address.address1 = data.get('address1', address.address1)
        address.address2 = data.get('address2', address.address2)
        address.city = data.get('city', address.city)
        address.state = data.get('state', address.state)
        address.zip_code = data.get('zip_code', address.zip_code)

        # Handle profile picture upload
        if 'profilePicture' in files:
            profile_picture = files['profilePicture']
            if profile_picture.size < 50 * 1024 or profile_picture.size > 400 * 1024:
                return JsonResponse({"status": "error", "message": "File size must be between 50KB and 400KB."}, status=400)
            if profile_picture.content_type not in ['image/jpeg', 'image/png']:
                return JsonResponse({"status": "error", "message": "Invalid file format. Please upload a .jpg or .png image."}, status=400)
            profile.profile_picture = profile_picture

        # Save both profile and address
        profile.save()
        address.save()

        # Return success response
        return JsonResponse({"status": "success", "message": "Profile updated successfully"})
    else:
        return JsonResponse({"status": "error", "message": "Invalid request method"}, status=400)
    
    



def add_expense_page(request):
    return render(request, "expense/add_expense.html")


def save_expense(request):
    # Generate a unique expense_id when the page is loaded
    expense_id = uuid.uuid4()
    print(f"Generated Expense ID: {expense_id}")
    employees = Employee.objects.all()

    context = {
        'expense_id': expense_id,
        'employees': employees,
    }

    if request.method == 'POST':
        try:
            # Extract form data
            employee_name = request.POST.get('employee_name')
            employee_code = request.POST.get('employee_code')
            work_location = request.POST.get('work_location')
            designation = request.POST.get('designation')
            email = request.POST.get('email')  # Employee's email
            mobile_number = request.POST.get('mobile_number')

            try:
                employee = Employee.objects.get(name=employee_name)
                employee_id = employee.employee_id
            except Employee.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Selected employee not found.'})

            # Extract table data
            sr_nos = request.POST.getlist('sr_no[]')
            start_dates = request.POST.getlist('start_date[]')
            end_dates = request.POST.getlist('end_date[]')
            travel_modes = request.POST.getlist('travel_mode[]')
            details_list = request.POST.getlist('details[]')
            total_amounts_inr = request.POST.getlist('total_amount_inr[]')
            total_amounts_fcy = request.POST.getlist('total_amount_foreign_currency[]')
            currency_codes = request.POST.getlist('currency_code[]')
            attached_file_yes_nos = request.POST.getlist('attached_file_yes_no[]')
            attached_files = request.FILES.getlist('attached_file[]')

            # Calculate the total of total_amount_inr across all rows
            total_calculation = Decimal(0)
            for amount_inr in total_amounts_inr:
                total_calculation += Decimal(amount_inr)

            # Save each row to the Expense model
            expenses = []  # To store all created expense objects
            for i in range(len(sr_nos)):
                # Convert string dates to date objects
                try:
                    from_date = datetime.strptime(start_dates[i], '%Y-%m-%d').date()
                    to_date = datetime.strptime(end_dates[i], '%Y-%m-%d').date()
                except (ValueError, TypeError) as e:
                    return JsonResponse({
                        'success': False, 
                        'error': f'Invalid date format: {str(e)}'
                    })

                # Handle foreign currency amount
                total_amount_fcy = total_amounts_fcy[i] if i < len(total_amounts_fcy) else None
                currency_code = currency_codes[i] if i < len(currency_codes) else None

                if total_amount_fcy and currency_code:
                    total_amount_fcy_formatted = f"{total_amount_fcy}-({currency_code})"
                elif currency_code:
                    total_amount_fcy_formatted = f"({currency_code})"
                else:
                    total_amount_fcy_formatted = None

                # Handle attached file
                attached_document = attached_file_yes_nos[i] == 'yes' if i < len(attached_file_yes_nos) else False
                attached_file = attached_files[i] if i < len(attached_files) else None

                # Create Expense object with proper date objects
                expense = Expense(
                    expense_id=expense_id,
                    from_date=from_date,  # Use converted date object
                    to_date=to_date,    # Use converted date object
                    sr_no=sr_nos[i],
                    details=details_list[i],
                    travel_modes=travel_modes[i],
                    total_amount_inr=Decimal(total_amounts_inr[i]),
                    total_amount_foreign_currency=total_amount_fcy_formatted,
                    total_calculation=total_calculation,
                    total_approved_amount=Decimal(0.00),
                    attached_document=attached_document,
                    employee_id=employee_id,
                    created_at=timezone.now(),
                )

                if attached_file:
                    file_name = default_storage.save(f'expense_documents/{attached_file.name}', attached_file)
                    expense.attached_document_file = file_name

                expense.save()
                expenses.append(expense)

            # After all expenses are saved, create the NewExpenseGenerated record
            if expenses:
                    # Generate the PDF
                    pdf_content = generate_expense_pdf_content(expenses[0])
                    
                    # Get date range for all expenses
                    grouped_dates = Expense.objects.filter(expense_id=expense_id).aggregate(
                        start_date=Min('from_date'),
                        end_date=Max('to_date')
                    )
                    
                    # Create the NewExpenseGenerated record
                    new_report = NewExpenseGenerated.objects.create(
                        expense=expenses[0],
                        employee=employee,
                        email_address=email
                    )
                    
                    # Generate PDF filename
                    employee_name = employee.name.replace(" ", "_")
                    start_date_str = grouped_dates['start_date'].strftime('%Y-%m-%d')
                    end_date_str = grouped_dates['end_date'].strftime('%Y-%m-%d')
                    pdf_filename = f"expense_report_{employee_name}_{start_date_str}_to_{end_date_str}.pdf"
                    
                    # Save the PDF file
                    new_report.new_expense_pdf.save(pdf_filename, ContentFile(pdf_content))

                    # Prepare email context
                    email_context = {
                        'employee_name': employee.name,
                        'expense_id': str(expense_id),
                        'total_claimed_amount': total_calculation,
                        'submission_date': timezone.now(),
                    }

                    # Render HTML email template
                    html_content = render_to_string('expense/email_template.html', email_context)
                    text_content = strip_tags(html_content)

                    # Create email message
                    subject = f"Expense Report Submission - {expense_id}"
                    from_email = settings.DEFAULT_FROM_EMAIL
                    to_email = [email]

                    email_msg = EmailMultiAlternatives(
                        subject=subject,
                        body=text_content,
                        from_email=from_email,
                        to=to_email,
                    )
                    email_msg.attach_alternative(html_content, "text/html")
                    
                    # Attach PDF
                    email_msg.attach(
                        filename=pdf_filename,
                        content=pdf_content,
                        mimetype='application/pdf'
                    )

                    # Send email
                    try:
                        email_msg.send()
                        new_report.mark_as_sent()  # Update sent status
                        print(f"Email sent successfully to {email}")
                    except Exception as e:
                        print(f"Failed to send email: {str(e)}")
                        # You might want to log this error

                # You can mark as sent here if you're sending immediately
                # new_report.mark_as_sent()

            return JsonResponse({'success': True, 'expense_id': str(expense_id)})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return render(request, 'expense/add_expense.html', context)




def generate_expense_pdf_content(expense):
    """Generate modern expense report with dual-column layout"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, inch
    from reportlab.lib import colors
    from io import BytesIO
    from datetime import datetime
    from reportlab.platypus import (Table, TableStyle, SimpleDocTemplate, NextPageTemplate,
                                   Paragraph, Spacer, Frame, Image, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import os
    from django.conf import settings
    from decimal import Decimal
    from PIL import Image as PILImage

    buffer = BytesIO()
    
    # Create PDF Document with wider margins for modern look
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=0.5*inch,
        rightMargin=0.5*inch,
        topMargin=1.25*inch,
        bottomMargin=1*inch
    )
    
    elements = []
    styles = getSampleStyleSheet()

    # Load header/footer images
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")
    footer_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_footer.png")
    
    # Get current date/time
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    current_date = datetime.now().strftime("%d-%b-%Y")

    # Define header/footer function for main document
    def on_page(canvas, doc):
        width, height = A4

        # Draw header image
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.25*inch)
            header_img.drawOn(canvas, 0, height - 1.25*inch)

        # Draw footer image
        if os.path.exists(footer_path):
            footer_img = Image(footer_path, width=width, height=0.75*inch)
            footer_img.drawOn(canvas, 0, 0)

        # Add report date
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(width - 0.5*inch, height - 1.4*inch, f"Generated: {report_date}")

    # Define empty header/footer function for attachment pages
    def on_attachment_page(canvas, doc):
        pass  # No headers/footers for attachment pages

    # Define page template with frames
    frame = Frame(
        x1=0.5*inch, y1=1.25*inch,
        width=A4[0] - 1*inch,
        height=A4[1] - 2*inch,
        id='normal'
    )
    template = PageTemplate(id='custom', frames=frame, onPage=on_page)
    attachment_template = PageTemplate(id='attachment', frames=frame, onPage=on_attachment_page)
    doc.addPageTemplates([template, attachment_template])

     # Custom styles for modern look
    title_style = ParagraphStyle(
        'title_style',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=14,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#333333'),
        spaceAfter=12
    )
    
    section_header_style = ParagraphStyle(
        'section_header',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor('#4472C4'),
        spaceAfter=8
    )
    
    bullet_style = ParagraphStyle(
        'bullet_style',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=12,
        leftIndent=0,
        spaceAfter=6
    )
    
    table_text_style = ParagraphStyle(
        'table_text',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=10,
        wordWrap='LTR',  # Enable word wrapping
        spaceBefore=2,
        spaceAfter=2
    )

    # ========== CONTENT ==========
    # Add spacer to account for header
    elements.append(Spacer(1, 30))

    # Modern title with border
    title_table = Table([
        [Paragraph("EXPENSE REIMBURSEMENT FORM", title_style)]
    ], colWidths=[doc.width])
    title_table.setStyle(TableStyle([
        ('LINEBELOW', (0,0), (-1,-1), 1, colors.HexColor('#4472C4')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    elements.append(title_table)
    
    # Get all expense entries at the beginning of the function
    expenses = Expense.objects.filter(expense_id=expense.expense_id).order_by('from_date')

    # Then later in your code where you create doc_info:
    doc_info = Table([
        [
            Paragraph(
                f"<b>TRISNOTA TECHNICAL SERVICES PVT. LTD.</b><br/>"
               f"<b>Address:</b> 1110, Mayuresh Cosmos, Plot no. 37<br/>"
                f"Sector - 11, CBD Belapur, Navi Mumbai - 400614<br/>"
                f"Maharashtra, India<br/>"
                f"<b>Phone:</b> +91 22 4605 5448<br/>"
                f"<b>Email:</b> accounts@ttspl.co.in",
                styles['Normal']
            ),
            Paragraph(
                f"<b>Expense ID:</b> {expense.expense_id}<br/>"
                f"<b>Expense From:</b> {expenses.first().from_date.strftime('%d-%b-%Y') if expenses.exists() else 'N/A'}<br/>"
                f"<b>Expense To:</b> {expenses.last().to_date.strftime('%d-%b-%Y') if expenses.exists() else 'N/A'}<br/>"
                f"<b>Created on:</b> {current_date}<br/>"
                f"<b>Current Status:</b> {expense.status}",
                styles['Normal']
            )
        ]
    ], colWidths=[doc.width/2]*2)

    elements.append(doc_info)
    elements.append(Spacer(1, 15))

    # ========== DUAL COLUMN EMPLOYEE & SUMMARY SECTION ==========
    # Calculate tax and net payable using Decimal
    tax_rate = Decimal('0.18')  # 18% tax rate
    tax_amount = expense.total_calculation * tax_rate
    net_payable = expense.total_approved_amount * (Decimal('1') - tax_rate)

    dual_column = Table([
    [
        # Left Column - Employee Information
        Paragraph("<b>EMPLOYEE INFORMATION</b>", section_header_style),
        # Right Column - Expense Summary
        Paragraph("<b>EXPENSE SUMMARY</b>", section_header_style)
    ],
    [
        # Employee details with bold labels
        Paragraph(
            f"<b>Name:</b> {expense.employee.name}<br/>"
            f"<b>Employee Code:</b> {expense.employee.employee_code}<br/>"
            f"<b>Designation:</b> {expense.employee.designation}<br/>"
            f"<b>Location:</b> {expense.employee.location}<br/>"
            f"<b>Work Location:</b> {str(expense.employee.work_location) if expense.employee.work_location else 'N/A'}<br/>"
            f"<b>Email:</b> {expense.employee.email}<br/>"
            f"<b>Mobile:</b> {expense.employee.mobile_number}",
            styles['Normal']
        ),
        # Expense summary with bold labels
        Paragraph(
            f"<b>Total Claimed:</b> Rs.{expense.total_calculation:,.2f}<br/>"
            f"<b>Approved Amount:</b> Rs.{expense.total_approved_amount:,.2f}<br/>",
            styles['Normal']
        )
    ]
    ], colWidths=[doc.width/2]*2)
    
    dual_column.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
    ]))
    
    elements.append(dual_column)
    elements.append(Spacer(1, 7))

    # ========== EXPENSE DETAILS TABLE ==========
    elements.append(Paragraph("<b>EXPENSE DETAILS</b>", section_header_style))
    elements.append(Spacer(1, 2))

    # Get all expense entries
    expenses = Expense.objects.filter(expense_id=expense.expense_id).order_by('sr_no')

    # Table headers with all requested columns
    data = [
        [
            "Sr.No", 
            "Start Date", 
            "End Date", 
            "Travel Mode", 
            "Details", 
            "Total Amt (INR)", 
            "Total Amt (FCY)", 
            "FA"
        ]
    ]

    # Add expense rows
    for exp in expenses:
        # Format foreign currency amount if exists
        fcy_amount = exp.total_amount_foreign_currency if exp.total_amount_foreign_currency else "-"
        
        # Determine file attachment status
        if exp.attached_document_file:
            file_status = "Yes" if exp.attached_document else "No"
        else:
            file_status = "No"
        
        data.append([
            f"{exp.sr_no:02d}",  # Zero-padded numbers
            exp.from_date.strftime("%d-%b-%y"),
            exp.to_date.strftime("%d-%b-%y"),
            exp.travel_modes,
            Paragraph(exp.details, table_text_style),  # Full details with wrapping
            f"Rs.{exp.total_amount_inr:,.2f}",
            fcy_amount,
            file_status
        ])

    # Table styling - modern look with auto-adjusted row heights
    col_widths = [
        0.4*inch,   # Sr.No
        0.7*inch,   # Start Date
        0.7*inch,   # End Date
        0.8*inch,   # Travel Mode
        2.0*inch,   # Details (increased width for better wrapping)
        1.1*inch,   # Total Amt (INR)
        1.1*inch,   # Total Amt (FCY)
        0.5*inch    # File Attached
    ]

    # Create table with automatic row heights
    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign='LEFT')

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (4, 0), (4, -1), 'LEFT'),  # Left align details
        ('ALIGN', (5, 0), (6, -1), 'RIGHT'), # Right align amounts
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('WORDWRAP', (4, 1), (4, -1), True),  # Enable word wrapping for details column
        ('MINIMUMHEIGHT', (0, 1), (-1, -1), 0.25*inch),  # Minimum row height
        ('LEADING', (0, 1), (-1, -1), 12),  # Line spacing
    ]))

    elements.append(table)
    elements.append(Spacer(1, 15))

    # ========== ATTACHMENTS SECTION ==========
    # Get attachments info with serial numbers
    attachments = []
    for exp in expenses:
        if exp.attached_document and exp.attached_document_file:
            file_name = os.path.basename(exp.attached_document_file.name)
            attachments.append(f"{exp.sr_no:02d} - {file_name}")

    # Only add attachments section if there are any attachments
    if attachments:
        elements.append(Paragraph("<b>ATTACHMENTS</b>", section_header_style))
        elements.append(Paragraph("<br/>".join(attachments), bullet_style))
        elements.append(Spacer(1, 15))

    # ========== APPROVAL WORKFLOW ==========
    elements.append(Paragraph("<b>APPROVAL WORKFLOW</b>", section_header_style))

    # Determine the approval status text
    approval_status = f"Status: {expense.status}"
    if expense.status.lower() == 'approved':
        approval_status = "Approved By Account Team"

    approval_data = [
        ["Submitted By", "Approved By", "Status"],
        [expense.employee.name, "Account Team", approval_status],
        [current_date, expense.approved_date.strftime('%d-%b-%Y') if hasattr(expense, 'approved_date') and expense.approved_date else "______/______/____", 
        expense.processed_date.strftime('%d-%b-%Y') if hasattr(expense, 'processed_date') and expense.processed_date else "______/______/____"]
    ]

    approval_table = Table(approval_data, colWidths=[doc.width/3]*3, rowHeights=[20, 25, 20])
    approval_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F1F1F1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('LINEBELOW', (0, 1), (-1, 1), 1, colors.lightgrey),
        ('FONTNAME', (2, 1), (2, 1), 'Helvetica-Bold'),  # Make status text bold
        ('TEXTCOLOR', (2, 1), (2, 1), colors.green if expense.status.lower() == 'approved' else colors.black),
    ]))

    elements.append(approval_table)
    elements.append(Spacer(1, 15))

    # ========== NOTES SECTION ==========
    notes = [
        "• Payment will be processed within 5-7 working days after approval",
        "• Original receipts must be submitted for auditing",
        "• Unapproved expenses will be deducted from salary"
    ]
    
    elements.append(Paragraph("<b>NOTES</b>", section_header_style))
    elements.append(Paragraph("<br/>".join(notes), bullet_style))


    # ========== ATTACHMENT PREVIEW SECTION ==========
    supported_extensions = ['.pdf', '.jpg', '.jpeg', '.png']
    has_attachments_to_preview = False

    # First check for attachments
    for exp in expenses:
        if exp.attached_document and exp.attached_document_file:
            file_ext = os.path.splitext(exp.attached_document_file.name)[1].lower()
            if file_ext in supported_extensions:
                has_attachments_to_preview = True
                break

    if has_attachments_to_preview:
        # Switch to attachment template (no headers/footers)
        elements.append(NextPageTemplate('attachment'))
        elements.append(Paragraph("<b>ATTACHMENT PREVIEW</b>", section_header_style))
        elements.append(Spacer(1, 10))
        
        # Define page dimensions with safe margins (more conservative margins)
        PAGE_WIDTH = A4[0] - 2.5 * inch  # Increased margins
        PAGE_HEIGHT = A4[1] - 2.5 * inch
        
        # Process each attachment
        for exp in expenses:
            if exp.attached_document and exp.attached_document_file:
                file_path = exp.attached_document_file.path
                file_ext = os.path.splitext(file_path)[1].lower()
                
                if file_ext in supported_extensions:
                    try:
                        file_name = os.path.basename(file_path)
                        elements.append(Paragraph(
                            f"<b>Attachment {exp.sr_no:02d}: {file_name}</b>", 
                            bullet_style
                        ))
                        
                        if file_ext == '.pdf':
                            try:
                                import fitz  # PyMuPDF
                                try:
                                    pdf_document = fitz.open(file_path)
                                    
                                    # Process ALL pages
                                    for page_num in range(len(pdf_document)):
                                        page = pdf_document.load_page(page_num)
                                        
                                        # Render page to image (lower DPI for smaller size)
                                        pix = page.get_pixmap(dpi=150)  # Reduced from 200 to 150
                                        
                                        # Convert to PIL Image
                                        img = PILImage.frombytes("RGB", [pix.width, pix.height], pix.samples)
                                        
                                        # Create output buffer
                                        img_data = BytesIO()
                                        img.save(img_data, format='JPEG', quality=85)
                                        img_data.seek(0)
                                        
                                        # Calculate dimensions with DPI conversion
                                        img_width = pix.width / 150 * inch
                                        img_height = pix.height / 150 * inch
                                        
                                        # Calculate scaling with maximum size constraints
                                        scale = min(
                                            PAGE_WIDTH / img_width,
                                            PAGE_HEIGHT / img_height,
                                            0.95  # Slightly smaller than available space
                                        )
                                        
                                        # Ensure minimum dimensions
                                        final_width = max(img_width * scale, 0.5 * inch)
                                        final_height = max(img_height * scale, 0.5 * inch)
                                        
                                        # Add image to elements
                                        elements.append(Image(
                                            img_data,
                                            width=final_width,
                                            height=final_height,
                                            hAlign='CENTER'
                                        ))
                                        
                                        elements.append(Paragraph(
                                            f"[Page {page_num+1} of {len(pdf_document)}]", 
                                            styles['Italic']
                                        ))
                                        
                                        # Add page break except for last page
                                        if page_num < len(pdf_document)-1:
                                            elements.append(PageBreak())
                                            
                                except Exception as pdf_error:
                                    elements.append(Paragraph(
                                        f"[PDF Error: {str(pdf_error)}]", 
                                        styles['Italic']
                                    ))
                                finally:
                                    if 'pdf_document' in locals():
                                        pdf_document.close()
                            except ImportError:
                                elements.append(Paragraph(
                                    "[Install PyMuPDF: pip install pymupdf]", 
                                    styles['Italic']
                                ))
                        else:
                            # Process image files (JPG, PNG, etc.)
                            try:
                                with open(file_path, 'rb') as f:
                                    img_data = BytesIO(f.read())
                                
                                with PILImage.open(img_data) as img:
                                    if img.mode in ('P', 'RGBA'):
                                        img = img.convert('RGB')
                                    
                                    output_buffer = BytesIO()
                                    img.save(output_buffer, format='JPEG', quality=90)
                                    output_buffer.seek(0)
                                    
                                    dpi = img.info.get('dpi', (150, 150))[0]
                                    img_width = img.width / dpi * inch
                                    img_height = img.height / dpi * inch
                                    
                                    # More conservative scaling for images
                                    scale = min(
                                        PAGE_WIDTH / img_width,
                                        PAGE_HEIGHT / img_height,
                                        0.9  # Smaller scaling factor
                                    )
                                    
                                    elements.append(Image(
                                        output_buffer,
                                        width=img_width * scale,
                                        height=img_height * scale,
                                        hAlign='CENTER'
                                    ))
                                    
                            except Exception as img_error:
                                elements.append(Paragraph(
                                    f"[Image Error: {str(img_error)}]", 
                                    styles['Italic']
                                ))
                        
                        elements.append(Spacer(1, 12))
                        
                    except Exception as e:
                        elements.append(Paragraph(
                            f"[Error: {str(e)}]", 
                            styles['Italic']
                        ))

        # Switch back to main template
        elements.append(NextPageTemplate('custom'))

    # ========== FOOTER ==========
    footer = Table([
        ["CONFIDENTIAL - Property of Trisnota Technical Services Pvt. Ltd."]
    ], colWidths=[doc.width])
    footer.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTSIZE', (0,0), (-1,-1), 7),
        ('TEXTCOLOR', (0,0), (-1,-1), colors.grey),
        ('LINEABOVE', (0,0), (-1,-1), 0.5, colors.lightgrey),
    ]))
    elements.append(footer)

    # Build PDF
    doc.build(elements)

    pdf_content = buffer.getvalue()
    buffer.close()
    return pdf_content



def download_expense_pdf(request, expense_id):
    # Get the latest generated report for this expense
    report = NewExpenseGenerated.objects.filter(
        expense__expense_id=expense_id
    ).order_by('-created_at').first()  # Changed to apply order_by on the queryset
    
    if not report or not report.new_expense_pdf:
        return HttpResponse("PDF not found", status=404)
    
    response = FileResponse(report.new_expense_pdf.open('rb'))
    response['Content-Disposition'] = f'attachment; filename="{report.get_pdf_filename()}"'
    return response


def get_employee_details(request):
    if request.method == 'GET':
        employee_name = request.GET.get('name')
        print(f"Fetching details for employee: {employee_name}")  # Debugging
        try:
            employee = Employee.objects.get(name=employee_name)
            return JsonResponse({
                'success': True,
                'employee_code': employee.employee_code,
                'work_location': employee.work_location.name if employee.work_location else 'N/A',  # Use the correct field name
                'designation': employee.designation,
                'email': employee.email,
                'mobile_number': employee.mobile_number,
            })
        except Employee.DoesNotExist:
            print(f"Employee not found: {employee_name}")  # Debugging
            return JsonResponse({'success': False, 'error': 'Employee not found.'})
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def track_status_page(request):
    return render(request, "expense/track_status.html")


def expense_list(request):
    # Get filter parameters from request
    status = request.GET.get('status', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Get DataTables parameters
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))  # Default to 10 items per page

    # Start with base queryset
    queryset = Expense.objects.select_related('employee')

    # Apply filters
    if status:
        queryset = queryset.filter(status=status)
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            queryset = queryset.filter(from_date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            queryset = queryset.filter(to_date__lte=end_date)
        except ValueError:
            pass

    # Group and annotate
    grouped_expenses = queryset.values(
        'expense_id',
        'employee__name',
        'employee__employee_code',
        'total_calculation',
        'total_approved_amount',
        'status'
    ).annotate(
        start_date=Min('from_date'),
        end_date=Max('to_date')
    ).order_by('-start_date')

    # Get total count before pagination
    total_records = grouped_expenses.count()

    # Apply pagination
    paginator = Paginator(grouped_expenses, length)
    page = (start // length) + 1
    paginated_expenses = paginator.get_page(page)

    # Convert to JSON format
    data = []
    for expense in paginated_expenses:
        data.append({
            'expense_id': str(expense['expense_id']),
            'employee_info': f"{expense['employee__employee_code']} - {expense['employee__name']}",
            'start_date': expense['start_date'].strftime('%Y-%m-%d'),
            'end_date': expense['end_date'].strftime('%Y-%m-%d'),
            'total_calculation': str(expense['total_calculation']),
            'total_approved_amount': str(expense.get('total_approved_amount', '0.00')),
            'status': expense['status']
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': total_records,
        'data': data
    })




def generate_expense_csv(request):
    # Get filter parameters
    status = request.GET.get('status', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Start with base queryset
    expenses = Expense.objects.select_related('employee')

    # Apply filters
    if status:
        expenses = expenses.filter(status=status)
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            expenses = expenses.filter(from_date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            expenses = expenses.filter(to_date__lte=end_date)
        except ValueError:
            pass

    # Group and annotate
    expenses = expenses.values(
        'expense_id',
        'employee__name',
        'employee__employee_code',
        'total_calculation',
        'total_approved_amount',
        'status'
    ).annotate(
        start_date=Min('from_date'),
        end_date=Max('to_date')
    ).order_by('-start_date')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="expense_report.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Expense ID', 'Employee Code', 'Employee Name', 'Start Date', 
        'End Date', 'Amount Claimed (₹)', 'Amount Approved (₹)', 'Status'
    ])

    for expense in expenses:
        writer.writerow([
            expense['expense_id'],
            expense['employee__employee_code'],
            expense['employee__name'],
            expense['start_date'].strftime('%Y-%m-%d'),
            expense['end_date'].strftime('%Y-%m-%d'),
            expense['total_calculation'],
            expense['total_approved_amount'],
            expense['status']
        ])

    return response


def generate_expense_excel(request):
    # Get filter parameters
    status = request.GET.get('status', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Start with base queryset
    expenses = Expense.objects.select_related('employee')

    # Apply filters
    if status:
        expenses = expenses.filter(status=status)
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            expenses = expenses.filter(from_date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            expenses = expenses.filter(to_date__lte=end_date)
        except ValueError:
            pass

    # Group and annotate
    expenses = expenses.values(
        'expense_id',
        'employee__name',
        'employee__employee_code',
        'total_calculation',
        'total_approved_amount',
        'status'
    ).annotate(
        start_date=Min('from_date'),
        end_date=Max('to_date')
    ).order_by('-start_date')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    headers = [
        'Expense ID', 'Employee Code', 'Employee Name', 'Start Date', 
        'End Date', 'Amount Claimed (₹)', 'Amount Approved (₹)', 'Status'
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header).font = openpyxl.styles.Font(bold=True)

    for row_num, expense in enumerate(expenses, 2):
        ws.cell(row=row_num, column=1, value=str(expense['expense_id']))
        ws.cell(row=row_num, column=2, value=expense['employee__employee_code'])
        ws.cell(row=row_num, column=3, value=expense['employee__name'])
        ws.cell(row=row_num, column=4, value=expense['start_date'].strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=5, value=expense['end_date'].strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=6, value=float(expense['total_calculation']))
        ws.cell(row=row_num, column=7, value=float(expense['total_approved_amount']))
        ws.cell(row=row_num, column=8, value=expense['status'])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="expense_report.xlsx"'

    wb.save(response)
    return response



def generate_expense_pdf(request):
    # Get filter parameters
    status = request.GET.get('status', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Start with base queryset
    expenses = Expense.objects.select_related('employee')

    # Apply filters
    if status:
        expenses = expenses.filter(status=status)
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            expenses = expenses.filter(from_date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            expenses = expenses.filter(to_date__lte=end_date)
        except ValueError:
            pass

    # Group and annotate
    expenses = expenses.values(
        'expense_id',
        'employee__name',
        'employee__employee_code',
        'total_calculation',
        'total_approved_amount',
        'status'
    ).annotate(
        start_date=Min('from_date'),
        end_date=Max('to_date')
    ).order_by('-start_date')

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="expense_report.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles for wrapping text
    custom_style = ParagraphStyle(
        'custom_style',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=12,
        spaceAfter=6
    )

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Report Title
    elements.append(Spacer(1, 80))
    elements.append(Paragraph('<para align="center"><b><font size=16>EXPENSE REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains details of expenses including employee information, dates, and amounts.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Table Headers
    data = [
        [
            "Expense ID", 
            "Employee", 
            "Period", 
            "Amount Claimed", 
            "Amount Approved", 
            "Status"
        ]
    ]
    
    # Fetch Expenses and Append to Table Data
    if expenses.exists():
        for expense in expenses:
            employee_info = f"{expense['employee__employee_code']} - {expense['employee__name']}"
            period = f"{expense['start_date'].strftime('%Y-%m-%d')} to {expense['end_date'].strftime('%Y-%m-%d')}"
            
            # Use Paragraph for text wrapping
            data.append([
                Paragraph(str(expense['expense_id']), custom_style),
                Paragraph(employee_info, custom_style),
                Paragraph(period, custom_style),
                Paragraph(f"₹{float(expense['total_calculation']):,.2f}", custom_style),
                Paragraph(f"₹{float(expense['total_approved_amount']):,.2f}", custom_style),
                Paragraph(expense['status'], custom_style)
            ])
    else:
        data.append(["No Data", "-", "-", "-", "-", "-"])  

    # Table Column Widths (adjust as needed)
    col_widths = [
        1.2 * inch,  # Expense ID
        1.5 * inch,  # Employee
        1.2 * inch,  # Period
        1.2 * inch,  # Amount Claimed
        1.2 * inch,  # Amount Approved
        1.0 * inch   # Status
    ]
    
    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),  # Blue header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#D9E1F2')),  # Light blue rows
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#8EA9DB')),  # Light grid
        ('WORDWRAP', (0, 0), (-1, -1), True),  # Enable word wrap
    ])
    table.setStyle(style)

    # Add table to elements
    elements.append(table)
    elements.append(Spacer(1, 15))

    # Footer Note
    elements.append(Paragraph("<strong>Note:</strong> This is a system-generated expense report.", custom_style))
    
    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response



def verify_process_page(request):
    return render(request, "expense/verify_and_process.html")

def view_verify_and_process_page(request):
    return render(request, "expense/view_verify_and_process.html")


def process_expense(request, expense_id):
    expense_items = Expense.objects.filter(expense_id=expense_id)
    
    if not expense_items.exists():
        return render(request, '404.html', status=404)
    
    if request.method == 'POST':
        try:
            print("\n===== FORM DATA RECEIVED =====")
            print("POST data:", request.POST)
            
            # Get form data
            total_approved = request.POST.get('total_approved', '0')
            approver_comments = request.POST.get('approver_comments', '')
            expense_status = request.POST.get('expense_status', 'Pending')
            send_email = request.POST.get('send_mail_notification') == 'on'
            
            print(f"\nTotal Approved: {total_approved}")
            print(f"Expense Status: {expense_status}")
            print(f"Send Email: {send_email}")
            print(f"Approver Comments: {approver_comments}")
            
            # Get the employee from the first expense item
            processing_employee = expense_items.first().employee
            employee = processing_employee  # For email context
            
            # Update each expense item
            for item in expense_items:
                print(f"\nProcessing Expense Item ID: {item.id}")
                
                # Get individual approved amount and remarks for each item
                approved_amount = request.POST.get(f'approved_amount_{item.id}', '0')
                remarks = request.POST.get(f'remarks_{item.id}', '')
                
                print(f"Approved Amount: {approved_amount}")
                print(f"Remarks: {remarks}")
                
                # Update or create VerifyProcessExpense record
                verification, created = VerifyProcessExpense.objects.update_or_create(
                    expense=item,
                    defaults={
                        'employee': processing_employee,
                        'approved_amount': approved_amount,
                        'remarks': remarks,
                        'approver_comments': approver_comments,
                        'verified_at': timezone.now()  # Update verification timestamp
                    }
                )
                
                # Update Expense model
                item.total_approved_amount = approved_amount
                item.status = expense_status
                item.save()
            
            # Update total approved amount in all related expense items
            expense_items.update(total_approved_amount=total_approved)
            
            # Send email only if status is Approved/Rejected AND send_email is checked
            if send_email and expense_status in ['Approved', 'Rejected']:
                try:
                    # Prepare email context
                    context = {
                        'expense_id': expense_id,
                        'employee_name': employee.name,
                        'total_claimed_amount': sum(float(item.total_amount_inr) for item in expense_items),
                        'total_approved_amount': total_approved,
                        'current_status': expense_status,
                        'approver_comments': approver_comments,
                        'expense_period': f"{min(item.from_date for item in expense_items).strftime('%b %d, %Y')} to {max(item.to_date for item in expense_items).strftime('%b %d, %Y')}",
                    }
                    
                    # Render HTML email template
                    html_message = render_to_string('expense/expense_status_email.html', context)
                    plain_message = strip_tags(html_message)
                    
                    # Create email
                    email = EmailMessage(
                      f'Expense #{expense_id} {expense_status}' + (' - Payment in Process' if expense_status != 'Rejected' else ''),
                        plain_message,
                        settings.DEFAULT_FROM_EMAIL,
                        [employee.email],  # Send to the employee's email
                    )
                    
                    # Attach HTML content
                    email.content_subtype = "html"
                    email.body = html_message
                    
                    # Send email
                    email.send()
                    
                    # Update verification records to mark email as sent
                    VerifyProcessExpense.objects.filter(expense__in=expense_items).update(
                        email_sent=True,
                        email_sent_date=timezone.now()
                    )
                    
                    print(f"Email notification sent to {employee.email}")
                except Exception as email_error:
                    print(f"Failed to send email: {str(email_error)}")
                    # Continue processing even if email fails
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            
            messages.success(request, 'Expense successfully verified and processed!')
            return redirect('expense_list')
            
        except Exception as e:
            print(f"\nERROR: {str(e)}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)}, status=400)
            messages.error(request, f'Error processing expense: {str(e)}')
            return redirect('process_expense', expense_id=expense_id)
    
    # GET request handling
    first_item = expense_items.first()
    employee = first_item.employee
    total_calculation = sum(item.total_amount_inr for item in expense_items)
    
    # Create a list to store items with attached verification data
    expense_items_with_verification = []
    
    for item in expense_items:
        try:
            verification = VerifyProcessExpense.objects.get(expense=item)
            # Attach verification data directly to the item
            item.approved_amount = str(verification.approved_amount)
            item.remarks = verification.remarks or ''
            item.approver_comments = verification.approver_comments or ''
            item.email_sent = verification.email_sent
        except VerifyProcessExpense.DoesNotExist:
            item.approved_amount = '0.00'
            item.remarks = ''
            item.approver_comments = ''
            item.email_sent = False
        
        expense_items_with_verification.append(item)
    
    date_range = expense_items.aggregate(
        start_date=Min('from_date'),
        end_date=Max('to_date')
    )
    
    # Calculate total approved amount
    total_approved = sum(
        float(item.approved_amount)
        for item in expense_items_with_verification
    )
    
    context = {
        'expense_id': expense_id,
        'employee': employee,
        'expense_items': expense_items_with_verification,
        'total_calculation': total_calculation,
        'created_date': first_item.created_at.strftime('%Y-%m-%d'),
        'expense_period': f"{date_range['start_date'].strftime('%Y-%m-%d')} to {date_range['end_date'].strftime('%Y-%m-%d')}",
        'date_range': date_range,
        'readonly': False,
        'first_verification': {
            'approved_amount': expense_items_with_verification[0].approved_amount if expense_items_with_verification else '0.00',
            'remarks': expense_items_with_verification[0].remarks if expense_items_with_verification else '',
            'approver_comments': expense_items_with_verification[0].approver_comments if expense_items_with_verification else '',
            'email_sent': expense_items_with_verification[0].email_sent if expense_items_with_verification else False
        },
        'first_item_total_approved': str(total_approved) if total_approved > 0 else str(first_item.total_approved_amount or '0.00'),
        'status': first_item.status
    }
    
    return render(request, 'expense/view_verify_and_process.html', context)



def view_process_expense(request, expense_id):
    expense_items = Expense.objects.filter(expense_id=expense_id)
    
    if not expense_items.exists():
        return render(request, '404.html', status=404)
    
    # Get verification data and attach to each expense item
    expense_items_with_verification = []
    for item in expense_items:
        try:
            verification = VerifyProcessExpense.objects.get(expense=item)
            item.verification = {
                'approved_amount': str(verification.approved_amount),
                'remarks': verification.remarks or '',
                'approver_comments': verification.approver_comments or '',
                'email_sent': verification.email_sent
            }
        except VerifyProcessExpense.DoesNotExist:
            item.verification = {
                'approved_amount': '0.00',
                'remarks': '',
                'approver_comments': '',
                'email_sent': False
            }
        expense_items_with_verification.append(item)
    
    first_item = expense_items.first()
    employee = first_item.employee
    total_calculation = sum(item.total_amount_inr for item in expense_items)
    
    date_range = expense_items.aggregate(
        start_date=Min('from_date'),
        end_date=Max('to_date')
    )
    
    # Calculate total approved amount
    total_approved = sum(
        float(item.verification['approved_amount'])
        for item in expense_items_with_verification
    )
    
    context = {
        'expense_id': expense_id,
        'employee': employee,
        'expense_items': expense_items_with_verification,
        'total_calculation': total_calculation,
        'created_date': first_item.created_at.strftime('%Y-%m-%d'),
        'expense_period': f"{date_range['start_date'].strftime('%Y-%m-%d')} to {date_range['end_date'].strftime('%Y-%m-%d')}",
        'date_range': date_range,
        'readonly': True,
        'first_verification': expense_items_with_verification[0].verification,
        'first_item_total_approved': str(total_approved) if total_approved > 0 else str(first_item.total_approved_amount or '0.00'),
        'status': first_item.status
    }
    
    return render(request, 'expense/view_process_expense.html', context)



def finance_process_list_page(request):
    return render(request, "expense/finance_process_list.html")


def expensefinanceprocess_list(request):
    # Get filter parameters from request
    employee_id = request.GET.get('employee_id', '')  # Changed from employee_name to employee_id
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Get DataTables parameters
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))

    # Start with base queryset - Approved or Payment Successful expenses
    queryset = Expense.objects.filter(status__in=['Approved', 'Payment Successful']).select_related('employee')

    # Apply filters
    if employee_id:
        queryset = queryset.filter(employee_id=employee_id)
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            queryset = queryset.filter(from_date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            queryset = queryset.filter(to_date__lte=end_date)
        except ValueError:
            pass

    # Group and annotate
    grouped_expenses = queryset.values(
        'expense_id',
        'employee__name',
        'employee__employee_code',
        'total_calculation',
        'total_approved_amount',
        'status'
    ).annotate(
        start_date=Min('from_date'),
        end_date=Max('to_date')
    ).order_by('-status', '-start_date')  # Payment Successful first, then Approved

    # Get total count before pagination
    total_records = grouped_expenses.count()

    # Apply pagination
    paginator = Paginator(grouped_expenses, length)
    page = (start // length) + 1
    paginated_expenses = paginator.get_page(page)

    # Convert to JSON format
    data = []
    for expense in paginated_expenses:
        data.append({
            'expense_id': str(expense['expense_id']),
            'employee_info': f"{expense['employee__employee_code']} - {expense['employee__name']}",
            'start_date': expense['start_date'].strftime('%Y-%m-%d'),
            'end_date': expense['end_date'].strftime('%Y-%m-%d'),
            'total_calculation': str(expense['total_calculation']),
            'total_approved_amount': str(expense.get('total_approved_amount', '0.00')),
            'status': expense['status']
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': total_records,
        'data': data
    })


def get_employee_list(request):
    """New view to provide employee data for Select2 dropdown"""
    search = request.GET.get('search', '')
    
    # Query employees with search term (both name and code)
    employees = Employee.objects.filter(
        Q(name__icontains=search) | Q(employee_code__icontains=search)
    ).values('employee_id', 'name', 'employee_code')[:20]  # Use employee_id instead of id
    
    # Format results for Select2
    results = [{
        'id': emp['employee_id'],  # Use employee_id here
        'text': f"{emp['employee_code']} - {emp['name']}"
    } for emp in employees]
    
    return JsonResponse({'results': results})


def view_finance_process_page(request):
    return render(request, "expense/view_finance_process.html")


def expense_finance_process(request, expense_id):
    # Get all expenses with this ID (should normally be just one)
    expenses = Expense.objects.filter(expense_id=expense_id)
    
    if not expenses.exists():
        return render(request, '404.html', status=404)
    
    # For safety, we'll use the first expense
    expense = expenses.first()
    
    # Get or create finance process for this expense
    finance_process, created = ExpenseFinanceProcess.objects.get_or_create(expense=expense)
    
    if request.method == 'POST':
        try:
            # Process form data
            finance_process.payment_mode = request.POST.get('payment_mode')
            finance_process.payment_date = request.POST.get('payment_date')
            finance_process.finance_comment = request.POST.get('finance_comment', '')
            send_email = 'payment_email_sent' in request.POST
            
            # Handle file upload
            if 'payment_file' in request.FILES:
                finance_process.payment_file = request.FILES['payment_file']
            
            finance_process.save()
            
            # Update ALL expense records with this ID
            expenses.update(status='Payment Successful')
            
            # Send email if checkbox was checked
            if send_email:
                # Prepare email content
                context = {
                    'expense_id': expense.expense_id,
                    'employee_name': expense.employee.name,
                    'total_claimed_amount': expense.total_calculation,
                    'total_approved_amount': expense.total_approved_amount,
                    'current_status': 'Payment Successful',
                    'payment_mode': finance_process.payment_mode,
                    'payment_date': finance_process.payment_date,
                    'finance_comment': finance_process.finance_comment,
                }
                
                # Render HTML email template
                html_message = render_to_string('expense/payment_confirmation_email.html', context)
                plain_message = strip_tags(html_message)
                
                # Create email
                email = EmailMessage(
                    f'Payment Confirmation for Expense #{expense.expense_id}',
                    plain_message,
                    settings.DEFAULT_FROM_EMAIL,
                    [expense.employee.email],  # Send to the employee's email
                )
                
                # Attach HTML content
                email.content_subtype = "html"
                email.body = html_message
                
                # Attach payment proof file if exists
                if finance_process.payment_file:
                    email.attach_file(finance_process.payment_file.path)
                
                # Send email
                email.send()
                
                # Update email sent status
                finance_process.payment_email_sent = True
                finance_process.email_sent_date = timezone.now()
                finance_process.save()
                
                messages.success(request, 'Payment details saved and confirmation email sent successfully!')
            else:
                messages.success(request, 'Payment details saved successfully!')
            
            return redirect('expense_detail', expense_id=expense.expense_id)
            
        except Exception as e:
            messages.error(request, f'Error processing payment: {str(e)}')
    
    context = {
        'expense': expense,
        'finance_process': finance_process,
        'payment_modes': ExpenseFinanceProcess.PAYMENT_MODES,
        # Add this to show warning if duplicates exist
        'has_duplicates': expenses.count() > 1  
    }
    
    return render(request, 'expense/process_finance_process.html', context)

def view_expense_finance_process(request, expense_id):
    """
    View-only version of the finance process details
    """
    # Get all expenses with this ID (should normally be just one)
    expenses = Expense.objects.filter(expense_id=expense_id)
    
    if not expenses.exists():
        return render(request, '404.html', status=404)
    
    # For safety, we'll use the first expense
    expense = expenses.first()
    
    # Get finance process for this expense (don't create if doesn't exist)
    try:
        finance_process = ExpenseFinanceProcess.objects.get(expense=expense)
    except ExpenseFinanceProcess.DoesNotExist:
        # If no finance process exists, create an empty one for display purposes
        finance_process = ExpenseFinanceProcess(expense=expense)
    
    context = {
        'expense': expense,
        'finance_process': finance_process,
        'payment_modes': ExpenseFinanceProcess.PAYMENT_MODES,
        'has_duplicates': expenses.count() > 1,
        'view_only': True  # Add this flag to indicate view-only mode
    }
    
    return render(request, 'expense/view_finance_process.html', context)


def verify_and_process_list(request):
    # Get filter parameters from request
    status = request.GET.get('status', '')
    employee_id = request.GET.get('employee_id', '')  # Changed to employee_id
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Get DataTables parameters
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    # Start with base queryset - exclude Payment Successful by default
    queryset = Expense.objects.select_related('employee').exclude(status='Payment Successful')

    # Apply filters
    if status:
        queryset = queryset.filter(status=status)
    
    if employee_id:
        queryset = queryset.filter(employee__employee_id=employee_id)
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            queryset = queryset.filter(from_date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            queryset = queryset.filter(to_date__lte=end_date)
        except ValueError:
            pass

    # Apply global search if needed
    if search_value:
        queryset = queryset.filter(
            Q(expense_id__icontains=search_value) |
            Q(employee__name__icontains=search_value) |
            Q(employee__employee_code__icontains=search_value) |
            Q(status__icontains=search_value)
        )

    # Group and annotate
    grouped_expenses = queryset.values(
        'expense_id',
        'employee__name',
        'employee__employee_code',
        'total_calculation',
        'total_approved_amount',
        'status'
    ).annotate(
        start_date=Min('from_date'),
        end_date=Max('to_date')
    ).order_by('-start_date')

    # Get total count before pagination
    total_records = grouped_expenses.count()

    # Apply pagination
    paginator = Paginator(grouped_expenses, length)
    page = (start // length) + 1
    try:
        paginated_expenses = paginator.get_page(page)
    except:
        paginated_expenses = paginator.get_page(1)

    # Convert to JSON format
    data = []
    for expense in paginated_expenses:
        data.append({
            'expense_id': str(expense['expense_id']),
            'employee_info': f"{expense['employee__employee_code']} - {expense['employee__name']}",
            'start_date': expense['start_date'].strftime('%Y-%m-%d'),
            'end_date': expense['end_date'].strftime('%Y-%m-%d'),
            'total_calculation': str(expense['total_calculation']),
            'total_approved_amount': str(expense.get('total_approved_amount', '0.00')),
            'status': expense['status']
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': total_records,
        'data': data
    })


def generate_expensefinanceprocess_csv(request):
    """Generate CSV report for Finance Process expenses"""
    # Get filter parameters
    employee_id = request.GET.get('employee_id', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Prefetch the finance process data
    finance_process_qs = ExpenseFinanceProcess.objects.select_related('processed_by')
    
    # Start with base queryset
    expenses = Expense.objects.select_related('employee') \
        .prefetch_related(
            Prefetch('finance_processes', queryset=finance_process_qs)
        ) \
        .filter(status='Payment Successful') \
        .exclude(finance_processes__isnull=True)

    # Apply filters
    if employee_id:
        expenses = expenses.filter(employee__employee_id=employee_id)
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            expenses = expenses.filter(from_date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            expenses = expenses.filter(to_date__lte=end_date)
        except ValueError:
            pass

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="finance_process_report.csv"'

    writer = csv.writer(response)
    # Write headers
    writer.writerow([
        'Expense ID', 'Employee Code', 'Employee Name', 
        'Payment Date', 'Payment Mode', 'Amount Claimed (₹)', 
        'Amount Approved (₹)', 'Payment Status', 'Finance Comments',
        'Processed By', 'Period Start', 'Period End'
    ])

    # Write data rows
    for expense in expenses:
        # Get the first finance process (there should be only one per expense)
        finance_process = expense.finance_processes.first()
        
        writer.writerow([
            expense.expense_id,
            expense.employee.employee_code,
            expense.employee.name,
            finance_process.payment_date.strftime('%Y-%m-%d') if finance_process.payment_date else '',
            finance_process.payment_mode,
            expense.total_calculation,
            expense.total_approved_amount,
            expense.status,
            finance_process.finance_comment or '',
            finance_process.processed_by.name if finance_process.processed_by else '',
            expense.from_date.strftime('%Y-%m-%d'),
            expense.to_date.strftime('%Y-%m-%d')
        ])

    return response

def generate_expensefinanceprocess_excel(request):
    """Generate Excel report for Finance Process expenses"""
    # Get filter parameters
    employee_id = request.GET.get('employee_id', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Prefetch the finance process data
    finance_process_qs = ExpenseFinanceProcess.objects.select_related('processed_by')
    
    # Start with base queryset
    expenses = Expense.objects.select_related('employee') \
        .prefetch_related(
            Prefetch('finance_processes', queryset=finance_process_qs)
        ) \
        .filter(status='Payment Successful') \
        .exclude(finance_processes__isnull=True)

    # Apply filters
    if employee_id:
        expenses = expenses.filter(employee__employee_id=employee_id)
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            expenses = expenses.filter(from_date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            expenses = expenses.filter(to_date__lte=end_date)
        except ValueError:
            pass

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Finance Process Report"

    # Define headers
    headers = [
        'Expense ID', 'Employee Code', 'Employee Name', 
        'Payment Date', 'Payment Mode', 'Amount Claimed (₹)', 
        'Amount Approved (₹)', 'Payment Status', 'Finance Comments',
        'Processed By', 'Period Start', 'Period End'
    ]

    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Write data rows
    for row_num, expense in enumerate(expenses, 2):
        finance_process = expense.finance_processes.first()
        
        ws.cell(row=row_num, column=1, value=str(expense.expense_id))
        ws.cell(row=row_num, column=2, value=expense.employee.employee_code)
        ws.cell(row=row_num, column=3, value=expense.employee.name)
        ws.cell(row=row_num, column=4, value=finance_process.payment_date.strftime('%Y-%m-%d') if finance_process.payment_date else '')
        ws.cell(row=row_num, column=5, value=finance_process.payment_mode)
        ws.cell(row=row_num, column=6, value=float(expense.total_calculation))
        ws.cell(row=row_num, column=7, value=float(expense.total_approved_amount))
        ws.cell(row=row_num, column=8, value=expense.status)
        ws.cell(row=row_num, column=9, value=finance_process.finance_comment or '')
        ws.cell(row=row_num, column=10, value=finance_process.processed_by.name if finance_process.processed_by else '')
        ws.cell(row=row_num, column=11, value=expense.from_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=12, value=expense.to_date.strftime('%Y-%m-%d'))

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Format currency columns
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = '₹#,##0.00'

    # Create response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="finance_process_report.xlsx"'

    wb.save(response)
    return response


def generate_expensefinanceprocess_pdf(request):
    # Get filter parameters
    employee_id = request.GET.get('employee_id', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Prefetch the finance process data
    finance_process_qs = ExpenseFinanceProcess.objects.select_related('processed_by')
    
    # Start with base queryset
    expenses = Expense.objects.select_related('employee') \
        .prefetch_related(
            Prefetch('finance_processes', queryset=finance_process_qs)
        ) \
        .filter(status='Payment Successful') \
        .exclude(finance_processes__isnull=True)

    # Apply filters
    if employee_id:
        expenses = expenses.filter(employee__employee_id=employee_id)
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            expenses = expenses.filter(from_date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            expenses = expenses.filter(to_date__lte=end_date)
        except ValueError:
            pass

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="finance_process_report.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles for wrapping text
    custom_style = ParagraphStyle(
        'custom_style',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        spaceAfter=6
    )

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Report Title
    elements.append(Spacer(1, 80))
    elements.append(Paragraph('<para align="center"><b><font size=16>FINANCE PROCESS REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains details of processed payments including payment modes and dates.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Table Headers
    data = [
        [
            "Expense ID", 
            "Employee", 
            "Period", 
            "Amount Claimed", 
            "Amount Approved", 
            "Payment Mode",
            "Payment Date",
            "Status"
        ]
    ]
    
    # Fetch Expenses and Append to Table Data
    if expenses.exists():
        for expense in expenses:
            finance_process = expense.finance_processes.first()
            employee_info = f"{expense.employee.employee_code} - {expense.employee.name}"
            period = f"{expense.from_date.strftime('%d-%b-%Y')} to {expense.to_date.strftime('%d-%b-%Y')}"
            
            # Use Paragraph for text wrapping
            data.append([
                Paragraph(str(expense.expense_id), custom_style),
                Paragraph(employee_info, custom_style),
                Paragraph(period, custom_style),
                Paragraph(f"Rs.{float(expense.total_calculation):,.2f}", custom_style),
                Paragraph(f"Rs.{float(expense.total_approved_amount):,.2f}", custom_style),
                Paragraph(finance_process.payment_mode if finance_process.payment_mode else "-", custom_style),
                Paragraph(finance_process.payment_date.strftime('%d-%b-%Y') if finance_process.payment_date else "-", custom_style),
                Paragraph(expense.status, custom_style)
            ])
    else:
        data.append(["No Data", "-", "-", "-", "-", "-", "-", "-"])  

    # Table Column Widths (adjust as needed)
    col_widths = [
        1.0 * inch,  # Expense ID
        1.0 * inch,  # Employee
        1.2 * inch,  # Period
        1.0 * inch,  # Amount Claimed
        1.0 * inch,  # Amount Approved
        1.0 * inch,  # Payment Mode
        1.0 * inch,  # Payment Date
        0.8 * inch   # Status
    ]
    
    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),  # Blue header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#D9E1F2')),  # Light blue rows
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#8EA9DB')),  # Light grid
        ('WORDWRAP', (0, 0), (-1, -1), True),  # Enable word wrap
    ])
    table.setStyle(style)

    # Add table to elements
    elements.append(table)
    elements.append(Spacer(1, 15))

    # Footer Note
    elements.append(Paragraph("<strong>Note:</strong> This is a system-generated finance process report.", custom_style))
    
    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response



def generate_verifyprocess_csv(request):
    """Generate CSV report for verified expenses (excluding Payment Successful)"""
    # Get filter parameters
    status = request.GET.get('status', '')
    employee_id = request.GET.get('employee_id', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Prefetch verification data
    verification_qs = VerifyProcessExpense.objects.select_related('employee')
    
    # Start with base queryset - exclude Payment Successful
    expenses = Expense.objects.select_related('employee') \
        .prefetch_related(
            Prefetch('verifications', queryset=verification_qs)
        ) \
        .exclude(status='Payment Successful')

    # Apply filters
    if status:
        expenses = expenses.filter(status=status)
    elif not status:
        # If no status filter is selected, only show Pending/Approved/Rejected
        expenses = expenses.filter(status__in=['Pending', 'Approved', 'Rejected'])
    
    if employee_id:
        expenses = expenses.filter(employee__employee_id=employee_id)
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            expenses = expenses.filter(from_date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            expenses = expenses.filter(to_date__lte=end_date)
        except ValueError:
            pass

    # Group expenses by expense_id and get min/max dates
    grouped_expenses = expenses.values('expense_id').annotate(
        min_from_date=Min('from_date'),
        max_to_date=Max('to_date')
    )

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="verified_expenses_report.csv"'

    writer = csv.writer(response)
    # Write headers
    writer.writerow([
        'Expense ID', 'Employee Code', 'Employee Name', 
        'Expense Status', 'Amount Claimed (₹)', 'Amount Approved (₹)', 
        'Approver Comments', 'Verification Remarks', 'Processed By',
        'Period Start', 'Period End', 'Created Date'
    ])

    # Get distinct expense IDs with their date ranges
    expense_ids = [group['expense_id'] for group in grouped_expenses]
    
    # Get the first expense record for each expense_id (for other fields)
    first_expenses = {}
    for expense in expenses.filter(expense_id__in=expense_ids):
        if expense.expense_id not in first_expenses:
            first_expenses[expense.expense_id] = expense

    # Write data rows
    for group in grouped_expenses:
        expense_id = group['expense_id']
        expense = first_expenses.get(expense_id)
        
        if expense:
            verification = expense.verifications.first()
            
            writer.writerow([
                expense_id,
                expense.employee.employee_code,
                expense.employee.name,
                expense.status,
                expense.total_calculation,
                expense.total_approved_amount,
                verification.approver_comments if verification else '',
                verification.remarks if verification else '',
                verification.employee.name if verification and verification.employee else '',
                group['min_from_date'].strftime('%Y-%m-%d'),  # Earliest from_date
                group['max_to_date'].strftime('%Y-%m-%d'),   # Latest to_date
                expense.created_at.strftime('%Y-%m-%d')
            ])

    return response



def generate_verifyprocess_excel(request):
    """Generate Excel report for verified expenses (excluding Payment Successful)"""
    # Get filter parameters
    status = request.GET.get('status', '')
    employee_id = request.GET.get('employee_id', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Prefetch verification data
    verification_qs = VerifyProcessExpense.objects.select_related('employee')
    
    # Start with base queryset - exclude Payment Successful
    expenses = Expense.objects.select_related('employee') \
        .prefetch_related(
            Prefetch('verifications', queryset=verification_qs)
        ) \
        .exclude(status='Payment Successful')

    # Apply filters
    if status:
        expenses = expenses.filter(status=status)
    elif not status:
        # If no status filter is selected, only show Pending/Approved/Rejected
        expenses = expenses.filter(status__in=['Pending', 'Approved', 'Rejected'])
    
    if employee_id:
        expenses = expenses.filter(employee__employee_id=employee_id)
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            expenses = expenses.filter(from_date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            expenses = expenses.filter(to_date__lte=end_date)
        except ValueError:
            pass

    # Group expenses by expense_id and get min/max dates
    grouped_expenses = expenses.values('expense_id').annotate(
        min_from_date=Min('from_date'),
        max_to_date=Max('to_date')
    )

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verified Expenses Report"

    # Define headers
    headers = [
        'Expense ID', 'Employee Code', 'Employee Name', 
        'Expense Status', 'Amount Claimed (₹)', 'Amount Approved (₹)', 
        'Approver Comments', 'Verification Remarks', 'Processed By',
        'Period Start', 'Period End', 'Created Date'
    ]

    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Get distinct expense IDs with their date ranges
    expense_ids = [group['expense_id'] for group in grouped_expenses]
    
    # Get the first expense record for each expense_id (for other fields)
    first_expenses = {}
    for expense in expenses.filter(expense_id__in=expense_ids):
        if expense.expense_id not in first_expenses:
            first_expenses[expense.expense_id] = expense

    # Write data rows
    row_num = 2
    for group in grouped_expenses:
        expense_id = group['expense_id']
        expense = first_expenses.get(expense_id)
        
        if expense:
            verification = expense.verifications.first()
            
            ws.cell(row=row_num, column=1, value=str(expense_id))
            ws.cell(row=row_num, column=2, value=expense.employee.employee_code)
            ws.cell(row=row_num, column=3, value=expense.employee.name)
            ws.cell(row=row_num, column=4, value=expense.status)
            ws.cell(row=row_num, column=5, value=float(expense.total_calculation))
            ws.cell(row=row_num, column=6, value=float(expense.total_approved_amount))
            ws.cell(row=row_num, column=7, value=verification.approver_comments if verification else '')
            ws.cell(row=row_num, column=8, value=verification.remarks if verification else '')
            ws.cell(row=row_num, column=9, value=verification.employee.name if verification and verification.employee else '')
            ws.cell(row=row_num, column=10, value=group['min_from_date'].strftime('%Y-%m-%d'))  # Earliest from_date
            ws.cell(row=row_num, column=11, value=group['max_to_date'].strftime('%Y-%m-%d'))     # Latest to_date
            ws.cell(row=row_num, column=12, value=expense.created_at.strftime('%Y-%m-%d'))
            
            row_num += 1

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Format currency columns
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = '₹#,##0.00'

    # Create response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="verified_expenses_report.xlsx"'

    wb.save(response)
    return response



def generate_verifyprocess_pdf(request):
    """Generate PDF report for verified expenses (excluding Payment Successful)"""
    # Get filter parameters
    status = request.GET.get('status', '')
    employee_id = request.GET.get('employee_id', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Prefetch verification data
    verification_qs = VerifyProcessExpense.objects.select_related('employee')
    
    # Start with base queryset - exclude Payment Successful
    expenses = Expense.objects.select_related('employee') \
        .prefetch_related(
            Prefetch('verifications', queryset=verification_qs)
        ) \
        .exclude(status='Payment Successful')

    # Apply filters
    if status:
        expenses = expenses.filter(status=status)
    elif not status:
        # If no status filter is selected, only show Pending/Approved/Rejected
        expenses = expenses.filter(status__in=['Pending', 'Approved', 'Rejected'])
    
    if employee_id:
        expenses = expenses.filter(employee__employee_id=employee_id)
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            expenses = expenses.filter(from_date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            expenses = expenses.filter(to_date__lte=end_date)
        except ValueError:
            pass

    # Group expenses by expense_id and get min/max dates
    grouped_expenses = expenses.values('expense_id').annotate(
        min_from_date=Min('from_date'),
        max_to_date=Max('to_date')
    )

    # Get distinct expense IDs with their date ranges
    expense_ids = [group['expense_id'] for group in grouped_expenses]
    
    # Get the first expense record for each expense_id (for other fields)
    first_expenses = {}
    for expense in expenses.filter(expense_id__in=expense_ids):
        if expense.expense_id not in first_expenses:
            first_expenses[expense.expense_id] = expense

    # Create response object for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="verified_expenses_report.pdf"'

    # Define PDF Document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles for wrapping text
    custom_style = ParagraphStyle(
        'custom_style',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        spaceAfter=6
    )

    # Header Image Path
    header_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf/report_header.png")

    # Get Current Date
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Header function
    def on_page(canvas, doc):
        width, height = A4
        if os.path.exists(header_path):
            header_img = Image(header_path, width=width, height=1.5 * inch)
            header_img.drawOn(canvas, 0, height - 1.5 * inch)

        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(width - 0.8 * inch, height - 1.7 * inch, f"Generated On: {report_date}")

    # Report Title
    elements.append(Spacer(1, 80))
    elements.append(Paragraph('<para align="center"><b><font size=16>VERIFIED EXPENSES REPORT</font></b></para>', styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Report Description
    elements.append(Paragraph('<font size=11>This report contains details of verified expenses including approval status and comments.</font>', custom_style))
    elements.append(Spacer(1, 10))

    # Add filter information if any filters are applied
    filter_info = []
    if status:
        filter_info.append(f"Status: {status}")
    if employee_id:
        filter_info.append(f"Employee ID: {employee_id}")
    if start_date:
        filter_info.append(f"From: {start_date}")
    if end_date:
        filter_info.append(f"To: {end_date}")
    
    if filter_info:
        elements.append(Paragraph(f"<b>Filters Applied:</b> {', '.join(filter_info)}", custom_style))
        elements.append(Spacer(1, 10))

    # Table Headers
    data = [
        [
            "Expense ID", 
            "Employee", 
            "Period", 
            "Claimed", 
            "Approved", 
            "Status",
            "Approver cmt",
            "Remarks"
        ]
    ]
    
    # Fetch Expenses and Append to Table Data
    if grouped_expenses:
        for group in grouped_expenses:
            expense_id = group['expense_id']
            expense = first_expenses.get(expense_id)
            
            if expense:
                verification = expense.verifications.first()
                employee_info = f"{expense.employee.employee_code} - {expense.employee.name}"
                period = f"{group['min_from_date'].strftime('%d-%b-%Y')} to {group['max_to_date'].strftime('%d-%b-%Y')}"
                
                # Use Paragraph for text wrapping
                data.append([
                    Paragraph(str(expense_id), custom_style),
                    Paragraph(employee_info, custom_style),
                    Paragraph(period, custom_style),
                    Paragraph(f"{float(expense.total_calculation):,.2f}", custom_style),
                    Paragraph(f"{float(expense.total_approved_amount):,.2f}", custom_style),
                    Paragraph(expense.status, custom_style),
                    Paragraph(verification.approver_comments if verification and verification.approver_comments else "-", custom_style),
                    Paragraph(verification.remarks if verification and verification.remarks else "-", custom_style)
                ])
    else:
        data.append(["No Data", "-", "-", "-", "-", "-", "-", "-"])  

    # Table Column Widths (adjust as needed)
    col_widths = [
        1.2 * inch,  # Expense ID
        1.0 * inch,  # Employee
        1.2 * inch,  # Period
        0.8 * inch,  # Claimed
        0.8 * inch,  # Approved
        0.8 * inch,  # Status
        1.0 * inch,  # Approver Comments
        1.0 * inch   # Remarks
    ]
    
    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Table Styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),  # Blue header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#D9E1F2')),  # Light blue rows
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#8EA9DB')),  # Light grid
        ('WORDWRAP', (0, 0), (-1, -1), True),  # Enable word wrap
    ])
    table.setStyle(style)

    # Add table to elements
    elements.append(table)
    elements.append(Spacer(1, 15))

    # Summary Information
    total_claimed = sum(expense.total_calculation for expense in first_expenses.values())
    total_approved = sum(float(expense.total_approved_amount) for expense in first_expenses.values())
    
    summary_data = [
        ["", "Total Claimed", "Total Approved"],
        ["Amount (₹)", f"{total_claimed:,.2f}", f"{total_approved:,.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
    summary_style = TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    summary_table.setStyle(summary_style)
    elements.append(summary_table)
    
    # Footer Note
    elements.append(Spacer(1, 15))
    elements.append(Paragraph("<strong>Note:</strong> This report includes only Pending, Approved, or Rejected expenses (Payment Successful records excluded).", custom_style))
    
    # Build PDF
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    return response



def get_user_statistics(request):
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    inactive_users = total_users - active_users

    data = {
        'total_users': total_users,
        'active_users': active_users,
        'inactive_users': inactive_users
    }
    return JsonResponse(data)


def get_stock_statistics(request):
    # Get total count of all stock entries
    total_stock = StockEntry.objects.count()
    
    # Get count of items grouped by product_status
    status_counts = StockEntry.objects.values('product_status')\
                                   .annotate(count=Count('product_status'))\
                                   .order_by('product_status')
    
    # Convert to a dictionary for easier access
    status_dict = {item['product_status']: item['count'] for item in status_counts}
    
    data = {
        'total_stock': total_stock,
        'status_counts': status_dict,
        'instock_count': status_dict.get('INSTOCK', 0),
        'assign_count': status_dict.get('ASSIGN', 0),
        'fixed_asset_count': status_dict.get('FIXED_ASSET', 0),
        'return_count': status_dict.get('RETURN', 0),
        'sell_count': status_dict.get('SELL', 0),  # Make sure this matches your model's choice value exactly
        'demo_count': status_dict.get('DEMO', 0),
        'inspection_count': status_dict.get('INSPECTION', 0),
    }
    print(data)  # Add this to debug what's being returned
    return JsonResponse(data)



def get_system_counts(request):
    # Get counts for all major models
    data = {
        # Location counts
        'total_locations': Location.objects.count(),
        'active_locations': Location.objects.filter(status='active').count(),
        'inactive_locations': Location.objects.filter(status='inactive').count(),
        
        # Employee counts
        'total_employees': Employee.objects.count(),
        
        # Company counts
        'total_companies': Company.objects.count(),
        
        # Category counts
        'total_categories': Category.objects.count(),
        
        # SubCategory counts
        'total_subcategories': SubCategory.objects.count(),
        'subcategories_by_category': list(
            Category.objects.annotate(subcategory_count=Count('subcategories'))
                          .values('category_name', 'subcategory_count')
        ),
        
        # Product counts
        'total_products': Product.objects.count(),
        'products_by_category': list(
            Category.objects.annotate(product_count=Count('product'))
                          .values('category_name', 'product_count')
        ),
        
        # Exhibition counts
        'total_exhibitions': Exhibition.objects.count(),
        'active_exhibitions': Exhibition.objects.filter(
            end_date__gte=timezone.now()
        ).count(),
        
        # Customer counts
        'total_customers': Customer.objects.count(),
        'customers_by_company': list(
            Company.objects.annotate(customer_count=Count('customers'))
                          .values('company_name', 'customer_count')
        ),
        
        # Supplier counts
        'total_suppliers': Supplier.objects.count(),
        'suppliers_by_company': list(
            Company.objects.annotate(supplier_count=Count('suppliers'))
                          .values('company_name', 'supplier_count')
        ),
        
        # Stock counts (from your existing implementation)
        'total_stock': StockEntry.objects.count(),
        'stock_by_status': dict(
            StockEntry.objects.values('product_status')
                            .annotate(count=Count('product_status'))
                            .values_list('product_status', 'count')
        )
    }
    
    return JsonResponse(data)

def get_stock_by_location(request):
    # Get total stock count for percentage calculations
    total_stock = StockEntry.objects.aggregate(total=Sum('quantity'))['total'] or 1
    
    # Get stock quantities grouped by location
    locations = Location.objects.annotate(
        stock_count=Sum('stockentry__quantity')
    ).values('name', 'shortcode', 'stock_count').order_by('-stock_count')
    
    # Calculate percentages
    location_data = []
    for loc in locations:
        percentage = round((loc['stock_count'] or 0) / total_stock * 100)
        location_data.append({
            'name': loc['name'],
            'shortcode': loc['shortcode'],
            'count': loc['stock_count'] or 0,
            'percentage': percentage
        })
    
    return JsonResponse({
        'locations': location_data,
        'total_stock': total_stock
    })



def stock_dashboard_data(request):
    # Stock Operations Data
    total_stock = StockEntry.objects.count()
    assign_stock = StockEntry.objects.filter(product_status='ASSIGN').count()
    instock = StockEntry.objects.filter(product_status='INSTOCK').count()
    fixed_asset_percentage = round((StockEntry.objects.filter(product_status='FIXED_ASSET').count() / total_stock * 100) if total_stock else 0, 2)
    
    # Stock by Locations Data
    locations = Location.objects.annotate(
        count=Count('stockentry'),
        assign=Sum(Case(When(stockentry__product_status='ASSIGN', then=1), output_field=IntegerField())),
        instock=Sum(Case(When(stockentry__product_status='INSTOCK', then=1), output_field=IntegerField())),
        fixed_asset=Sum(Case(When(stockentry__product_status='FIXED_ASSET', then=1), output_field=IntegerField())
    )).order_by('-count')
    
    # Prepare data for charts
    stock_operations_data = {
        'total_stock': total_stock,
        'assign_stock': assign_stock,
        'instock': instock,
        'fixed_asset_percentage': fixed_asset_percentage,
    }
    
    locations_data = [
        {
            'id': loc.id,
            'name': loc.name,
            'shortcode': loc.shortcode,
            'count': loc.count,
            'assign': loc.assign or 0,
            'instock': loc.instock or 0,
            'fixed_asset': loc.fixed_asset or 0,
        }
        for loc in locations
    ]
    
    return JsonResponse({
        'stock_operations': stock_operations_data,
        'locations': locations_data,
    })